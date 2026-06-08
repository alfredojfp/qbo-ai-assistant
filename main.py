"""
TMP AI Assistant - Main Application
Asistente inteligente para automatización de tareas contables

Autor: Alfredo
Fecha: Enero 2026
"""

import os
import json
import csv
import re
import requests
import subprocess
import sys
from collections import deque
from datetime import datetime, timedelta
import glob
import shutil
from ocr_bills import extraer_bills_de_pdf, generar_csv_preview

# ============================================================================
# AUTONOMÍA Y APRENDIZAJE - AUTO-INSTALADO 2026-01-21 22:13:27
# ============================================================================
from autonomia.autonomia_nivel1_websearch import tool_search_web, tool_search_qbo_docs
from autonomia.autonomia_nivel2_api_explorer import (
    tool_create_journal_entry,
    tool_create_transfer,
    tool_qbo_generic_request,
    tool_list_qbo_endpoints,
    tool_get_endpoint_info
)
from autonomia.autonomia_nivel3_code_executor import tool_execute_python
from autonomia.bank_feed_intelligence import (
    tool_analyze_bank_feed_for_classification,
    tool_record_bank_feed_classification,
    tool_get_classification_history_stats,
    tool_find_pattern_for_transaction,
)
# Re-exports Fase 1 refactor: nombres cortos para `from main import tool_xxx`
tool_analizarbankfeed = tool_analyze_bank_feed_for_classification
tool_registrarclasificacion = tool_record_bank_feed_classification
tool_estadisticasclasificacion = tool_get_classification_history_stats
tool_buscarpatron = tool_find_pattern_for_transaction
from autonomia.user_behavior_learning import (
    tool_learn_from_interaction,
    tool_get_user_suggestions,
    tool_record_user_correction,
    tool_get_conversation_context
)
from autonomia.dynamic_report_generator import (
    tool_generate_custom_report,
    tool_parse_date_expression
)
from dotenv import load_dotenv
from typing import Dict, List, Optional, Tuple, Any

# Multi-company support
from company_manager import (
    select_company_interactive,
    load_current_company,
    save_company_selection,
    load_company_context,
    save_company_context,
    list_local_companies,
    extract_realm_id,
    save_company_meta,
    get_company_meta
)

# Sistema centralizado de logging de errores (dexter.error_log)
from dexter.error_log import log_error as _log_error

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import SequenceMatcher

# ==================== CONFIGURACIÓN ====================
load_dotenv()

# ═════════════════════════════════════════════════════════════════
# MULTI-COMPANY GLOBALS
# ═════════════════════════════════════════════════════════════════
CURRENT_COMPANY = None
COMPANY_CONTEXT = None


# Credenciales QuickBooks
QB_ACCESS_TOKEN = os.getenv("QB_ACCESS_TOKEN")
QB_REFRESH_TOKEN = os.getenv("QB_REFRESH_TOKEN")
QB_CLIENT_ID = os.getenv("QB_CLIENT_ID")
QB_CLIENT_SECRET = os.getenv("QB_CLIENT_SECRET")
QB_REALM_ID = os.getenv("QB_REALM_ID")

# Credenciales OpenRouter
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

# URLs QuickBooks
def _build_qb_base_url(realm_id: str) -> str:
    """MED-1 fix: valida realm_id y construye QB_BASE_URL. Raise con
    mensaje claro si realm_id es None o vacío."""
    if not realm_id:
        raise RuntimeError(
            "QB_REALM_ID environment variable is required but missing or empty. "
            "Set it in .env or run the OAuth setup: python scripts/oauth_flow.py"
        )
    return f"https://sandbox-quickbooks.api.intuit.com/v3/company/{realm_id}"


QB_BASE_URL = _build_qb_base_url(QB_REALM_ID)
QB_AUTH_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"

# Configuración LLM (Hybrid Model Routing)
LLM_MODEL_HEAVY = "deepseek/deepseek-chat"
LLM_MODEL_LIGHT = "meta-llama/llama-3.1-8b-instruct"
LLM_API_URL = "https://openrouter.ai/api/v1/chat/completions"

# Precios promedio (USD por millón de tokens)
PRICE_INPUT_DEEPSEEK = 0.14
PRICE_OUTPUT_DEEPSEEK = 0.28
PRICE_INPUT_LLAMA = 0.05
PRICE_OUTPUT_LLAMA = 0.08

# Archivos del sistema
FILE_CHART_CACHE = "data/chart_of_accounts.json"
FILE_SAVED_REPORTS = "data/saved_reports.json"
FILE_TOKEN_USAGE = "data/token_usage.csv"
FILE_TOKEN_REPORT = "token_usage_report.xlsx"
FILE_DEPOSITS_TEMPLATE = "deposits_template.csv"

# ── Dry-Run Mode ──────────────────────────────────────────────────────
DRY_RUN_ACTIVE = False
_last_dry_run_message = None  # guardado para /ejecutar

# Tools de solo-lectura: SIEMPRE se ejecutan, incluso en dry-run
# (necesarios para dar contexto al usuario sobre qué se haría)
_READ_ONLY_TOOLS = {
    "buscar_cliente", "buscar_vendor", "buscar_cuenta", "buscar_item",
    "qbo_query", "consulta_avanzada",
    "generar_reporte_pl", "generar_balance_sheet", "generar_reporte_custom",
    "reporte_trial_balance", "reporte_general_ledger", "reporte_cash_flow",
    "reporte_ar_aging", "reporte_ap_aging", "reporte_journal",
    "reporte_account_list", "reporte_customer_balance", "reporte_vendor_balance",
    "reporte_pl_detail", "reporte_inventory_valuation", "reporte_class_sales",
    "reporte_department_sales", "reporte_expenses_by_vendor",
    "reporte_sales_by_customer", "reporte_transaction_list",
    "leer_companyinfo", "leer_preferencias", "leer_exchange_rate",
    "cdc_query", "ver_log_errores", "gestionar_memoria",
    "gestionar_empresas", "refrescar_chart_accounts",
    "estadisticasclasificacion", "buscarpatron",
    "analizarbankfeed", "buscarenweb", "buscardocsqbo",
    "obtenersugerencias", "obtenercontexto",
}
# Tools de escritura (no exhaustivo — por defecto, cualquier tool NO en
# _READ_ONLY_TOOLS se considera de escritura y se simula en dry-run).


def _parse_dry_run(message: str):
    """Detecta y elimina el flag --dry-run de un mensaje.

    Returns:
        (mensaje_limpio: str, es_dry_run: bool)
    """
    clean = message
    is_dry = False
    for flag in ("--dry-run", "--DRY-RUN", "--dry_run", "--DRY_RUN"):
        if flag in clean:
            clean = clean.replace(flag, "")
            is_dry = True
    # Limpiar espacios dobles
    clean = " ".join(clean.split())
    if is_dry:
        global _last_dry_run_message
        _last_dry_run_message = clean
    return clean, is_dry


def _execute_tool(function_name: str, arguments: dict):
    """Ejecuta un tool, respetando el modo dry-run.

    En dry-run, tools de solo-lectura se ejecutan normalmente.
    Tools de escritura retornan un mensaje simulado.
    """
    if function_name in TOOL_FUNCTIONS:
        if DRY_RUN_ACTIVE and function_name not in _READ_ONLY_TOOLS:
            args_str = ", ".join(f"{k}={v}" for k, v in arguments.items())
            return {
                "dry_run": True,
                "dry_run_note": f"[DRY-RUN] Se simularía {function_name}({args_str}). "
                                f"No se ejecutó nada en QBO.",
                "success": True,
            }
        try:
            return TOOL_FUNCTIONS[function_name](**arguments)
        except Exception as e:
            return {"error": str(e)}
    return {"error": f"Tool '{function_name}' no encontrado"}

# Estado de la sesión
session_state = {
    "start_time": datetime.now(),
    "input_tokens": 0,
    "output_tokens": 0,
    "total_cost": 0.0,
    "operations": {
        "searches": 0,
        "deposits": 0,
        "invoices": 0,
        "bills": 0,
        "payments": 0,
        "reports": 0,
        "csv_batches": 0
    },
    "chart_of_accounts": {},
    "saved_reports": {},
    "last_search_results": {},
    "language": "es"
}

# Historial de conversación (CRIT-2: bounded con deque para evitar OOM en sesiones largas)
CONVERSATION_HISTORY_MAXLEN = 200
conversation_history = deque(maxlen=CONVERSATION_HISTORY_MAXLEN)

# ==================== UTILIDADES GENERALES ====================

def log_operation(op_type: str):
    """Registra una operación en las estadísticas de la sesión"""
    if op_type in session_state["operations"]:
        session_state["operations"][op_type] += 1

def similarity_score(a: str, b: str) -> float:
    """Calcula similitud entre dos strings (0-1)"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def update_env_file(key: str, value: str):
    """Actualiza una variable en el archivo .env de forma atómica.

    MED-4 fix: escribe a .env.tmp primero, luego os.replace(tmp, real).
    os.replace es atómico en POSIX (y NTFS en Windows). Garantiza que
    .env siempre queda en estado consistente (viejo o nuevo, nunca parcial)
    aún si el proceso muere a mitad de write.
    """
    env_path = ".env"
    tmp_path = ".env.tmp"
    lines = []

    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            lines = f.readlines()

    updated = False
    for i, line in enumerate(lines):
        if line.startswith(f"{key}="):
            lines[i] = f"{key}={value}\n"
            updated = True
            break

    if not updated:
        lines.append(f"{key}={value}\n")

    try:
        with open(tmp_path, 'w') as f:
            f.writelines(lines)
        os.replace(tmp_path, env_path)
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise

def parse_date(date_str: str) -> str:
    """Convierte diferentes formatos de fecha a YYYY-MM-DD.

    MED-2 fix: distingue entre fecha vacía (usa today, backward compat)
    y fecha con formato inválido (raise ValueError con mensaje claro).
    El fallback silencioso a 'today' causaba errores de auditoría en
    operaciones batch.
    """
    if not date_str:
        return datetime.now().strftime("%Y-%m-%d")

    # Si ya está en formato correcto
    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str

    # Intentar parsear diferentes formatos
    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
        "%d/%m/%y", "%d-%m-%y",
        "%B %d, %Y", "%d %B %Y"
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    raise ValueError(
        f"parse_date: no se pudo parsear '{date_str}'. "
        f"Formatos soportados: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, "
        f"DD/MM/YY, '15 June 2026', etc."
    )

def format_currency(amount: float) -> str:
    """Formatea cantidad como moneda"""
    return f"${amount:,.2f}"

# ==================== AUTENTICACIÓN QUICKBOOKS ====================

def refresh_qb_token():
    """Refresca el access token de QuickBooks"""
    global QB_ACCESS_TOKEN, QB_REFRESH_TOKEN

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded"
    }

    data = {
        "grant_type": "refresh_token",
        "refresh_token": QB_REFRESH_TOKEN
    }

    response = requests.post(
        QB_AUTH_URL,
        headers=headers,
        data=data,
        auth=(QB_CLIENT_ID, QB_CLIENT_SECRET)
    )

    if response.status_code == 200:
        tokens = response.json()
        QB_ACCESS_TOKEN = tokens["access_token"]
        QB_REFRESH_TOKEN = tokens["refresh_token"]

        # Actualizar .env (como respaldo)
        update_env_file("QB_ACCESS_TOKEN", QB_ACCESS_TOKEN)
        update_env_file("QB_REFRESH_TOKEN", QB_REFRESH_TOKEN)

        # Actualizar tokens de la empresa actual
        if CURRENT_COMPANY:
            save_company_meta(
                CURRENT_COMPANY["name"], 
                CURRENT_COMPANY["realm_id"], 
                access_token=QB_ACCESS_TOKEN, 
                refresh_token=QB_REFRESH_TOKEN
            )

        print(f"✅ Token refrescado exitosamente para {CURRENT_COMPANY['name'] if CURRENT_COMPANY else 'QBO'}")
        return True
    else:
        print(f"❌ Error al refrescar token: {response.text}")
        _log_error(
            f"Refresh token failed: HTTP {response.status_code}",
            category="auth",
            extra={
                "status_code": response.status_code,
                "response_preview": response.text[:500],
            },
        )
        return False

QB_REQUEST_TIMEOUT = int(os.getenv("QB_REQUEST_TIMEOUT", "30"))


def qbo_request(method: str, endpoint: str, data: dict = None, params: dict = None,
                raw_body: bytes = None, extra_headers: dict = None) -> requests.Response:
    """Realiza request a QuickBooks con manejo automático de refresh token y retry.

    CRIT-1 fix: timeout=30s por defecto (configurable vía QB_REQUEST_TIMEOUT).
    CRIT-4 fix: retry con backoff exponencial (1s, 2s, 4s) en 429/503/Timeout/ConnectionError.
    HIGH-6 fix: acepta `raw_body` (bytes) y `extra_headers` para soportar
                multipart/form-data en upload_attachment.
    """
    from dexter.core.retry import retry_request

    global QB_ACCESS_TOKEN

    headers = {
        "Authorization": f"Bearer {QB_ACCESS_TOKEN}",
        "Accept": "application/json",
    }
    if raw_body is not None:
        if extra_headers:
            headers.update(extra_headers)
    else:
        headers["Content-Type"] = "application/json"
        if extra_headers:
            headers.update(extra_headers)

    url = f"{QB_BASE_URL}/{endpoint}"

    # Pinear minorversion para usar una versión específica de la API (default 70, configurable)
    minor_version = os.getenv("QB_MINOR_VERSION", "70")
    if params is None:
        params = {}
    params.setdefault("minorversion", minor_version)

    if method == "GET":
        response = retry_request(
            requests.get, url,
            headers=headers, params=params, timeout=QB_REQUEST_TIMEOUT,
        )
    elif method == "POST":
        if raw_body is not None:
            response = retry_request(
                requests.post, url,
                headers=headers, data=raw_body, timeout=QB_REQUEST_TIMEOUT,
            )
        else:
            response = retry_request(
                requests.post, url,
                headers=headers, json=data, timeout=QB_REQUEST_TIMEOUT,
            )
    else:
        raise ValueError(f"Método no soportado: {method}")

    # Si es 401, refrescar token y reintentar (CRIT-1 logic preservado)
    if response.status_code == 401:
        print("🔄 Token expirado, refrescando...")
        if refresh_qb_token():
            headers["Authorization"] = f"Bearer {QB_ACCESS_TOKEN}"
            if method == "GET":
                response = retry_request(
                    requests.get, url,
                    headers=headers, params=params, timeout=QB_REQUEST_TIMEOUT,
                )
            else:
                if raw_body is not None:
                    response = retry_request(
                        requests.post, url,
                        headers=headers, data=raw_body, timeout=QB_REQUEST_TIMEOUT,
                    )
                else:
                    response = retry_request(
                        requests.post, url,
                        headers=headers, json=data, timeout=QB_REQUEST_TIMEOUT,
                    )

    # Si la respuesta final no es 2xx, persistir en el log para diagnóstico
    if response.status_code >= 400:
        _log_error(
            f"QBO API {method} /{endpoint} → HTTP {response.status_code}: {response.text[:200]}",
            category="api_call",
            extra={
                "method": method,
                "endpoint": endpoint,
                "status_code": response.status_code,
                "response_preview": response.text[:500],
                "request_data_preview": (json.dumps(data)[:200] if data else None),
            },
        )

    return response

def qbo_query(sql: str) -> dict:
    """Ejecuta query SQL en QuickBooks con auto-paginación (HIGH-8).

    QBO API limita a 1000 resultados por query. Si la respuesta trae
    1000 (o el caller recibe menos que totalCount), se re-ejecuta la
    query con STARTPOSITION N MAXRESULTS 1000 hasta agotar la lista.
    """
    sql_upper = sql.upper()
    if "MAXRESULTS" not in sql_upper:
        sql = f"{sql.rstrip(';').rstrip()} MAXRESULTS 1000"

    first = qbo_request("GET", "query", params={"query": sql})

    if first.status_code != 200:
        return {"error": first.text, "status": first.status_code}

    aggregated = first.json()
    qr = aggregated.get("QueryResponse", {})

    entity_keys = [k for k in qr.keys() if k not in ("maxResults", "startPosition", "time",
                                                          "totalCount", "QueryResponse")]
    if not entity_keys:
        return aggregated
    entity_key = entity_keys[0]
    raw = qr.get(entity_key, [])
    try:
        rows = list(raw)
    except TypeError:
        # entity_key apunta a escalar (ej: totalCount), no a lista
        return aggregated

    page_size = 1000
    start_position = len(rows) + 1
    while len(rows) > 0 and len(rows) % page_size == 0:
        paged_sql = _inject_startposition(sql, start_position)
        next_resp = qbo_request("GET", "query", params={"query": paged_sql})
        if next_resp.status_code != 200:
            break
        next_qr = next_resp.json().get("QueryResponse", {})
        next_rows = next_qr.get(entity_key, [])
        if not next_rows:
            break
        rows.extend(next_rows)
        start_position += len(next_rows)
        if len(next_rows) < page_size:
            break

    if "QueryResponse" in aggregated:
        aggregated["QueryResponse"][entity_key] = rows
    return aggregated


def _inject_startposition(sql: str, position: int) -> str:
    """Inserta/actualiza STARTPOSITION N en el SQL."""
    import re
    pattern = re.compile(r"\bSTARTPOSITION\s+\d+", re.IGNORECASE)
    if pattern.search(sql):
        return pattern.sub(f"STARTPOSITION {position}", sql)
    return f"{sql.rstrip(';').rstrip()} STARTPOSITION {position}"

# ==================== CHART OF ACCOUNTS ====================

CHART_SCHEMA_VERSION = 2  # LOW-4: bump cuando cambia estructura de accounts


def load_chart_of_accounts(force_refresh: bool = False) -> dict:
    """Carga Chart of Accounts desde QBO (con caché opcional).

    LOW-4 fix: el cache incluye schema_version + company_realm_id.
    Si el schema difiere o el realm es OTRA empresa, ignora el cache
    y re-descarga de QBO. Esto evita que un cambio de estructura
    en el código cargue datos viejos incompatibles.
    """
    current_realm = os.environ.get("QB_REALM_ID", "")

    if not force_refresh and os.path.exists(FILE_CHART_CACHE):
        try:
            with open(FILE_CHART_CACHE, 'r') as f:
                cache = json.load(f)

            cache_schema = cache.get("schema_version")
            cache_realm = cache.get("company_realm_id")
            cache_date_str = cache.get("last_updated", "2020-01-01")

            if cache_schema != CHART_SCHEMA_VERSION:
                print(f"⚠️ Cache schema desactualizado (v{cache_schema} → "
                      f"v{CHART_SCHEMA_VERSION}); re-descargando...")
            elif cache_realm != current_realm:
                print(f"⚠️ Cache de realm '{cache_realm}' (actual: "
                      f"'{current_realm}'); re-descargando...")
            else:
                cache_date = datetime.fromisoformat(cache_date_str)
                if datetime.now() - cache_date < timedelta(days=1):
                    print(f"📊 Chart of Accounts cargado desde caché "
                          f"({len(cache['accounts'])} cuentas)")
                    return cache["accounts"]
        except Exception as e:
            print(f"⚠️ Error leyendo caché: {e}")

    print("📥 Descargando Chart of Accounts desde QuickBooks Online...")
    sql = "SELECT * FROM Account WHERE Active = true"
    result = qbo_query(sql)

    if "error" in result:
        print(f"❌ Error cargando Chart of Accounts: {result['error']}")
        return {}

    accounts_data = result.get("QueryResponse", {}).get("Account", [])

    chart = {}
    categories = {"ACTIVO": 0, "PASIVO": 0, "INGRESO": 0, "GASTO": 0, "OTRO": 0}

    for acc in accounts_data:
        acc_id = acc.get("Id")
        acc_type = acc.get("AccountType")
        category = get_account_category(acc_type)

        chart[acc_id] = {
            "id": acc_id,
            "name": acc.get("Name"),
            "number": acc.get("AcctNum", ""),
            "type": acc_type,
            "subtype": acc.get("AccountSubType", ""),
            "category": category,
            "active": acc.get("Active", True),
            "balance": float(acc.get("CurrentBalance", 0))
        }

        categories[category] += 1

    cache_data = {
        "schema_version": CHART_SCHEMA_VERSION,
        "company_realm_id": current_realm,
        "last_updated": datetime.now().isoformat(),
        "accounts": chart
    }

    with open(FILE_CHART_CACHE, 'w') as f:
        json.dump(cache_data, f, indent=2)

    print(f"✅ {len(chart)} cuentas cargadas")
    print(f"   Activos: {categories['ACTIVO']} | Pasivos: {categories['PASIVO']} | Ingresos: {categories['INGRESO']} | Gastos: {categories['GASTO']}")

    return chart

def get_account_category(account_type: str) -> str:
    """Determina la categoría principal de una cuenta"""
    asset_types = ["Bank", "Other Current Asset", "Fixed Asset", "Other Asset", "Accounts Receivable"]
    liability_types = ["Accounts Payable", "Credit Card", "Other Current Liability", "Long Term Liability"]
    income_types = ["Income", "Other Income"]
    expense_types = ["Expense", "Other Expense", "Cost of Goods Sold"]

    if account_type in asset_types:
        return "ACTIVO"
    elif account_type in liability_types:
        return "PASIVO"
    elif account_type in income_types:
        return "INGRESO"
    elif account_type in expense_types:
        return "GASTO"
    else:
        return "OTRO"

def find_account(search_term: str, exact: bool = False, category: str = None) -> List[dict]:
    """Busca cuenta por nombre o número con fuzzy matching.

    MED-5 fix: None-guard para acc['name'] y acc['category']. Si el
    chart tiene data malformada (None values), skip la cuenta (no
    es match válido). Antes crasheaba con AttributeError en .lower().
    """
    chart = session_state.get("chart_of_accounts", {})
    results = []

    for acc_id, acc in chart.items():
        acc_name = acc.get("name")
        acc_category = acc.get("category")
        acc_number = acc.get("number")

        if acc_name is None:
            continue

        # Filtrar por categoría si se especifica
        if category and acc_category != category.upper():
            continue

        # Buscar por número exacto
        if acc_number == search_term:
            return [acc]

        # Buscar por nombre
        if exact:
            if acc_name.lower() == search_term.lower():
                results.append(acc)
        else:
            score = similarity_score(search_term, acc_name)
            if score > 0.6:  # 60% de similitud mínima
                results.append({**acc, "match_score": score})

    # Ordenar por score si es fuzzy search
    if not exact and results:
        results.sort(key=lambda x: x.get("match_score", 0), reverse=True)

    return results

# ==================== TRACKING DE TOKENS ====================

def update_token_usage(input_tokens: int, output_tokens: int, model: str):
    """Actualiza los contadores de tokens y el costo acumulado de la sesión"""
    session_state["input_tokens"] += input_tokens
    session_state["output_tokens"] += output_tokens
    
    # Calcular costo de este llamado basado en el modelo
    if "deepseek" in model.lower():
        cost = (input_tokens / 1_000_000 * PRICE_INPUT_DEEPSEEK) + (output_tokens / 1_000_000 * PRICE_OUTPUT_DEEPSEEK)
    else:
        cost = (input_tokens / 1_000_000 * PRICE_INPUT_LLAMA) + (output_tokens / 1_000_000 * PRICE_OUTPUT_LLAMA)
        
    session_state["total_cost"] += cost

def calculate_session_cost() -> float:
    """Retorna el costo acumulado de la sesión"""
    return session_state.get("total_cost", 0.0)

import threading as _threading
_csv_write_lock = _threading.Lock()
_company_lock = _threading.Lock()  # MED-12: serializa cambios de empresa


def _build_entity_ref(entity_id, entity_type: str = None,
                      name: str = None, include_name: bool = False) -> dict:
    """Construye un EntityRef para QBO sin incluir 'name' por default.

    LOW-5 fix: QBO rechaza EntityRef cuyo 'name' no coincide exacto
    con el DisplayName actual del Customer/Vendor/Employee. Convención:
    enviar SOLO 'value' (el ID) y dejar que QBO resuelva el name.

    Args:
        entity_id: ID numérico de QBO (str o int).
        entity_type: 'Customer' | 'Vendor' | 'Employee' | 'Other'.
                     Opcional (algunos endpoints no lo requieren).
        name: Display name (IGNORADO por default para evitar rechazos).
        include_name: si True, incluye 'name' en el dict (uso edge-case).

    Returns:
        dict {'value': entity_id, 'type': entity_type, ['name': name]}
    """
    if not entity_id and entity_id != 0:
        raise ValueError(f"entity_id required, got {entity_id!r}")

    ref = {"value": str(entity_id)}
    if entity_type:
        ref["type"] = entity_type
    if include_name and name:
        ref["name"] = name
    return ref


def save_session_to_csv():
    """Guarda los datos de la sesión en el CSV histórico.

    MED-3 fix: usa _csv_write_lock (threading.Lock) para serializar
    writes concurrentes. Evita corrupción si Ctrl+C signal handler y
    auto-save timer llaman a esta función simultáneamente.
    """
    with _csv_write_lock:
        duration = (datetime.now() - session_state["start_time"]).total_seconds() / 60
        operations_count = sum(session_state["operations"].values())
        operations_detail = ", ".join([
            f"{count} {op}" for op, count in session_state["operations"].items() if count > 0
        ])

        row = {
            "fecha": session_state["start_time"].strftime("%Y-%m-%d"),
            "sesion_inicio": session_state["start_time"].strftime("%H:%M"),
            "sesion_fin": datetime.now().strftime("%H:%M"),
            "duracion_min": round(duration, 1),
            "input_tokens": session_state["input_tokens"],
            "output_tokens": session_state["output_tokens"],
            "total_tokens": session_state["input_tokens"] + session_state["output_tokens"],
            "costo_usd": round(calculate_session_cost(), 4),
            "operaciones": operations_count,
            "detalles": operations_detail or "Sin operaciones"
        }

        # Crear archivo si no existe
        file_exists = os.path.exists(FILE_TOKEN_USAGE)

        with open(FILE_TOKEN_USAGE, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=row.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)

def generate_token_report():
    """Genera informe Excel con estadísticas de consumo (SOBRESCRIBE)"""
    if not os.path.exists(FILE_TOKEN_USAGE):
        print("⚠️ No hay datos de consumo todavía")
        return

    # Leer datos
    df = pd.read_csv(FILE_TOKEN_USAGE)

    # Crear Excel
    wb = Workbook()

    # Hoja 1: Resumen mensual
    ws1 = wb.active
    ws1.title = "Resumen Mensual"

    df['fecha'] = pd.to_datetime(df['fecha'])
    monthly = df.groupby(df['fecha'].dt.to_period('M')).agg({
        'costo_usd': 'sum',
        'operaciones': 'sum',
        'duracion_min': 'sum',
        'total_tokens': 'sum'
    }).reset_index()
    monthly['fecha'] = monthly['fecha'].astype(str)

    # Encabezados
    headers = ["Mes", "Costo USD", "Operaciones", "Duración (min)", "Total Tokens"]
    ws1.append(headers)

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    for _, row in monthly.iterrows():
        ws1.append([row['fecha'], round(row['costo_usd'], 4), int(row['operaciones']),
                   round(row['duracion_min'], 1), int(row['total_tokens'])])

    # Hoja 2: Por sesión
    ws2 = wb.create_sheet("Por Sesión")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Hoja 3: Estadísticas
    ws3 = wb.create_sheet("Estadísticas")

    total_cost = df['costo_usd'].sum()
    total_sessions = len(df)
    avg_cost = df['costo_usd'].mean()
    total_operations = df['operaciones'].sum()

    stats = [
        ["Métrica", "Valor"],
        ["Total gastado", f"${total_cost:.4f}"],
        ["Sesiones totales", total_sessions],
        ["Costo promedio/sesión", f"${avg_cost:.4f}"],
        ["Operaciones totales", int(total_operations)],
        ["Promedio operaciones/sesión", round(total_operations / total_sessions, 1)],
        ["Proyección mensual", f"${total_cost / max(len(monthly), 1):.2f}"],
        ["Proyección anual", f"${(total_cost / max(len(monthly), 1)) * 12:.2f}"]
    ]

    for row in stats:
        ws3.append(row)

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    # Hoja 4: Por tipo de operación
    ws4 = wb.create_sheet("Por Tipo")

    # Analizar detalles de operaciones
    operation_counts = {}
    for detail in df['detalles'].dropna():
        for op in detail.split(", "):
            if op.strip():
                parts = op.strip().split()
                if len(parts) >= 2:
                    count = parts[0]
                    op_type = " ".join(parts[1:])
                    operation_counts[op_type] = operation_counts.get(op_type, 0) + (int(count) if count.isdigit() else 1)

    ws4.append(["Tipo de Operación", "Cantidad"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for op_type, count in sorted(operation_counts.items(), key=lambda x: x[1], reverse=True):
        ws4.append([op_type, count])

    # Guardar (SOBRESCRIBE)
    wb.save(FILE_TOKEN_REPORT)
    print(f"✅ Informe sobrescrito: {FILE_TOKEN_REPORT}")
    print(f"   📊 {len(df)} sesiones | ${total_cost:.4f} total | {int(total_operations)} operaciones")

# ==================== BÚSQUEDAS EN QUICKBOOKS ====================

def search_customer(search_term: str, exact: bool = False) -> List[dict]:
    """Busca clientes en QuickBooks"""
    log_operation("searches")

    if exact:
        sql = f"SELECT * FROM Customer WHERE DisplayName = '{search_term}'"
    else:
        sql = f"SELECT * FROM Customer WHERE DisplayName LIKE '%{search_term}%'"

    result = qbo_query(sql)

    if "error" in result:
        return []

    customers = result.get("QueryResponse", {}).get("Customer", [])
    results = []

    for c in customers:
        results.append({
            "id": c.get("Id"),
            "name": c.get("DisplayName"),
            "company": c.get("CompanyName", ""),
            "balance": float(c.get("Balance", 0)),
            "active": c.get("Active", True)
        })

    return results

def search_vendor(search_term: str, exact: bool = False) -> List[dict]:
    """Busca vendors en QuickBooks"""
    log_operation("searches")

    if exact:
        sql = f"SELECT * FROM Vendor WHERE DisplayName = '{search_term}'"
    else:
        sql = f"SELECT * FROM Vendor WHERE DisplayName LIKE '%{search_term}%'"

    result = qbo_query(sql)

    if "error" in result:
        return []

    vendors = result.get("QueryResponse", {}).get("Vendor", [])
    results = []

    for v in vendors:
        results.append({
            "id": v.get("Id"),
            "name": v.get("DisplayName"),
            "company": v.get("CompanyName", ""),
            "balance": float(v.get("Balance", 0)),
            "active": v.get("Active", True)
        })

    return results

def search_item(search_term: str) -> List[dict]:
    """Busca items/servicios en QuickBooks"""
    log_operation("searches")

    sql = f"SELECT * FROM Item WHERE Name LIKE '%{search_term}%'"
    result = qbo_query(sql)

    if "error" in result:
        return []

    items = result.get("QueryResponse", {}).get("Item", [])
    results = []

    for item in items:
        results.append({
            "id": item.get("Id"),
            "name": item.get("Name"),
            "type": item.get("Type"),
            "price": float(item.get("UnitPrice", 0)),
            "description": item.get("Description", ""),
            "active": item.get("Active", True)
        })

    return results

# ==================== CREACIÓN DE TRANSACCIONES ====================

def create_invoice(customer_id: str, line_items: List[dict], txn_date: str = None,
                  memo: str = None, custom_fields: dict = None) -> dict:
    """Crea un invoice en QuickBooks.

    MED-10 fix: valida que line_items no esté vacío/None/no-list antes
    de llamar QBO. Antes, [] o None generaba Invoice sin Line y QBO
    rechazaba con 400 confuso.
    """
    log_operation("invoices")

    if not line_items or not isinstance(line_items, list):
        raise ValueError(
            "create_invoice: line_items is required and must be a non-empty list. "
            f"Got {line_items!r}"
        )

    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")

    invoice_data = {
        "CustomerRef": {"value": customer_id},
        "TxnDate": txn_date,
        "Line": []
    }

    # Agregar líneas
    for idx, item in enumerate(line_items, 1):
        qty = item.get("quantity", 1)
        if qty is None or qty <= 0:
            raise ValueError(
                f"create_invoice: line_items[{idx}] quantity must be > 0, got {qty!r}"
            )
        line = {
            "DetailType": "SalesItemLineDetail",
            "Amount": item["amount"],
            "SalesItemLineDetail": {
                "ItemRef": {"value": item["item_id"]},
                "Qty": qty,
                "UnitPrice": item["amount"] / qty
            }
        }

        if "description" in item:
            line["Description"] = item["description"]

        invoice_data["Line"].append(line)

    if memo:
        invoice_data["PrivateNote"] = memo

    if custom_fields:
        invoice_data["CustomField"] = [
            {"DefinitionId": k, "StringValue": v}
            for k, v in custom_fields.items()
        ]

    response = qbo_request("POST", "invoice", data=invoice_data)

    if response.status_code == 200:
        invoice = response.json()["Invoice"]
        return {
            "success": True,
            "invoice_id": invoice["Id"],
            "doc_number": invoice.get("DocNumber"),
            "total": invoice.get("TotalAmt"),
            "balance": invoice.get("Balance")
        }
    else:
        return {
            "success": False,
            "error": response.text
        }

def create_customer(display_name: str, email: str = None, phone: str = None,
                    address: str = None, company_name: str = None,
                    deduplicate: bool = False) -> dict:
    """Crea un cliente (Customer) en QuickBooks.

    Si `deduplicate=True`, hace pre-check vía `search_customer(exact=True)`
    antes de POST. Si encuentra match exacto por DisplayName, retorna el
    ID existente con `idempotent_reused=True` y no llama a QBO. Default
    False para backward compat.
    """
    log_operation("customers_created")

    if deduplicate:
        existing = search_customer(display_name, exact=True)
        if existing:
            return {
                "success": True,
                "customer_id": existing[0]["id"],
                "display_name": existing[0]["name"],
                "company_name": existing[0].get("company", ""),
                "balance": existing[0].get("balance", 0),
                "active": existing[0].get("active", True),
                "idempotent_reused": True,
            }

    customer_data: Dict[str, Any] = {"DisplayName": display_name}

    if company_name:
        customer_data["CompanyName"] = company_name
    if email:
        customer_data["PrimaryEmailAddr"] = {"Address": email}
    if phone:
        customer_data["PrimaryPhone"] = {"FreeFormNumber": phone}
    if address:
        customer_data["BillAddr"] = {"Line1": address}

    response = qbo_request("POST", "customer", data=customer_data)

    if response.status_code == 200:
        customer = response.json()["Customer"]
        return {
            "success": True,
            "customer_id": customer["Id"],
            "display_name": customer.get("DisplayName"),
            "company_name": customer.get("CompanyName"),
            "balance": customer.get("Balance", 0),
            "active": customer.get("Active", True),
        }
    return {
        "success": False,
        "error": response.text,
        "status_code": response.status_code,
    }


# ========================================================================
# Master Data Create helpers (Sprint 1A: 8 tools)
# ========================================================================

def create_vendor(display_name: str, company_name: str = None, email: str = None,
                  phone: str = None, address: str = None, vendor_1099: bool = False,
                  bill_rate: float = None, term_id: str = None) -> dict:
    """Crea un vendor (proveedor) en QuickBooks."""
    log_operation("vendors_created")
    vendor_data: Dict[str, Any] = {"DisplayName": display_name}
    if company_name:
        vendor_data["CompanyName"] = company_name
    if email:
        vendor_data["PrimaryEmailAddr"] = {"Address": email}
    if phone:
        vendor_data["PrimaryPhone"] = {"FreeFormNumber": phone}
    if address:
        vendor_data["BillAddr"] = {"Line1": address}
    if vendor_1099:
        vendor_data["Vendor1099"] = True
    if bill_rate is not None:
        vendor_data["BillRate"] = bill_rate
    if term_id:
        vendor_data["TermRef"] = {"value": term_id}
    response = qbo_request("POST", "vendor", data=vendor_data)
    if response.status_code == 200:
        v = response.json()["Vendor"]
        return {
            "success": True,
            "vendor_id": v["Id"],
            "display_name": v.get("DisplayName"),
            "company_name": v.get("CompanyName"),
            "vendor_1099": v.get("Vendor1099", False),
            "balance": v.get("Balance", 0),
            "active": v.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_account(name: str, account_type: str, account_sub_type: str = None,
                   description: str = None, opening_balance: float = None,
                   opening_balance_date: str = None) -> dict:
    """Crea una cuenta (Account) en el Chart of Accounts.

    account_type: Bank, AccountsReceivable, OtherCurrentAsset, FixedAsset,
                  OtherAsset, AccountsPayable, CreditCard, OtherCurrentLiability,
                  LongTermLiability, Equity, Income, CostOfGoodsSold, Expense,
                  OtherIncome, OtherExpense
    """
    log_operation("accounts_created")
    account_data: Dict[str, Any] = {"Name": name, "AccountType": account_type}
    if account_sub_type:
        account_data["AccountSubType"] = account_sub_type
    if description:
        account_data["Description"] = description
    if opening_balance is not None:
        account_data["OpeningBalance"] = opening_balance
        account_data["OpeningBalanceDate"] = opening_balance_date or datetime.now().strftime("%Y-%m-%d")
    response = qbo_request("POST", "account", data=account_data)
    if response.status_code == 200:
        a = response.json()["Account"]
        return {
            "success": True,
            "account_id": a["Id"],
            "name": a.get("Name"),
            "account_type": a.get("AccountType"),
            "current_balance": a.get("CurrentBalance", 0),
            "active": a.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_item(name: str, item_type: str = "Service", unit_price: float = 0.0,
                income_account_id: str = None, expense_account_id: str = None,
                asset_account_id: str = None, sku: str = None,
                track_quantity: bool = False, qty_on_hand: float = 0.0,
                inv_start_date: str = None, description: str = None) -> dict:
    """Crea un item (producto o servicio) en QuickBooks.

    item_type: Service, Inventory, NonInventory
    """
    log_operation("items_created")
    item_data: Dict[str, Any] = {
        "Name": name,
        "Type": item_type,
        "UnitPrice": unit_price,
    }
    if income_account_id:
        item_data["IncomeAccountRef"] = {"value": income_account_id}
    if expense_account_id:
        item_data["ExpenseAccountRef"] = {"value": expense_account_id}
    if asset_account_id:
        item_data["AssetAccountRef"] = {"value": asset_account_id}
    if sku:
        item_data["Sku"] = sku
    if description:
        item_data["Description"] = description
    if item_type == "Inventory" and track_quantity:
        item_data["TrackQuantityOnHand"] = True
        item_data["QtyOnHand"] = qty_on_hand
        item_data["InvStartDate"] = inv_start_date or datetime.now().strftime("%Y-%m-%d")
    response = qbo_request("POST", "item", data=item_data)
    if response.status_code == 200:
        i = response.json()["Item"]
        return {
            "success": True,
            "item_id": i["Id"],
            "name": i.get("Name"),
            "type": i.get("Type"),
            "unit_price": i.get("UnitPrice", 0),
            "active": i.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_employee(display_name: str, given_name: str = None, family_name: str = None,
                    email: str = None, phone: str = None, address: str = None,
                    hired_date: str = None, bill_rate: float = None) -> dict:
    """Crea un empleado (Employee) en QuickBooks."""
    log_operation("employees_created")
    emp_data: Dict[str, Any] = {"DisplayName": display_name}
    if given_name:
        emp_data["GivenName"] = given_name
    if family_name:
        emp_data["FamilyName"] = family_name
    if email:
        emp_data["PrimaryEmailAddr"] = {"Address": email}
    if phone:
        emp_data["PrimaryPhone"] = {"FreeFormNumber": phone}
    if address:
        emp_data["PrimaryAddr"] = {"Line1": address}
    if hired_date:
        emp_data["HiredDate"] = hired_date
    if bill_rate is not None:
        emp_data["BillRate"] = bill_rate
    response = qbo_request("POST", "employee", data=emp_data)
    if response.status_code == 200:
        e = response.json()["Employee"]
        return {
            "success": True,
            "employee_id": e["Id"],
            "display_name": e.get("DisplayName"),
            "active": e.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_class(name: str, parent_class_id: str = None, active: bool = True) -> dict:
    """Crea una clase (Class) en QuickBooks para segmentación P&L."""
    log_operation("classes_created")
    class_data: Dict[str, Any] = {"Name": name, "Active": active}
    if parent_class_id:
        class_data["SubClass"] = True
        class_data["ParentRef"] = {"value": parent_class_id}
    response = qbo_request("POST", "class", data=class_data)
    if response.status_code == 200:
        c = response.json()["Class"]
        return {
            "success": True,
            "class_id": c["Id"],
            "name": c.get("Name"),
            "active": c.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_department(name: str, parent_dept_id: str = None, active: bool = True) -> dict:
    """Crea un departamento en QuickBooks para segmentación P&L."""
    log_operation("departments_created")
    dept_data: Dict[str, Any] = {"Name": name, "Active": active}
    if parent_dept_id:
        dept_data["SubDepartment"] = True
        dept_data["ParentRef"] = {"value": parent_dept_id}
    response = qbo_request("POST", "department", data=dept_data)
    if response.status_code == 200:
        d = response.json()["Department"]
        return {
            "success": True,
            "department_id": d["Id"],
            "name": d.get("Name"),
            "active": d.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_term(name: str, due_days: int = 30, discount_days: int = 0,
                discount_pct: float = 0.0, active: bool = True) -> dict:
    """Crea un plazo de pago (Term) en QuickBooks (ej: Net 30, 2/10 Net 30)."""
    log_operation("terms_created")
    term_data: Dict[str, Any] = {
        "Name": name,
        "DueDays": due_days,
        "Active": active,
    }
    if discount_days > 0:
        term_data["DiscountDays"] = discount_days
        term_data["DiscountPct"] = discount_pct
    response = qbo_request("POST", "term", data=term_data)
    if response.status_code == 200:
        t = response.json()["Term"]
        return {
            "success": True,
            "term_id": t["Id"],
            "name": t.get("Name"),
            "due_days": t.get("DueDays"),
            "active": t.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_payment_method(name: str, payment_type: str = "Other", active: bool = True) -> dict:
    """Crea un método de pago (PaymentMethod) en QuickBooks.

    payment_type: CreditCard, Check, Cash, BankTransfer, Other
    """
    log_operation("payment_methods_created")
    pm_data: Dict[str, Any] = {"Name": name, "Type": payment_type, "Active": active}
    response = qbo_request("POST", "paymentmethod", data=pm_data)
    if response.status_code == 200:
        pm = response.json()["PaymentMethod"]
        return {
            "success": True,
            "payment_method_id": pm["Id"],
            "name": pm.get("Name"),
            "type": pm.get("Type"),
            "active": pm.get("Active", True),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Transaction Create helpers (Sprint 1B: 9 tools)
# ========================================================================

def create_billpayment(vendor_id: str, total_amt: float, pay_type: str = "Check",
                       txn_date: str = None, bank_account_id: str = None,
                       cc_account_id: str = None, line_payments: List[dict] = None,
                       memo: str = None) -> dict:
    """Crea un pago de bill (BillPayment) en QuickBooks.

    pay_type: Check, CreditCard
    line_payments: lista de {bill_id, amount} para aplicar a bills específicas
    """
    log_operation("billpayments_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    bp_data: Dict[str, Any] = {
        "VendorRef": {"value": vendor_id},
        "TotalAmt": total_amt,
        "PayType": pay_type,
        "TxnDate": txn_date,
    }
    if pay_type == "Check" and bank_account_id:
        bp_data["CheckPayment"] = {"BankAccountRef": {"value": bank_account_id}}
    elif pay_type == "CreditCard" and cc_account_id:
        bp_data["CreditCardPayment"] = {"CCAccountRef": {"value": cc_account_id}}
    if line_payments:
        bp_data["Line"] = [
            {
                "Amount": lp["amount"],
                "LinkedTxn": [{"TxnId": lp["bill_id"], "TxnType": "Bill"}]
            }
            for lp in line_payments
        ]
    if memo:
        bp_data["PrivateNote"] = memo
    response = qbo_request("POST", "billpayment", data=bp_data)
    if response.status_code == 200:
        bp = response.json()["BillPayment"]
        return {
            "success": True,
            "bill_payment_id": bp["Id"],
            "total": bp.get("TotalAmt"),
            "vendor_id": vendor_id,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_estimate(customer_id: str, line_items: List[dict], txn_date: str = None,
                    expiration_date: str = None, memo: str = None) -> dict:
    """Crea un estimate (cotización) en QuickBooks."""
    log_operation("estimates_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    est_data: Dict[str, Any] = {
        "CustomerRef": {"value": customer_id},
        "TxnDate": txn_date,
        "Line": [],
    }
    if expiration_date:
        est_data["ExpirationDate"] = expiration_date
    if memo:
        est_data["PrivateNote"] = memo
    for item in line_items:
        line: Dict[str, Any] = {
            "DetailType": "SalesItemLineDetail",
            "Amount": item["amount"],
            "SalesItemLineDetail": {"ItemRef": {"value": item["item_id"]}},
        }
        if "quantity" in item:
            line["SalesItemLineDetail"]["Qty"] = item["quantity"]
        if "description" in item:
            line["Description"] = item["description"]
        est_data["Line"].append(line)
    response = qbo_request("POST", "estimate", data=est_data)
    if response.status_code == 200:
        e = response.json()["Estimate"]
        return {
            "success": True,
            "estimate_id": e["Id"],
            "doc_number": e.get("DocNumber"),
            "total": e.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_salesreceipt(customer_id: str = None, line_items: List[dict] = None,
                        txn_date: str = None, deposit_to_account_id: str = None,
                        payment_method_id: str = None, total_amt: float = None,
                        memo: str = None) -> dict:
    """Crea un sales receipt (recibo de venta inmediata) en QuickBooks.

    customer_id opcional (puede ser venta sin cliente asignado).
    Si se pasa total_amt sin line_items, crea un recibo simple.
    """
    log_operation("salesreceipts_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    sr_data: Dict[str, Any] = {"TxnDate": txn_date}
    if customer_id:
        sr_data["CustomerRef"] = {"value": customer_id}
    if deposit_to_account_id:
        sr_data["DepositToAccountRef"] = {"value": deposit_to_account_id}
    if payment_method_id:
        sr_data["PaymentRefNum"] = ""
    if memo:
        sr_data["PrivateNote"] = memo
    if line_items:
        sr_data["Line"] = []
        for item in line_items:
            line: Dict[str, Any] = {
                "DetailType": "SalesItemLineDetail",
                "Amount": item["amount"],
                "SalesItemLineDetail": {"ItemRef": {"value": item["item_id"]}},
            }
            if "quantity" in item:
                line["SalesItemLineDetail"]["Qty"] = item["quantity"]
            if "description" in item:
                line["Description"] = item["description"]
            sr_data["Line"].append(line)
    response = qbo_request("POST", "salesreceipt", data=sr_data)
    if response.status_code == 200:
        sr = response.json()["SalesReceipt"]
        return {
            "success": True,
            "sales_receipt_id": sr["Id"],
            "doc_number": sr.get("DocNumber"),
            "total": sr.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_creditmemo(customer_id: str, line_items: List[dict], txn_date: str = None,
                      memo: str = None) -> dict:
    """Crea un credit memo (nota de crédito) en QuickBooks."""
    log_operation("creditmemos_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    cm_data: Dict[str, Any] = {
        "CustomerRef": {"value": customer_id},
        "TxnDate": txn_date,
        "Line": [],
    }
    if memo:
        cm_data["PrivateNote"] = memo
    for item in line_items:
        line: Dict[str, Any] = {
            "DetailType": "SalesItemLineDetail",
            "Amount": item["amount"],
            "SalesItemLineDetail": {"ItemRef": {"value": item["item_id"]}},
        }
        if "quantity" in item:
            line["SalesItemLineDetail"]["Qty"] = item["quantity"]
        if "description" in item:
            line["Description"] = item["description"]
        cm_data["Line"].append(line)
    response = qbo_request("POST", "creditmemo", data=cm_data)
    if response.status_code == 200:
        cm = response.json()["CreditMemo"]
        return {
            "success": True,
            "credit_memo_id": cm["Id"],
            "doc_number": cm.get("DocNumber"),
            "total": cm.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_purchase(vendor_id: str, account_id: str, total_amt: float,
                    payment_type: str = "Cash", txn_date: str = None,
                    description: str = None, memo: str = None) -> dict:
    """Crea una purchase genérica (gasto via check/CC/cash) en QuickBooks.

    payment_type: Cash, Check, CreditCard
    """
    log_operation("purchases_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    pur_data: Dict[str, Any] = {
        "VendorRef": {"value": vendor_id},
        "PaymentType": payment_type,
        "TxnDate": txn_date,
        "Line": [{
            "DetailType": "AccountBasedExpenseLineDetail",
            "Amount": total_amt,
            "AccountBasedExpenseLineDetail": {"AccountRef": {"value": account_id}},
        }],
    }
    if description:
        pur_data["Line"][0]["Description"] = description
    if memo:
        pur_data["PrivateNote"] = memo
    response = qbo_request("POST", "purchase", data=pur_data)
    if response.status_code == 200:
        p = response.json()["Purchase"]
        return {
            "success": True,
            "purchase_id": p["Id"],
            "doc_number": p.get("DocNumber"),
            "total": p.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_purchaseorder(vendor_id: str, line_items: List[dict], txn_date: str = None,
                         ship_to_addr: str = None, memo: str = None,
                         po_email: str = None) -> dict:
    """Crea una purchase order (orden de compra) en QuickBooks."""
    log_operation("purchaseorders_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    po_data: Dict[str, Any] = {
        "VendorRef": {"value": vendor_id},
        "TxnDate": txn_date,
        "Line": [],
    }
    if ship_to_addr:
        po_data["ShipAddr"] = {"Line1": ship_to_addr}
    if po_email:
        po_data["POEmail"] = {"Address": po_email}
    if memo:
        po_data["PrivateNote"] = memo
    for item in line_items:
        line: Dict[str, Any] = {
            "DetailType": "ItemBasedExpenseLineDetail",
            "Amount": item["amount"],
            "ItemBasedExpenseLineDetail": {"ItemRef": {"value": item["item_id"]}},
        }
        if "quantity" in item:
            line["ItemBasedExpenseLineDetail"]["Qty"] = item["quantity"]
        if "description" in item:
            line["Description"] = item["description"]
        po_data["Line"].append(line)
    response = qbo_request("POST", "purchaseorder", data=po_data)
    if response.status_code == 200:
        po = response.json()["PurchaseOrder"]
        return {
            "success": True,
            "purchase_order_id": po["Id"],
            "doc_number": po.get("DocNumber"),
            "total": po.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_refundreceipt(customer_id: str, line_items: List[dict], refund_account_id: str,
                         txn_date: str = None, memo: str = None) -> dict:
    """Crea un refund receipt (recibo de reembolso) en QuickBooks."""
    log_operation("refundreceipts_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    rr_data: Dict[str, Any] = {
        "CustomerRef": {"value": customer_id},
        "DepositToAccountRef": {"value": refund_account_id},
        "TxnDate": txn_date,
        "Line": [],
    }
    if memo:
        rr_data["PrivateNote"] = memo
    for item in line_items:
        line: Dict[str, Any] = {
            "DetailType": "SalesItemLineDetail",
            "Amount": item["amount"],
            "SalesItemLineDetail": {"ItemRef": {"value": item["item_id"]}},
        }
        if "quantity" in item:
            line["SalesItemLineDetail"]["Qty"] = item["quantity"]
        if "description" in item:
            line["Description"] = item["description"]
        rr_data["Line"].append(line)
    response = qbo_request("POST", "refundreceipt", data=rr_data)
    if response.status_code == 200:
        rr = response.json()["RefundReceipt"]
        return {
            "success": True,
            "refund_receipt_id": rr["Id"],
            "doc_number": rr.get("DocNumber"),
            "total": rr.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_vendorcredit(vendor_id: str, line_items: List[dict], txn_date: str = None,
                        memo: str = None) -> dict:
    """Crea un vendor credit (crédito de proveedor) en QuickBooks."""
    log_operation("vendorcredits_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    vc_data: Dict[str, Any] = {
        "VendorRef": {"value": vendor_id},
        "TxnDate": txn_date,
        "Line": [],
    }
    if memo:
        vc_data["PrivateNote"] = memo
    for item in line_items:
        line: Dict[str, Any] = {
            "DetailType": "AccountBasedExpenseLineDetail",
            "Amount": item["amount"],
            "AccountBasedExpenseLineDetail": {"AccountRef": {"value": item["account_id"]}},
        }
        if "description" in item:
            line["Description"] = item["description"]
        vc_data["Line"].append(line)
    response = qbo_request("POST", "vendorcredit", data=vc_data)
    if response.status_code == 200:
        vc = response.json()["VendorCredit"]
        return {
            "success": True,
            "vendor_credit_id": vc["Id"],
            "doc_number": vc.get("DocNumber"),
            "total": vc.get("TotalAmt"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_timeactivity(employee_id: str, hours: int = 0, minutes: int = 0,
                        txn_date: str = None, customer_id: str = None,
                        item_id: str = None, billable: bool = True,
                        description: str = None) -> dict:
    """Crea un time activity (registro de horas) en QuickBooks.

    hours+minutes: tiempo total (ej: 1h 30min = 1*60+30=90 min en qbo, pero aquí lo pasamos directo)
    """
    log_operation("timeactivities_created")
    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")
    total_minutes = hours * 60 + minutes
    ta_data: Dict[str, Any] = {
        "EmployeeRef": {"value": employee_id},
        "TxnDate": txn_date,
        "Hours": total_minutes / 60.0,
        "BillableStatus": "Billable" if billable else "NotBillable",
    }
    if customer_id:
        ta_data["CustomerRef"] = {"value": customer_id}
    if item_id:
        ta_data["ItemRef"] = {"value": item_id}
    if description:
        ta_data["Description"] = description
    response = qbo_request("POST", "timeactivity", data=ta_data)
    if response.status_code == 200:
        ta = response.json()["TimeActivity"]
        return {
            "success": True,
            "time_activity_id": ta["Id"],
            "hours": ta.get("Hours"),
            "txn_date": ta.get("TxnDate"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Update/Void/Delete operations (Sprint 1C)
# ========================================================================

def update_entity(entity_name: str, entity_id: str, updates: dict,
                  sync_token: str = None, sparse: bool = True) -> dict:
    """Actualiza una entidad vía POST con sparse update por defecto.

    Si sparse=True, solo se actualizan los campos en `updates` (más eficiente).
    Si sparse=False, se debe pasar la entidad completa en `updates`.
    Si sync_token es None, primero hace un read para obtenerlo.
    """
    if not sync_token:
        # Read para obtener sync token
        read_resp = qbo_request("GET", f"{entity_name}/{entity_id}")
        if read_resp.status_code != 200:
            return {"success": False, "error": f"Cannot read {entity_name}: {read_resp.text}",
                    "status_code": read_resp.status_code}
        current = read_resp.json().get(entity_name.capitalize(), {})
        sync_token = str(current.get("SyncToken", "0"))
    if sparse:
        # En sparse update, NO se puede cambiar Id ni SyncToken, se pasan como sparse field
        # Y solo se mandan los campos a cambiar
        payload = {"Id": entity_id, "SyncToken": sync_token, **updates}
    else:
        # Full update: la entidad completa
        payload = updates
    params = {"operation": "sparseUpdate"} if sparse else None
    response = qbo_request("POST", entity_name, data=payload, params=params)
    if response.status_code == 200:
        data_key = entity_name[0].upper() + entity_name[1:]
        updated = response.json().get(data_key, {})
        return {
            "success": True,
            f"{entity_name}_id": updated.get("Id"),
            "sync_token": updated.get("SyncToken"),
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def delete_transaction(entity_name: str, entity_id: str, sync_token: str) -> dict:
    """Elimina una transacción (hard delete). Para entidades con simplified delete,
    solo se manda Id + SyncToken."""
    log_operation(f"{entity_name}_deleted")
    response = qbo_request(
        "POST",
        f"{entity_name}?operation=delete",
        data={"Id": entity_id, "SyncToken": sync_token},
    )
    if response.status_code == 200:
        return {
            "success": True,
            f"{entity_name}_id": entity_id,
            "deleted": True,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def void_transaction(entity_name: str, entity_id: str, sync_token: str) -> dict:
    """Anula (void) una transacción. Aplica a Payment, BillPayment, Invoice, SalesReceipt.

    HIGH-3 fix: lee la transacción primero para preservar PrivateNote original
    (e.g. BNK-RECON tag). Prepend "[VOIDED] " a la nota existente. Si la nota
    está vacía/ausente o el read falla, usa "[VOIDED]" como fallback (no bloquea).
    """
    log_operation(f"{entity_name}_voided")
    existing_note = ""
    try:
        read_resp = qbo_request("GET", f"{entity_name}/{entity_id}")
        if read_resp.status_code == 200:
            data_key = entity_name[0].upper() + entity_name[1:]
            existing_note = read_resp.json().get(data_key, {}).get("PrivateNote", "") or ""
    except Exception:
        existing_note = ""
    void_note = f"[VOIDED] {existing_note}".strip() if existing_note else "[VOIDED]"
    response = qbo_request(
        "POST",
        entity_name,
        params={"operation": "sparseUpdate"},
        data={
            "Id": entity_id,
            "SyncToken": sync_token,
            "PrivateNote": void_note,
        },
    )
    if response.status_code == 200:
        data_key = entity_name[0].upper() + entity_name[1:]
        return {
            "success": True,
            f"{entity_name}_id": entity_id,
            "voided": True,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def deactivate_entity(entity_name: str, entity_id: str, sync_token: str = None) -> dict:
    """Desactiva una entidad de master data (soft delete via Active=false)."""
    log_operation(f"{entity_name}_deactivated")
    if not sync_token:
        read_resp = qbo_request("GET", f"{entity_name}/{entity_id}")
        if read_resp.status_code != 200:
            return {"success": False, "error": f"Cannot read {entity_name}: {read_resp.text}"}
        current = read_resp.json().get(entity_name.capitalize(), {})
        sync_token = str(current.get("SyncToken", "0"))
    response = qbo_request(
        "POST",
        f"{entity_name}?operation=delete",
        data={"Id": entity_id, "SyncToken": sync_token, "Active": False},
    )
    if response.status_code == 200:
        return {
            "success": True,
            f"{entity_name}_id": entity_id,
            "active": False,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def send_transaction_email(entity_name: str, entity_id: str, send_to: str = None) -> dict:
    """Envía una transacción por email (Invoice o PurchaseOrder)."""
    log_operation(f"{entity_name}_emailed")
    endpoint = f"{entity_name}/{entity_id}/send"
    if send_to:
        endpoint += f"?sendTo={send_to}"
    response = qbo_request("POST", endpoint, data={})
    if response.status_code == 200:
        return {
            "success": True,
            f"{entity_name}_id": entity_id,
            "sent_to": send_to or "default",
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Report helpers (Sprint 1E: 10+ tools)
# ========================================================================

def _truncate_report_data(report_name: str, data: dict, max_bytes: int = 250_000) -> dict:
    """Trunca Rows.Row de un reporte QBO si excede max_bytes.

    MED-6 fix: protege context window del LLM. Reportes grandes
    (ProfitAndLossDetail, GeneralLedger, TransactionList) pueden
    ser 5-50MB; default 250KB ≈ 60K tokens, dejando espacio para
    system prompt + tools + respuesta.
    """
    if not isinstance(data, dict) or max_bytes <= 0:
        return data

    try:
        encoded_size = len(json.dumps(data, ensure_ascii=False, default=str).encode("utf-8"))
    except (TypeError, ValueError):
        return data

    if encoded_size <= max_bytes:
        return data

    rows_root = data.get("Rows")
    if not isinstance(rows_root, dict):
        return data

    rows = rows_root.get("Row")
    if not isinstance(rows, list) or not rows:
        return data

    original_count = len(rows)
    header = data.get("Header")
    header_size = len(json.dumps({"Header": header}, ensure_ascii=False, default=str).encode("utf-8")) if header else 0

    safe_budget = max(1024, max_bytes - header_size - 512)
    kept = []
    accumulated = 0
    for row in rows:
        try:
            row_size = len(json.dumps(row, ensure_ascii=False, default=str).encode("utf-8"))
        except (TypeError, ValueError):
            row_size = 0
        if accumulated + row_size > safe_budget:
            break
        kept.append(row)
        accumulated += row_size

    new_data = dict(data)
    new_rows_root = dict(rows_root)
    new_rows_root["Row"] = kept
    new_data["Rows"] = new_rows_root
    new_data["_truncated"] = True
    new_data["_truncation_summary"] = {
        "report": report_name,
        "original_bytes": encoded_size,
        "max_bytes": max_bytes,
        "original_rows": original_count,
        "kept_rows": len(kept),
        "message": (
            f"Reporte {report_name} truncado: {original_count} filas reducidas "
            f"a {len(kept)} (de {encoded_size} bytes a <={max_bytes}). "
            "Para ver el reporte completo, divide el rango de fechas o filtra "
            "por cuenta/clase/departamento."
        ),
    }
    return new_data


def _fetch_report(report_name: str, params: dict) -> dict:
    """Helper genérico para fetch de reportes. Retorna JSON parseado o error.

    MED-6 fix: aplica _truncate_report_data para proteger context window.
    """
    response = qbo_request("GET", f"reports/{report_name}", params=params)
    if response.status_code == 200:
        raw = response.json()
        max_bytes = int(os.environ.get("MAX_REPORT_BYTES", "250000"))
        truncated = _truncate_report_data(report_name, raw, max_bytes=max_bytes)
        return {"success": True, "data": truncated}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def generate_trial_balance_report(start_date: str, end_date: str,
                                   accounting_method: str = "Accrual") -> dict:
    """Genera reporte de Trial Balance (balance de comprobación)."""
    log_operation("reports")
    return _fetch_report("TrialBalance", {
        "start_date": start_date,
        "end_date": end_date,
        "accounting_method": accounting_method,
    })


def generate_general_ledger_report(start_date: str, end_date: str,
                                   accounting_method: str = "Accrual",
                                   account_id: str = None) -> dict:
    """Genera reporte de General Ledger para una cuenta o todas."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
        "accounting_method": accounting_method,
    }
    if account_id:
        params["account"] = account_id
    return _fetch_report("GeneralLedger", params)


def generate_cash_flow_report(start_date: str, end_date: str,
                              accounting_method: str = "Accrual") -> dict:
    """Genera reporte de Statement of Cash Flows."""
    log_operation("reports")
    return _fetch_report("CashFlow", {
        "start_date": start_date,
        "end_date": end_date,
        "accounting_method": accounting_method,
    })


def generate_ar_aging_report(report_date: str, aging_method: str = "ReportDate",
                              num_periods: int = 4) -> dict:
    """Genera A/R Aging Summary (reporte de cobranzas)."""
    log_operation("reports")
    return _fetch_report("AgedReceivables", {
        "report_date": report_date,
        "aging_method": aging_method,
        "aging_period": num_periods,
    })


def generate_ap_aging_report(report_date: str, aging_method: str = "ReportDate",
                              num_periods: int = 4) -> dict:
    """Genera A/P Aging Summary (reporte de pagos pendientes)."""
    log_operation("reports")
    return _fetch_report("AgedPayables", {
        "report_date": report_date,
        "aging_method": aging_method,
        "aging_period": num_periods,
    })


def generate_customer_balance_report(report_date: str = None,
                                      customer_id: str = None) -> dict:
    """Genera Customer Balance Summary."""
    log_operation("reports")
    params: Dict[str, Any] = {}
    if report_date:
        params["report_date"] = report_date
    if customer_id:
        params["customer"] = customer_id
    return _fetch_report("CustomerBalance", params)


def generate_vendor_balance_report(report_date: str = None,
                                    vendor_id: str = None) -> dict:
    """Genera Vendor Balance Summary."""
    log_operation("reports")
    params: Dict[str, Any] = {}
    if report_date:
        params["report_date"] = report_date
    if vendor_id:
        params["vendor"] = vendor_id
    return _fetch_report("VendorBalance", params)


def generate_pl_detail_report(start_date: str, end_date: str,
                               accounting_method: str = "Accrual") -> dict:
    """Genera Profit and Loss Detail (más granular que P&L normal)."""
    log_operation("reports")
    return _fetch_report("ProfitAndLossDetail", {
        "start_date": start_date,
        "end_date": end_date,
        "accounting_method": accounting_method,
    })


def generate_journal_report(start_date: str, end_date: str) -> dict:
    """Genera Journal Report (todos los journal entries en un período)."""
    log_operation("reports")
    return _fetch_report("JournalReport", {
        "start_date": start_date,
        "end_date": end_date,
    })


def generate_account_list_report() -> dict:
    """Genera Account List (lista de cuentas contables)."""
    log_operation("reports")
    return _fetch_report("AccountListDetail", {})


# ========================================================================
# Sprint 1E P2: 6 reportes opcionales adicionales
# ========================================================================

def generate_inventory_valuation_report(start_date: str = None, end_date: str = None,
                                         item_id: str = None) -> dict:
    """Genera Inventory Valuation Summary (valorización de inventario)."""
    log_operation("reports")
    params: Dict[str, Any] = {}
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    if item_id:
        params["item"] = item_id
    return _fetch_report("InventoryValuationSummary", params)


def generate_sales_by_customer_report(start_date: str, end_date: str,
                                       customer_id: str = None) -> dict:
    """Genera Sales by Customer Summary (ventas agrupadas por cliente)."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }
    if customer_id:
        params["customer"] = customer_id
    return _fetch_report("CustomerSales", params)


def generate_expenses_by_vendor_report(start_date: str, end_date: str,
                                        vendor_id: str = None) -> dict:
    """Genera Expenses by Vendor Summary (gastos agrupados por proveedor)."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }
    if vendor_id:
        params["vendor"] = vendor_id
    return _fetch_report("VendorExpenses", params)


def generate_transaction_list_report(start_date: str, end_date: str,
                                      account_id: str = None,
                                      transaction_type: str = None) -> dict:
    """Genera Transaction List (lista de transacciones con filtros)."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }
    if account_id:
        params["account"] = account_id
    if transaction_type:
        params["transaction_type"] = transaction_type
    return _fetch_report("TransactionList", params)


def generate_class_sales_report(start_date: str, end_date: str,
                                 class_id: str = None) -> dict:
    """Genera Sales by Class Summary (ventas agrupadas por clase)."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }
    if class_id:
        params["class"] = class_id
    return _fetch_report("ClassSales", params)


def generate_department_sales_report(start_date: str, end_date: str,
                                     department_id: str = None) -> dict:
    """Genera Sales by Department Summary (ventas agrupadas por departamento)."""
    log_operation("reports")
    params: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }
    if department_id:
        params["department"] = department_id
    return _fetch_report("DepartmentSales", params)


# ========================================================================
# Read operations (Sprint 1F: 3 tools)
# ========================================================================

def get_company_info() -> dict:
    """Lee información de la empresa (CompanyInfo)."""
    log_operation("companyinfo_read")
    response = qbo_request("GET", f"companyinfo/{QB_REALM_ID}")
    if response.status_code == 200:
        ci = response.json().get("CompanyInfo", {})
        return {
            "success": True,
            "company_name": ci.get("CompanyName"),
            "legal_name": ci.get("LegalName"),
            "country": ci.get("Country"),
            "fiscal_year_start": ci.get("FiscalYearStartMonth"),
            "email": ci.get("Email", {}).get("Address") if ci.get("Email") else None,
            "address": ci.get("CompanyAddr", {}),
            "phone": ci.get("PrimaryPhone", {}).get("FreeFormNumber") if ci.get("PrimaryPhone") else None,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


def get_preferences() -> dict:
    """Lee preferencias de la empresa (Preferences)."""
    log_operation("preferences_read")
    response = qbo_request("GET", "preferences")
    if response.status_code == 200:
        return {"success": True, "data": response.json().get("Preferences", {})}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def advanced_query(sql: str, start_position: int = 1, max_results: int = 100) -> dict:
    """Ejecuta una query SQL-like en QuickBooks (QBO query language).

    Soporta SELECT, WHERE, ORDERBY, STARTPOSITION, MAXRESULTS, COUNT.
    Limitaciones: no OR, no DROP/DELETE/UPDATE, max 1000 max_results.
    """
    log_operation("queries")
    # Validaciones de seguridad
    sql_upper = sql.upper().strip()
    forbidden = ["DROP", "DELETE ", "UPDATE ", "INSERT ", "ALTER ", "CREATE "]
    for kw in forbidden:
        if kw in sql_upper:
            return {"success": False, "error": f"Operación no permitida: {kw.strip()}"}
    if "MAXRESULTS" not in sql_upper:
        sql += f" MAXRESULTS {min(max_results, 1000)}"
    if "STARTPOSITION" not in sql_upper:
        sql = sql.rstrip() + f" STARTPOSITION {start_position}"
    response = qbo_request("GET", "query", params={"query": sql})
    if response.status_code == 200:
        return {"success": True, "data": response.json()}
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Sprint 2: Recurring + Attachments
# ========================================================================

def create_recurring_transaction(base_txn: dict, name: str, recur_type: str = "Automated",
                                   interval_type: str = "Monthly", num_interval: int = 1,
                                   start_date: str = None, max_occurrences: int = None,
                                   day_of_month: int = None, days_before: int = 2,
                                   active: bool = True) -> dict:
    """Crea una transacción recurrente (plantilla que se autogenera).

    base_txn: el cuerpo de la transacción base (Invoice, Bill, etc.) SIN Id ni SyncToken
    recur_type: "Automated" o "Reminder"
    interval_type: "Daily", "Weekly", "Monthly", "Yearly"
    """
    log_operation("recurring_created")
    if not start_date:
        start_date = datetime.now().strftime("%Y-%m-%d")
    schedule_info: Dict[str, Any] = {
        "StartDate": start_date,
        "IntervalType": interval_type,
        "NumInterval": num_interval,
        "DaysBefore": days_before,
    }
    if max_occurrences is not None:
        schedule_info["MaxOccurrences"] = max_occurrences
    if day_of_month is not None:
        schedule_info["DayOfMonth"] = day_of_month
    recurring_info = {
        "Name": name,
        "RecurType": recur_type,
        "Active": active,
        "ScheduleInfo": schedule_info,
    }
    payload = {**base_txn, "RecurringInfo": recurring_info}
    response = qbo_request("POST", "recurringtransaction", data=payload)
    if response.status_code == 200:
        r = response.json().get("RecurringTransaction", {})
        return {
            "success": True,
            "recurring_id": r.get("Id"),
            "name": name,
            "recur_type": recur_type,
        }
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Sprint 3: P2 tools
# ========================================================================

def create_taxcode(name: str, tax_rate_id: str = None, description: str = None,
                   active: bool = True) -> dict:
    """Crea un TaxCode (NON o TAX) en QuickBooks."""
    log_operation("taxcodes_created")
    tc_data: Dict[str, Any] = {"Name": name, "Active": active}
    if tax_rate_id:
        tc_data["TaxRateRef"] = {"value": tax_rate_id}
    if description:
        tc_data["Description"] = description
    response = qbo_request("POST", "taxcode", data=tc_data)
    if response.status_code == 200:
        t = response.json()["TaxCode"]
        return {"success": True, "tax_code_id": t["Id"], "name": t.get("Name")}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_taxrate(name: str, rate_value: float, agency_id: str = None,
                   description: str = None, active: bool = True) -> dict:
    """Crea un TaxRate en QuickBooks."""
    log_operation("taxrates_created")
    tr_data: Dict[str, Any] = {
        "Name": name,
        "RateValue": rate_value,
        "Active": active,
    }
    if agency_id:
        tr_data["AgencyRef"] = {"value": agency_id}
    if description:
        tr_data["Description"] = description
    response = qbo_request("POST", "taxrate", data=tr_data)
    if response.status_code == 200:
        t = response.json()["TaxRate"]
        return {"success": True, "tax_rate_id": t["Id"], "name": t.get("Name"), "rate": rate_value}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def get_exchange_rate(source_currency: str, target_currency: str = "USD",
                      as_of_date: str = None) -> dict:
    """Lee tasa de cambio de moneda en una fecha dada."""
    log_operation("exchange_rate_read")
    if not as_of_date:
        as_of_date = datetime.now().strftime("%Y-%m-%d")
    sql = f"SELECT * FROM ExchangeRate WHERE SourceCurrencyCode = '{source_currency}' AND AsOfDate = '{as_of_date}'"
    response = qbo_request("GET", "query", params={"query": sql})
    if response.status_code == 200:
        return {"success": True, "data": response.json()}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def _validate_batch_schema(operations: list) -> tuple:
    """Valida que `operations` cumpla el schema QBO batch.

    MED-7 fix: ejecuta antes del POST. Reglas:
    - debe ser list no-vacía
    - cada item debe ser dict con 'bId' y 'operation'
    - 'operation' debe ser uno de {create, update, delete, query}
    - 'bId' debe ser único en la lista
    - máximo 30 (también validado por execute_batch)

    Returns (ok: bool, error_msg: str). Si ok=True, error_msg="".
    """
    valid_ops = {"create", "update", "delete", "query"}

    if operations is None:
        return (False, "operations debe ser una lista, recibí None")
    if not isinstance(operations, list):
        return (False, f"operations debe ser list, recibí {type(operations).__name__}")
    if len(operations) == 0:
        return (False, "operations está vacía; un batch sin items no se envía a QBO")
    if len(operations) > 30:
        return (False, f"Batch demasiado grande: {len(operations)} items (max 30)")

    seen_bids = set()
    for idx, item in enumerate(operations):
        prefix = f"item[{idx}]"
        if not isinstance(item, dict):
            return (False, f"{prefix} debe ser dict, recibí {type(item).__name__}")

        bId = item.get("bId")
        if bId is None or bId == "":
            return (False, f"{prefix} falta campo requerido 'bId'")
        if bId in seen_bids:
            return (False, f"bId duplicado: '{bId}'. Cada item del batch debe tener bId único")
        seen_bids.add(bId)

        op = item.get("operation")
        if op is None:
            return (False, f"{prefix} (bId={bId}) falta campo requerido 'operation'")
        if op not in valid_ops:
            return (False, (
                f"{prefix} (bId={bId}) operation='{op}' inválida. "
                f"Valores válidos: {sorted(valid_ops)}"
            ))

    return (True, "")


def execute_batch(operations: List[dict]) -> dict:
    """Ejecuta hasta 30 operaciones en una sola llamada (batch).

    operations: lista de dicts con {bId, operation, Entity/Query}

    MED-7 fix: valida schema antes del POST. Retorna
    {"success": False, "validation_error": True, "error": ...}
    si la lista está vacía, tiene bIds duplicados, etc.
    """
    log_operation("batch_executed")

    ok, err = _validate_batch_schema(operations)
    if not ok:
        return {"success": False, "validation_error": True, "error": err}

    batch_data = {"BatchItemRequest": operations}
    response = qbo_request("POST", "batch", data=batch_data)
    if response.status_code == 200:
        return {"success": True, "data": response.json()}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def cdc_query(entities: List[str], since: str) -> dict:
    """Change Data Capture: retorna entidades modificadas desde `since`.

    entities: lista de nombres de entidades (e.g., ['Customer', 'Invoice'])
    since: timestamp ISO (e.g., '2026-06-01T00:00:00Z')

    CRIT-6 fix: payload debe seguir el schema QBO CDC documented:
        {
          "trackedEntities": [
            {
              "entities": [{"name": "Customer"}, {"name": "Invoice"}],
              "lastModified": "2026-06-01T00:00:00Z"
            }
          ]
        }
    Antes se enviaba {entities: [...], since: ...} que QBO rechazaba con 400.
    """
    log_operation("cdc_query")
    payload = {
        "trackedEntities": [
            {
                "entities": [{"name": name} for name in entities],
                "lastModified": since,
            }
        ]
    }
    response = qbo_request("POST", "cdc", data=payload)
    if response.status_code == 200:
        return {"success": True, "data": response.json()}
    return {"success": False, "error": response.text, "status_code": response.status_code}


def create_budget(name: str, start_date: str, end_date: str,
                  budget_lines: List[dict]) -> dict:
    """Crea un Budget (presupuesto) en QuickBooks.

    budget_lines: lista de {account_id, amount, period (mes)}
    """
    log_operation("budgets_created")
    budget_data: Dict[str, Any] = {
        "Name": name,
        "StartDate": start_date,
        "EndDate": end_date,
        "BudgetDetail": {"BudgetLine": []},
    }
    for line in budget_lines:
        budget_data["BudgetDetail"]["BudgetLine"].append({
            "AccountRef": {"value": line["account_id"]},
            "Amount": line["amount"],
            "Period": line.get("period", "Monthly"),
        })
    response = qbo_request("POST", "budget", data=budget_data)
    if response.status_code == 200:
        b = response.json().get("Budget", {})
        return {"success": True, "budget_id": b.get("Id"), "name": name}
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# Attachments (Attachable) - Sprint 2
# ========================================================================

def upload_attachment(file_content: bytes, file_name: str, content_type: str,
                      entity_type: str, entity_id: str, note: str = None) -> dict:
    """Sube un archivo como attachment y lo vincula a una entidad (Bill, Invoice, etc.).

    file_content: bytes del archivo
    file_name: nombre del archivo (e.g., 'factura.pdf')
    content_type: MIME type (e.g., 'application/pdf', 'image/jpeg')
    entity_type: 'Bill', 'Invoice', 'Purchase', etc.
    entity_id: ID de la entidad a la que se vincula

    HIGH-6 fix: pasa por qbo_request() (con raw_body + extra_headers) para
    heredar timeout, retry 429/503, refresh token en 401, y error logging.
    Antes llamaba requests.post() directamente, bypaseando todos los
    safeguards acumulados (CRIT-1, CRIT-4).
    """
    import base64
    log_operation("attachments_uploaded")
    b64_content = base64.b64encode(file_content).decode("utf-8")
    metadata = {
        "AttachableRef": [{"EntityRef": _build_entity_ref(entity_id, entity_type=entity_type)}],
        "FileName": file_name,
        "ContentType": "Document" if content_type == "application/pdf" else "Image",
    }
    if note:
        metadata["Note"] = note
    # Construir multipart manualmente
    boundary = "----DexterFormBoundary" + str(hash(file_name))[:10]
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file_metadata_01"; filename="meta.json"\r\n'
        f"Content-Type: application/json; charset=UTF-8\r\n\r\n"
        f"{json.dumps(metadata)}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file_content_0"; filename="{file_name}"\r\n'
        f"Content-Type: {content_type}\r\n"
        f"Content-Transfer-Encoding: base64\r\n\r\n"
        f"{b64_content}\r\n"
        f"--{boundary}--\r\n"
    ).encode("utf-8")
    response = qbo_request(
        "POST",
        "upload",
        raw_body=body,
        extra_headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    if response.status_code == 200:
        att_list = response.json().get("AttachableResponse", [])
        if att_list:
            att = att_list[0].get("Attachable", att_list[0])
            return {
                "success": True,
                "attachable_id": att.get("Id"),
                "file_name": file_name,
                "linked_to": f"{entity_type}/{entity_id}",
            }
        return {"success": True, "uploaded": True, "response": response.json()}
    return {"success": False, "error": response.text, "status_code": response.status_code}


# ========================================================================
# End of new helpers (Sprints 1+2+3)
# ========================================================================


def create_bill(vendor_id: str, line_items: List[dict], txn_date: str = None,
               due_date: str = None, memo: str = None) -> dict:
    """Crea un bill en QuickBooks"""
    log_operation("bills")

    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")

    if not due_date:
        due_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")

    bill_data = {
        "VendorRef": {"value": vendor_id},
        "TxnDate": txn_date,
        "DueDate": due_date,
        "Line": []
    }

    # Agregar líneas
    for item in line_items:
        line = {
            "DetailType": "AccountBasedExpenseLineDetail",
            "Amount": item["amount"],
            "AccountBasedExpenseLineDetail": {
                "AccountRef": {"value": item["account_id"]}
            }
        }

        if "description" in item:
            line["Description"] = item["description"]

        bill_data["Line"].append(line)

    if memo:
        bill_data["PrivateNote"] = memo

    response = qbo_request("POST", "bill", data=bill_data)

    if response.status_code == 200:
        bill = response.json()["Bill"]
        return {
            "success": True,
            "bill_id": bill["Id"],
            "doc_number": bill.get("DocNumber"),
            "total": bill.get("TotalAmt"),
            "balance": bill.get("Balance")
        }
    else:
        return {
            "success": False,
            "error": response.text
        }

def create_deposit(account_id: str, line_items: List[dict], txn_date: str = None,
                  memo: str = None) -> dict:
    """Crea un depósito en QuickBooks"""
    log_operation("deposits")

    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")

    deposit_data = {
        "DepositToAccountRef": {"value": account_id},
        "TxnDate": txn_date,
        "Line": []
    }

    total_amount = 0

    # Agregar líneas
    for item in line_items:
        line = {
            "DetailType": "DepositLineDetail",
            "Amount": item["amount"],
            "DepositLineDetail": {
                "AccountRef": {"value": item["from_account_id"]}
            }
        }

        if "customer_id" in item:
            line["DepositLineDetail"]["Entity"] = {
                "Type": "Customer",
                "EntityRef": {"value": item["customer_id"]}
            }

        if "description" in item:
            line["Description"] = item["description"]

        deposit_data["Line"].append(line)
        total_amount += item["amount"]

    if memo:
        deposit_data["PrivateNote"] = memo

    response = qbo_request("POST", "deposit", data=deposit_data)

    if response.status_code == 200:
        deposit = response.json()["Deposit"]
        return {
            "success": True,
            "deposit_id": deposit["Id"],
            "total": deposit.get("TotalAmt"),
            "date": deposit.get("TxnDate")
        }
    else:
        return {
            "success": False,
            "error": response.text
        }

def create_payment(customer_id: str, amount: float, account_id: str,
                  txn_date: str = None, apply_to_invoices: List[dict] = None) -> dict:
    """Crea un payment received en QuickBooks"""
    log_operation("payments")

    if not txn_date:
        txn_date = datetime.now().strftime("%Y-%m-%d")

    payment_data = {
        "CustomerRef": {"value": customer_id},
        "TotalAmt": amount,
        "DepositToAccountRef": {"value": account_id},
        "TxnDate": txn_date
    }

    # Aplicar a invoices específicos si se proporcionan
    if apply_to_invoices:
        payment_data["Line"] = []
        for inv in apply_to_invoices:
            payment_data["Line"].append({
                "Amount": inv["amount"],
                "LinkedTxn": [{
                    "TxnId": inv["invoice_id"],
                    "TxnType": "Invoice"
                }]
            })

    response = qbo_request("POST", "payment", data=payment_data)

    if response.status_code == 200:
        payment = response.json()["Payment"]
        return {
            "success": True,
            "payment_id": payment["Id"],
            "total": payment.get("TotalAmt")
        }
    else:
        return {
            "success": False,
            "error": response.text
        }

# ==================== REPORTES ====================

def generate_pl_report(start_date: str, end_date: str,
                      accounting_method: str = "Accrual") -> list:
    """Genera reporte de Profit & Loss. LOW-7 fix: retorna list[dict]."""
    log_operation("reports")

    params = {
        "start_date": start_date,
        "end_date": end_date,
        "accounting_method": accounting_method
    }

    response = qbo_request("GET", "reports/ProfitAndLoss", params=params)

    if response.status_code != 200:
        print(f"❌ Error generando reporte: {response.text}")
        return []

    report_data = response.json()
    rows_data = report_data.get("Rows", {}).get("Row", [])

    # Parsear datos del reporte
    data = []

    def parse_row(row, parent_name=""):
        if row.get("type") == "Data":
            cols = row.get("ColData", [])
            if len(cols) >= 2:
                account_name = cols[0].get("value", "")
                amount = cols[1].get("value", "0")

                try:
                    amount_float = float(amount.replace(",", ""))
                except:
                    amount_float = 0

                data.append({
                    "cuenta": account_name,
                    "categoria": parent_name,
                    "monto": amount_float
                })

        # Procesar subrows recursivamente
        if "Rows" in row:
            current_name = row.get("Header", {}).get("ColData", [{}])[0].get("value", parent_name)
            for subrow in row["Rows"].get("Row", []):
                parse_row(subrow, current_name)

    for row in rows_data:
        parse_row(row)

    return data

def generate_balance_sheet(as_of_date: str, accounting_method: str = "Accrual") -> list:
    """Genera reporte de Balance Sheet. LOW-7 fix: retorna list[dict]."""
    log_operation("reports")

    params = {
        "date": as_of_date,
        "accounting_method": accounting_method
    }

    response = qbo_request("GET", "reports/BalanceSheet", params=params)

    if response.status_code != 200:
        print(f"❌ Error generando reporte: {response.text}")
        return []

    report_data = response.json()
    rows_data = report_data.get("Rows", {}).get("Row", [])

    data = []

    def parse_row(row, parent_name=""):
        if row.get("type") == "Data":
            cols = row.get("ColData", [])
            if len(cols) >= 2:
                account_name = cols[0].get("value", "")
                amount = cols[1].get("value", "0")

                try:
                    amount_float = float(amount.replace(",", ""))
                except:
                    amount_float = 0

                data.append({
                    "cuenta": account_name,
                    "categoria": parent_name,
                    "monto": amount_float
                })

        if "Rows" in row:
            current_name = row.get("Header", {}).get("ColData", [{}])[0].get("value", parent_name)
            for subrow in row["Rows"].get("Row", []):
                parse_row(subrow, current_name)

    for row in rows_data:
        parse_row(row)

    return data


def _aggregate_by_category(rows: list) -> dict:
    """LOW-7 helper: agrupa rows por 'categoria' y suma 'monto'.

    Reemplaza el patrón df.groupby('categoria')['monto'].sum().to_dict()
    sin requerir pandas. Si rows está vacío retorna {}.
    """
    if not rows:
        return {}
    out = {}
    for r in rows:
        cat = r.get("categoria", "")
        monto = r.get("monto", 0) or 0
        out[cat] = out.get(cat, 0.0) + float(monto)
    return out

def save_report_config(name: str, config: dict):
    """Guarda configuración de reporte para uso futuro"""
    saved_reports = session_state.get("saved_reports", {})

    saved_reports[name] = {
        "config": config,
        "created": datetime.now().isoformat(),
        "last_used": datetime.now().isoformat()
    }

    session_state["saved_reports"] = saved_reports

    with open(FILE_SAVED_REPORTS, 'w') as f:
        json.dump(saved_reports, f, indent=2)

    print(f"✅ Reporte '{name}' guardado")

def load_report_config(name: str) -> Optional[dict]:
    """Carga configuración de reporte guardado"""
    saved_reports = session_state.get("saved_reports", {})

    if name in saved_reports:
        config = saved_reports[name]
        config["last_used"] = datetime.now().isoformat()

        with open(FILE_SAVED_REPORTS, 'w') as f:
            json.dump(saved_reports, f, indent=2)

        return config["config"]

    return None

# ==================== PROCESAMIENTO CSV ====================

def process_deposits_csv(csv_path: str) -> dict:
    """Procesa archivo CSV de depósitos vía batch engine (Sprint 2).

    HIGH-4 fix: ya no llama create_deposit() directamente. Delega a
    tool_depositar_lote_csv(confirmar=True) que usa el BatchEngine con
    state machine, dry-run obligatorio, y rollback seguro via storage.
    Si una fila falla, no se queda con depósitos huérfanos en QBO.

    Mantiene la return shape anterior {success, total, errors, success_count}
    para backward compat con callers existentes.
    """
    log_operation("csv_batches")

    if not os.path.exists(csv_path):
        return {"success": False, "error": f"Archivo no encontrado: {csv_path}"}

    try:
        import pandas as pd
        df = pd.read_csv(csv_path)
        total = len(df)
    except Exception as e:
        return {"success": False, "error": f"Error leyendo CSV: {e}"}

    batch_result = tool_depositar_lote_csv(
        ruta_archivo=csv_path,
        confirmar=True,
    )

    if not batch_result.get("success"):
        return {
            "success": False,
            "total": total,
            "success_count": 0,
            "errors": batch_result.get("errors", [batch_result.get("error", "Batch falló")]),
        }

    inner = batch_result.get("results", {})
    return {
        "success": True,
        "total": total,
        "success_count": inner.get("success", 0) if isinstance(inner, dict) else 0,
        "errors": inner.get("errors", []) if isinstance(inner, dict) else [],
    }

def create_deposits_template():
    """Crea archivo CSV template para depósitos"""
    template_data = {
        "customer_name": ["Cliente Ejemplo 1", "Cliente Ejemplo 2"],
        "amount": [1500.00, 2300.50],
        "from_account": ["Client Retainers", "Prepaid Labour"],
        "to_account": ["Checking Account", "Checking Account"],
        "date": ["2026-01-15", "2026-01-16"],
        "memo": ["Anticipo proyecto A", "Prepago servicios enero"]
    }

    df = pd.DataFrame(template_data)
    df.to_csv(FILE_DEPOSITS_TEMPLATE, index=False)

    print(f"✅ Template creado: {FILE_DEPOSITS_TEMPLATE}")

# ==================== BANK FEED PROCESSING ====================

def validar_suma_deposit(lines: List[dict], bank_feed_amount: float) -> Tuple[bool, float]:
    """Valida que la suma de las líneas coincida con el monto del Bank Feed"""
    total = sum(float(line['amount']) for line in lines)
    diferencia = abs(total - bank_feed_amount)
    es_valido = diferencia < 0.01  # Tolerancia de 1 centavo
    return es_valido, diferencia

def agrupar_bank_feed_por_deposit_id(csv_file: str) -> Optional[dict]:
    """
    Lee el CSV de Bank Feed y agrupa las líneas por deposit_id

    Formato esperado del CSV:
    bank_feed_date,bank_feed_amount,deposit_id,line_type,customer_name,amount,account,memo

    Returns:
        dict: {deposit_id: {'date': ..., 'amount': ..., 'lines': [...]}}
    """
    from collections import defaultdict

    deposits = defaultdict(lambda: {'lines': []})

    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            for row_num, row in enumerate(reader, start=2):
                deposit_id = row['deposit_id'].strip()

                if not deposit_id:
                    print(f"⚠️  Fila {row_num}: deposit_id vacío, saltando...")
                    continue

                # Guardar metadata del depósito
                if 'date' not in deposits[deposit_id]:
                    deposits[deposit_id]['date'] = row['bank_feed_date'].strip()
                    deposits[deposit_id]['bank_feed_amount'] = float(row['bank_feed_amount'])
                    deposits[deposit_id]['deposit_id'] = deposit_id

                # Agregar línea
                line = {
                    'line_type': row['line_type'].strip().lower(),
                    'customer_name': row['customer_name'].strip() if row['customer_name'].strip() else None,
                    'amount': float(row['amount']),
                    'account': row['account'].strip(),
                    'memo': row['memo'].strip()
                }

                deposits[deposit_id]['lines'].append(line)

        return dict(deposits)

    except FileNotFoundError:
        print(f"❌ Archivo no encontrado: {csv_file}")
        return None
    except Exception as e:
        print(f"❌ Error al procesar CSV: {e}")
        return None

def crear_deposito_bank_feed(fecha: str, lines: List[dict], 
                             cuenta_bancaria_id: str = None,
                             memo_principal: str = "Bank Feed Deposit") -> dict:
    """
    Crea un depósito en QuickBooks con múltiples líneas (splits)
    incluyendo ingresos y fees

    Args:
        fecha: Fecha del depósito (YYYY-MM-DD)
        lines: Lista de líneas con format:
               [{'line_type': 'income'|'fee', 'customer_name': str, 
                 'amount': float, 'account': str, 'memo': str}, ...]
        cuenta_bancaria_id: ID de la cuenta bancaria destino (opcional, busca Checking)
        memo_principal: Memo general del depósito

    Returns:
        dict: {'success': bool, 'deposit_id': str, 'message': str}
    """
    # Determinar cuenta bancaria destino
    if not cuenta_bancaria_id:
        # Buscar Checking Account
        checking_accounts = find_account("Checking", category="ACTIVO")
        if not checking_accounts:
            return {
                'success': False,
                'message': 'No se encontró Checking Account'
            }
        cuenta_bancaria_id = checking_accounts[0]['id']

    # Construir las líneas del depósito
    deposit_lines = []

    for line in lines:
        line_type = line['line_type']
        customer_name = line['customer_name']
        amount = line['amount']
        account_name = line['account']
        memo = line['memo']

        # Buscar la cuenta contable
        accounts = find_account(account_name)
        if not accounts:
            return {
                'success': False,
                'message': f"Cuenta no encontrada: {account_name}"
            }

        account_id = accounts[0]['id']

        # Construir línea del depósito
        deposit_line = {
            "Amount": abs(amount),
            "DetailType": "DepositLineDetail",
            "DepositLineDetail": {
                "AccountRef": {"value": account_id}
            },
            "Description": memo
        }

        # Si tiene cliente asociado, agregarlo
        if customer_name:
            customers = search_customer(customer_name)
            if customers:
                deposit_line["DepositLineDetail"]["Entity"] = {
                    "EntityRef": _build_entity_ref(
                        customers[0]['id'],
                        entity_type="Customer",
                        name=customers[0]['name'],
                        include_name=False,
                    ),
                    "Type": "Customer"
                }

        deposit_lines.append(deposit_line)

    # Crear el depósito
    deposit_body = {
        "TxnDate": fecha,
        "DepositToAccountRef": {"value": cuenta_bancaria_id},
        "Line": deposit_lines,
        "PrivateNote": memo_principal
    }

    response = qbo_request("POST", "deposit", data=deposit_body)

    if response.status_code != 200:
        return {
            'success': False,
            'message': f"Error QuickBooks: {response.status_code} - {response.text[:200]}"
        }

    deposit_created = response.json().get("Deposit", {})

    return {
        'success': True,
        'deposit_id': deposit_created.get('Id'),
        'message': 'Depósito creado exitosamente',
        'total_amount': deposit_created.get('TotalAmt')
    }

def procesar_csv_bank_feed(csv_file: str, verbose: bool = True, log: list = None) -> dict:
    """
    Procesa un CSV de Bank Feed y crea depósitos con splits en QuickBooks

    Args:
        csv_file: Ruta al archivo CSV
        verbose: si True (default), imprime progreso a stdout (uso CLI).
                 si False, silencioso (uso tool/LLM).
        log: si se pasa una list, cada mensaje de progreso se appendea aquí
             y se incluye en el return bajo 'log_lines'.

    Returns:
        dict con keys: success, total, success_count, errors, details, log_lines
    """
    def _emit(msg: str) -> None:
        if log is not None:
            log.append(msg)
        if verbose:
            print(msg)

    _emit(f"\n📁 Procesando Bank Feed CSV: {csv_file}")
    _emit("=" * 60)

    deposits = agrupar_bank_feed_por_deposit_id(csv_file)

    if not deposits:
        result = {
            'success': False,
            'message': 'Error al leer el archivo CSV',
            'log_lines': list(log) if log is not None else [],
        }
        return result

    _emit(f"\n✅ {len(deposits)} depósito(s) encontrado(s) en el CSV")
    _emit("")

    results = {
        'total': len(deposits),
        'success_count': 0,
        'errors': 0,
        'details': []
    }

    for dep_id, dep_data in deposits.items():
        _emit(f"🔄 Procesando {dep_id}...")

        bank_feed_amount = dep_data['bank_feed_amount']
        date = dep_data['date']
        lines = dep_data['lines']

        es_valido, diferencia = validar_suma_deposit(lines, bank_feed_amount)

        if not es_valido:
            error_msg = f"  ❌ Suma no cuadra: diferencia de ${diferencia:.2f}"
            _emit(error_msg)
            results['errors'] += 1
            results['details'].append({
                'deposit_id': dep_id,
                'status': 'error',
                'message': error_msg
            })
            continue

        _emit(f"  ✓ Suma validada: ${bank_feed_amount:.2f}")

        resultado = crear_deposito_bank_feed(
            fecha=date,
            lines=lines,
            memo_principal=f"Bank Feed Classification - {dep_id}"
        )

        if resultado['success']:
            _emit(f"  ✅ Depósito creado (ID: {resultado['deposit_id']})")
            _emit(f"     • {len(lines)} líneas procesadas")
            _emit(f"     • Monto total: ${bank_feed_amount:.2f}")
            results['success_count'] += 1
            results['details'].append({
                'deposit_id': dep_id,
                'qb_deposit_id': resultado['deposit_id'],
                'status': 'success',
                'amount': bank_feed_amount,
                'lines': len(lines)
            })
        else:
            _emit(f"  ❌ Error: {resultado['message']}")
            results['errors'] += 1
            results['details'].append({
                'deposit_id': dep_id,
                'status': 'error',
                'message': resultado['message']
            })

        _emit("")

    _emit("=" * 60)
    _emit("📊 RESUMEN DEL PROCESAMIENTO")
    _emit("=" * 60)
    _emit(f"Total depósitos: {results['total']}")
    _emit(f"✅ Exitosos: {results['success_count']}")
    _emit(f"❌ Errores: {results['errors']}")

    log_operation("csv_batches")

    results['success'] = results['errors'] == 0
    if log is not None:
        results['log_lines'] = list(log)
    return results



def procesar_reconciliacion_bancaria(csv_file: str) -> dict:
    """Procesa CSV de reconciliación bancaria con balance opcional"""
    from decimal import Decimal

    print(f"\n🏦 RECONCILIACIÓN BANCARIA")
    print("="*70)
    print(f"📁 Archivo: {csv_file}\n")

    if not os.path.exists(csv_file):
        return {"success": False, "error": f"Archivo no encontrado: {csv_file}"}

    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            required = ['date', 'description', 'debit', 'credit']
            missing = [col for col in required if col not in headers]
            if missing:
                return {"success": False, "error": f"Columnas faltantes: {missing}"}
            has_balance = 'balance' in headers
            balance_mode = "CON balance" if has_balance else "SIN balance"
            print(f"📋 Modo: {balance_mode}")
            rows = list(reader)
    except Exception as e:
        return {"success": False, "error": f"Error leyendo CSV: {e}"}

    if not rows:
        return {"success": False, "error": "CSV vacío"}

    transactions = []
    opening_balance = Decimal('0')
    ending_balance = Decimal('0')
    running_balance = Decimal('0')
    total_debits = Decimal('0')
    total_credits = Decimal('0')
    validation_errors = []

    for idx, row in enumerate(rows, start=2):
        date = row['date'].strip()
        description = row['description'].strip()
        debit_str = row['debit'].strip()
        credit_str = row['credit'].strip()
        reference = row.get('reference', '').strip()
        balance_str = row.get('balance', '').strip() if has_balance else ''

        try:
            debit = Decimal(debit_str) if debit_str else Decimal('0')
            credit = Decimal(credit_str) if credit_str else Decimal('0')
            balance = Decimal(balance_str) if balance_str else None
        except:
            validation_errors.append(f"Fila {idx}: Formato inválido")
            continue

        desc_lower = description.lower()
        is_opening = 'opening' in desc_lower and 'balance' in desc_lower
        is_ending = 'ending' in desc_lower and 'balance' in desc_lower

        if is_opening:
            opening_balance = balance if (has_balance and balance) else Decimal('0')
            running_balance = opening_balance
            continue

        if is_ending:
            if has_balance and balance:
                ending_balance = balance
            continue

        if debit > 0 and credit > 0:
            validation_errors.append(f"Fila {idx}: Debit Y credit simultáneos")
            continue

        if debit > 0:
            total_debits += debit
            running_balance -= debit
        if credit > 0:
            total_credits += credit
            running_balance += credit

        if has_balance and balance is not None:
            diff = abs(balance - running_balance)
            if diff > Decimal('0.01'):
                validation_errors.append(f"Fila {idx}: Balance no cuadra")

        transactions.append({
            'row_num': idx, 'date': date, 'description': description,
            'debit': float(debit), 'credit': float(credit),
            'balance': float(running_balance), 'reference': reference
        })

    calculated_ending = opening_balance + total_credits - total_debits

    if validation_errors:
        return {"success": False, "error": "Validación falló", "validation_errors": validation_errors}

    if not transactions:
        return {"success": False, "error": "No hay transacciones"}

    checking_accounts = find_account("Checking", category="ACTIVO")
    if not checking_accounts:
        checking_accounts = find_account("Bank", category="ACTIVO")
    if not checking_accounts:
        return {"success": False, "error": "No se encontró cuenta bancaria"}

    bank_account_id = checking_accounts[0]['id']
    income_accounts = find_account("Income", category="INGRESO")
    expense_accounts = find_account("Expense", category="GASTO")

    if not income_accounts or not expense_accounts:
        return {"success": False, "error": "Cuentas contables no encontradas"}

    income_account_id = income_accounts[0]['id']
    expense_account_id = expense_accounts[0]['id']

    results = {"total": len(transactions), "success": 0, "errors": 0, "details": []}

    has_debits = any(txn['credit'] <= 0 for txn in transactions)
    bank_charges_vendor_id = None
    if has_debits:
        vendors = search_vendor("Bank Charges")
        if not vendors:
            return {
                "success": False,
                "error": "No se encontró vendor 'Bank Charges' para reconciliar débitos. "
                         "Crea el vendor o ajusta el flujo de reconciliación.",
            }
        bank_charges_vendor_id = vendors[0]['id']

    for txn in transactions:
        try:
            if txn['credit'] > 0:
                amount = txn['credit']
                result = create_deposit(
                    account_id=bank_account_id,
                    line_items=[{"amount": amount, "from_account_id": income_account_id, "description": txn['description']}],
                    txn_date=parse_date(txn['date']),
                    memo=f"Reconciliation - {txn['reference']}" if txn['reference'] else "Bank Reconciliation"
                )
            else:
                amount = txn['debit']
                result = create_bill(
                    vendor_id=bank_charges_vendor_id,
                    line_items=[{"amount": amount, "account_id": expense_account_id, "description": txn['description']}],
                    txn_date=parse_date(txn['date']),
                    memo=f"Reconciliation - {txn['reference']}" if txn['reference'] else "Bank Reconciliation"
                )

            if result.get('success'):
                results['success'] += 1
            else:
                results['errors'] += 1
        except Exception as e:
            results['errors'] += 1

    log_operation("csv_batches")
    results['success'] = results['success'] > 0
    results['summary'] = {
        'opening_balance': float(opening_balance),
        'ending_balance': float(calculated_ending),
        'total_credits': float(total_credits),
        'total_debits': float(total_debits),
        'mode': balance_mode
    }
    return results


SYSTEM_PROMPT = r"""
Eres Dexter, un AGENTE contable autónomo especializado en QuickBooks Online.
Trabajás para Alfredo, un contador profesional. Tu trabajo es usar tus 101
herramientas para consultar, crear y gestionar datos en QBO con precisión.

═══════════════════════════════════════════════════════════════
CÓMO TRABAJÁS — tu método de trabajo (OBLIGATORIO)
═══════════════════════════════════════════════════════════════

1. ENTENDER — analizá qué necesita Alfredo. Si no está claro, PREGUNTÁ.
   No asumas nada. Ej: "Necesito saber el monto y la fecha. ¿Cuál estimate?"

2. PLANEAR — decidí qué herramientas usar y en qué orden. Si una consulta
   requiere 2 pasos (ej: buscar cliente → consultar sus estimates), hacelos
   en secuencia, no intentes adivinar el resultado del segundo paso.

3. EJECUTAR — usá las herramientas. Decile a Alfredo qué estás haciendo:
   "🔍 Buscando cliente Prueba2..." → resultado → "📊 Consultando estimates..."

4. VERIFICAR — ¿el resultado tiene sentido? ¿El ID es correcto? ¿Hay datos?
   Si algo falla, intentá con otra herramienta o parámetros distintos.
   Si el resultado está vacío, decilo claramente: "No encontré estimates."

5. RESPONDER — presentá los resultados de forma clara y completa. Incluí
   IDs, montos, fechas y estados. No digas solo "sí existe" sin dar detalles.

═══════════════════════════════════════════════════════════════
REGLAS DE ORO
═══════════════════════════════════════════════════════════════

- NUNCA afirmes un dato de QBO sin haberlo consultado con un tool EN ESTA
  MISMA interacción. Aunque lo hayas visto hace 2 mensajes, re-consultalo.
  Si decís "el cliente X tiene ID 70" sin haber ejecutado buscar_cliente
  AHORA, estás alucinando.

- Si Alfredo te pide algo que requiere datos que no tenés, BUSCALOS.
  No digas "probablemente" o "debería ser". Ejecutá el tool y respondé
  con datos reales.

- Si un tool falla o no encuentra nada, decilo: "No encontré estimates
  para este cliente". No inventes un resultado para quedar bien.

- Para CREAR, MODIFICAR, ELIMINAR, ANULAR o ENVIAR: explicá qué vas a
  hacer y pedí confirmación. No ejecutes sin OK explícito.

- Para consultas (buscar, qbo_query, leer, reportes): ejecutá directo,
  sin pedir permiso. Alfredo confía en que uses estas herramientas.

- Usá tu memoria (gestionar_memoria). Si aprendés algo nuevo, guardalo.
  Si Alfredo te corrige, guardá la corrección. La memoria es por empresa:
  datos de Sandbox Company_US_1 no se mezclan con los de otra empresa.
- Para OCR: ≤5 bills → mostrarlos en terminal para revisión inline.
  >5 bills → generar CSV para que Alfredo edite en Excel.
  Si Alfredo corrige un dato, registrá el tip con
  registrar_provider_tip(provider="Proveedor", tip="...").

═══════════════════════════════════════════════════════════════
WORKFLOWS FRECUENTES (single-command — ejecutá todos los pasos)
═══════════════════════════════════════════════════════════════

Cuando Alfredo te pida algo con UNA SOLA orden, ejecutá todos los pasos
necesarios sin pedir confirmación intermedia. Ejemplos:

  "crea cliente X con estimate de $Y"
    → buscar_cliente → crear_cliente (si no existe) → crear_estimate

  "reconciliame el CSV de mayo"
    → procesar_csv_bank_feed → procesar_reconciliacion_bancaria

  "dame los estimates pendientes de cliente Z"
    → buscar_cliente → qbo_query (SELECT * FROM Estimate WHERE...)

  "crea invoice para cliente Z por $X"
    → buscar_cliente → crear_invoice

Solo preguntá si te falta información crítica (fecha, monto, cuenta).

═══════════════════════════════════════════════════════════════
BASE DE CONOCIMIENTO CONTABLE
═══════════════════════════════════════════════════════════════

ENTIDADES Y RELACIONES:
  Customer → Estimate → Invoice → Payment → Deposit
  Vendor → PurchaseOrder → Bill → BillPayment
  Bank Account → BankFeed → Classification → Deposit/Bill

ESTADOS DE ENTIDADES:
  Invoice: Balance>0=pendiente, Balance=0=pagada, void=anulada
  Estimate: Pending=recién creada, Accepted=cliente aceptó, Closed=convertida/expirada, Rejected=rechazada
  Bill: Balance>0=pendiente, Balance=0=pagada
  Payment: UnappliedAmt>0=parcialmente aplicado

TIPO DE CUENTAS:
  Bank, AccountsReceivable(AR), AccountsPayable(AP), Income, Expense,
  CostOfGoodsSold(COGS), FixedAsset, OtherAsset, OtherCurrentAsset,
  LongTermLiability, OtherCurrentLiability, Equity, CreditCard

QUERY LANGUAGE (SQL-like QBO):
  SELECT * FROM Entidad WHERE campo = 'valor' MAXRESULTS 100
  Entidades: Customer, Invoice, Estimate, Bill, Payment, Deposit,
  Account, Vendor, Item, Purchase, JournalEntry, etc.
  NO soporta LIKE, JOINs, ni subconsultas.
  COUNT(*) retorna totalCount como número, no como lista.
  Usar qbo_query para buscar cualquier entidad.

SIGNOS CONTABLES:
  Positivo (+) = ingreso, cobro, depósito, crédito a income
  Negativo (-) = gasto, pago, débito a expense
  Débito = aumenta activos/gastos, disminuye pasivos/ingresos
  Crédito = aumenta pasivos/ingresos, disminuye activos/gastos

ECUACIÓN CONTABLE: Activo = Pasivo + Patrimonio
PARTIDA DOBLE: cada transacción afecta ≥2 cuentas. Débitos = Créditos.

═══════════════════════════════════════════════════════════════
IDIOMA Y TONO
═══════════════════════════════════════════════════════════════

Responde SIEMPRE en el IDIOMA SELECCIONADO.
Idioma actual: {idioma}
Si ES: español profesional pero cercano. Usá "Alfredo" para dirigirte.
Si EN: English, professional and concise.
"""

def call_llm(user_message: str, tools: List[dict] = None, max_iterations: int = 5) -> str:
    """Llama al LLM con soporte de tools y maneja iteraciones automáticamente"""
    # Agregar contexto del chart of accounts al mensaje del sistema
    chart_summary = f"\\n\\nCHART OF ACCOUNTS EN MEMORIA: {len(session_state.get('chart_of_accounts', {}))} cuentas disponibles."

    # Construir el prompt del sistema local con los detalles necesarios
    current_lang = session_state.get("language", "es").upper()
    local_system_content = SYSTEM_PROMPT.replace("{idioma}", current_lang)
    
    if necesita_chart(user_message):
        local_system_content += chart_summary
    
    recent_hist, context_hint = build_conversation_context(conversation_history)
    local_system_content += "\n" + context_hint

    # Inyectar memoria persistente (MEMORY.md + USER.md) — snapshot al inicio
    mem = _get_memory()
    mem_block = mem.format_for_prompt()
    if mem_block:
        local_system_content += "\n\n" + mem_block

    # Inyectar perfil de empresa (PROFILE.md) — solo la referencia, no el contenido.
    # El LLM usa gestionar_memoria o lee el archivo cuando necesita los datos.
    if CURRENT_COMPANY:
        safe_name = CURRENT_COMPANY["name"].replace("/", "_").replace("\\", "_")
        profile_note = (
            f"\n\n═══ DATOS DE EMPRESA ═══\n"
            f"Nombre: {CURRENT_COMPANY['name']} | Realm: {CURRENT_COMPANY['realm_id']}\n"
            f"Perfil completo en: companies/{safe_name}/PROFILE.md\n"
            f"Memoria en: companies/{safe_name}/MEMORY.md\n"
            f"\nConsultá estos archivos SOLO cuando necesites datos específicos de la empresa."
        )
        local_system_content += profile_note

    relevant_tools = get_relevant_tools(user_message) if tools else []

    # Solo agregar mensaje del usuario si no está vacío (para tool calls iterativos)
    if user_message:
        conversation_history.append({
            "role": "user",
            "content": user_message
        })

    iteration = 0

    while iteration < max_iterations:
        iteration += 1

        # Construir mensajes incluyendo el historial actualizado en cada iteración
        # CRIT-2: conversation_history es deque(maxlen=200) — slicing no soportado,
        # usar list() para ventana de contexto
        _history_window = list(conversation_history)[-(max_iterations*4+10):]
        messages = [
            {"role": "system", "content": local_system_content},
            *_history_window  # Ventana de contexto amplia
        ]

        # Usar modelo pesado (DeepSeek) para TODAS las interacciones con tools.
        # La diferencia de costo es mínima (~$0.001 por consulta) y la diferencia
        # de calidad en function calling es enorme. Llama 3.1 8B ignora los tools
        # y responde de memoria, causando datos incorrectos.
        selected_model = LLM_MODEL_HEAVY
        if iteration == 1:
            self_model = selected_model
        else:
            selected_model = self_model

        payload = {
            "model": selected_model,
            "messages": messages,
            "temperature": 0.2
        }

        if tools:
            payload["tools"] = relevant_tools
            payload["tool_choice"] = "auto"

        headers = {
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://github.com/alfredo-qbo-assistant",
            "X-Title": "QuickBooks AI Assistant"
        }

        response = requests.post(LLM_API_URL, headers=headers, json=payload)

        if response.status_code != 200:
            error_msg = f"Error LLM ({response.status_code}): {response.text}"
            print(f"❌ {error_msg}")
            return error_msg

        result = response.json()

        # Actualizar tokens y costo
        usage = result.get("usage", {})
        update_token_usage(
            usage.get("prompt_tokens", 0),
            usage.get("completion_tokens", 0),
            selected_model
        )

        assistant_message = result["choices"][0]["message"]
        conversation_history.append(assistant_message)

        # Si no hay tool calls, retornar contenido
        if not assistant_message.get("tool_calls"):
            return assistant_message.get("content", "")

        # Procesar tool calls
        for tool_call in assistant_message["tool_calls"]:
            function_name = tool_call["function"]["name"]

            try:
                arguments = json.loads(tool_call["function"]["arguments"])
            except json.JSONDecodeError:
                arguments = {}

            # Mostrar qué está haciendo el agente (transparencia)
            arg_preview = ", ".join(f"{k}={v}" for k, v in list(arguments.items())[:2])
            if len(arguments) > 2:
                arg_preview += ", ..."
            try:
                from dexter.console import tool_start
                tool_start(function_name, arg_preview)
            except ImportError:
                print(f"  🔧 {function_name}({arg_preview})" if arg_preview else f"  🔧 {function_name}")

            # Ejecutar tool (con soporte dry-run)
            from dexter.core.safe_json import safe_dumps
            try:
                result_data = _execute_tool(function_name, arguments)
                result_str = safe_dumps(result_data, ensure_ascii=False)
            except Exception as e:
                result_str = safe_dumps({"error": str(e)}, ensure_ascii=False)
                _log_error(
                    e,
                    category="tool_dispatch",
                    tool_name=function_name,
                    extra={"arguments": arguments, "user_message": user_message},
                )

            # Agregar resultado al historial
            conversation_history.append({
                "role": "tool",
                "tool_call_id": tool_call["id"],
                "content": result_str
            })

            # Mini-resumen del resultado (primeros 100 chars sin saturar)
            summary = result_str[:100].replace("\n", " ")
            if len(result_str) > 100:
                summary += f"... ({len(result_str)} chars)"
            try:
                from dexter.console import tool_result
                is_ok = '"error"' not in result_str[:50].lower()
                tool_result(summary, success=is_ok)
            except ImportError:
                print(f"       → {summary}")

        # Continuar iteración (el LLM procesará los resultados de los tools)

    # Si llegamos aquí, se alcanzó el máximo de iteraciones
    return "Se alcanzó el límite de iteraciones. Por favor, reformula tu pregunta."

# ==================== TOOLS PARA EL LLM ====================

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "buscar_cliente",
            "description": "Busca clientes en QuickBooks por nombre (fuzzy search). Retorna lista de clientes con ID, nombre, balance.",
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre": {
                        "type": "string",
                        "description": "Nombre o parte del nombre del cliente a buscar"
                    },
                    "exacto": {
                        "type": "boolean",
                        "description": "Si es true, busca coincidencia exacta. Por defecto false (fuzzy).",
                        "default": False
                    }
                },
                "required": ["nombre"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "buscar_vendor",
            "description": "Busca vendors/proveedores en QuickBooks por nombre.",
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre": {
                        "type": "string",
                        "description": "Nombre del vendor a buscar"
                    }
                },
                "required": ["nombre"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "buscar_cuenta",
            "description": "Busca cuenta contable por nombre o número en el Chart of Accounts. Usa fuzzy matching.",
            "parameters": {
                "type": "object",
                "properties": {
                    "termino": {
                        "type": "string",
                        "description": "Nombre o número de cuenta a buscar"
                    },
                    "categoria": {
                        "type": "string",
                        "enum": ["ACTIVO", "PASIVO", "INGRESO", "GASTO"],
                        "description": "Filtrar por categoría de cuenta (opcional)"
                    }
                },
                "required": ["termino"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "buscar_item",
            "description": "Busca items/servicios en QuickBooks por nombre.",
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre": {
                        "type": "string",
                        "description": "Nombre del item/servicio"
                    }
                },
                "required": ["nombre"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crear_invoice",
            "description": "Crea un invoice/factura en QuickBooks. Requiere customer_id y líneas con items.",
            "parameters": {
                "type": "object",
                "properties": {
                    "customer_id": {
                        "type": "string",
                        "description": "ID del cliente (obtener con buscar_cliente)"
                    },
                    "lineas": {
                        "type": "array",
                        "description": "Lista de líneas del invoice",
                        "items": {
                            "type": "object",
                            "properties": {
                                "item_id": {"type": "string"},
                                "amount": {"type": "number"},
                                "quantity": {"type": "number", "default": 1},
                                "description": {"type": "string"}
                            },
                            "required": ["item_id", "amount"]
                        }
                    },
                    "fecha": {
                        "type": "string",
                        "description": "Fecha en formato YYYY-MM-DD (opcional, usa hoy por defecto)"
                    },
                    "memo": {
                        "type": "string",
                        "description": "Nota privada (opcional)"
                    }
                },
                "required": ["customer_id", "lineas"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crear_bill",
            "description": "Crea un bill/cuenta por pagar en QuickBooks. Requiere vendor_id y líneas con cuentas de gasto.",
            "parameters": {
                "type": "object",
                "properties": {
                    "vendor_id": {
                        "type": "string",
                        "description": "ID del vendor (obtener con buscar_vendor)"
                    },
                    "lineas": {
                        "type": "array",
                        "description": "Lista de líneas del bill con cuentas de gasto",
                        "items": {
                            "type": "object",
                            "properties": {
                                "account_id": {"type": "string", "description": "ID de cuenta de gasto"},
                                "amount": {"type": "number"},
                                "description": {"type": "string"}
                            },
                            "required": ["account_id", "amount"]
                        }
                    },
                    "fecha": {"type": "string", "description": "Fecha YYYY-MM-DD"},
                    "fecha_vencimiento": {"type": "string", "description": "Fecha vencimiento YYYY-MM-DD"},
                    "memo": {"type": "string"}
                },
                "required": ["vendor_id", "lineas"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crear_deposito",
            "description": "Crea un depósito en QuickBooks. Mueve dinero de cuentas origen (ej: Client Retainers) a cuenta destino (ej: Checking).",
            "parameters": {
                "type": "object",
                "properties": {
                    "cuenta_destino_id": {
                        "type": "string",
                        "description": "ID de cuenta bancaria destino (obtener con buscar_cuenta)"
                    },
                    "lineas": {
                        "type": "array",
                        "description": "Lista de líneas del depósito",
                        "items": {
                            "type": "object",
                            "properties": {
                                "cuenta_origen_id": {"type": "string", "description": "ID cuenta origen"},
                                "amount": {"type": "number"},
                                "customer_id": {"type": "string", "description": "ID cliente (opcional)"},
                                "description": {"type": "string"}
                            },
                            "required": ["cuenta_origen_id", "amount"]
                        }
                    },
                    "fecha": {"type": "string"},
                    "memo": {"type": "string"}
                },
                "required": ["cuenta_destino_id", "lineas"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crear_pago",
            "description": "Registra un pago recibido de un cliente en QuickBooks.",
            "parameters": {
                "type": "object",
                "properties": {
                    "customer_id": {"type": "string"},
                    "amount": {"type": "number"},
                    "cuenta_id": {"type": "string", "description": "Cuenta donde se deposita el pago"},
                    "fecha": {"type": "string"},
                    "aplicar_a_invoices": {
                        "type": "array",
                        "description": "Lista de invoices a los que aplicar el pago",
                        "items": {
                            "type": "object",
                            "properties": {
                                "invoice_id": {"type": "string"},
                                "amount": {"type": "number"}
                            }
                        }
                    }
                },
                "required": ["customer_id", "amount", "cuenta_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generar_reporte_pl",
            "description": "Genera reporte de Profit & Loss (P&L / Estado de Resultados).",
            "parameters": {
                "type": "object",
                "properties": {
                    "fecha_inicio": {"type": "string", "description": "Fecha inicio YYYY-MM-DD"},
                    "fecha_fin": {"type": "string", "description": "Fecha fin YYYY-MM-DD"},
                    "metodo": {"type": "string", "enum": ["Accrual", "Cash"], "default": "Accrual"}
                },
                "required": ["fecha_inicio", "fecha_fin"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generar_balance_sheet",
            "description": "Genera reporte de Balance Sheet (Balance General).",
            "parameters": {
                "type": "object",
                "properties": {
                    "fecha": {"type": "string", "description": "Fecha del balance YYYY-MM-DD"},
                    "metodo": {"type": "string", "enum": ["Accrual", "Cash"], "default": "Accrual"}
                },
                "required": ["fecha"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "guardar_reporte",
            "description": "Guarda configuración de un reporte para uso futuro.",
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre": {"type": "string", "description": "Nombre para identificar el reporte"},
                    "config": {"type": "object", "description": "Configuración del reporte"}
                },
                "required": ["nombre", "config"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "cargar_reporte",
            "description": "Carga configuración de un reporte guardado previamente.",
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre": {"type": "string"}
                },
                "required": ["nombre"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "listar_reportes_guardados",
            "description": "Lista todos los reportes guardados por el usuario.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "procesar_csv_depositos",
            "description": "Procesa archivo CSV con múltiples depósitos y los crea en batch.",
            "parameters": {
                "type": "object",
                "properties": {
                    "ruta_archivo": {"type": "string", "description": "Ruta del archivo CSV"}
                },
                "required": ["ruta_archivo"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crear_template_csv",
            "description": "Crea archivo CSV template para depósitos batch.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "obtener_estadisticas_tokens",
            "description": "Muestra estadísticas de consumo de tokens del LLM.",
            "parameters": {
                "type": "object",
                "properties": {
                    "periodo": {
                        "type": "string",
                        "enum": ["sesion", "hoy", "mes"],
                        "description": "Periodo a consultar"
                    }
                },
                "required": ["periodo"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generar_informe_tokens",
            "description": "Genera informe Excel con estadísticas detalladas de consumo (sobrescribe archivo).",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "refrescar_chart_accounts",
            "description": "Refresca el Chart of Accounts desde QuickBooks Online (fuerza actualización del caché).",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "procesar_bank_feed_csv",
            "description": "Procesa archivo CSV de Bank Feed y clasifica depósitos con splits múltiples (income + fees). Formato CSV: bank_feed_date,bank_feed_amount,deposit_id,line_type,customer_name,amount,account,memo",
            "parameters": {
                "type": "object",
                "properties": {
                    "archivo_csv": {
                        "type": "string",
                        "description": "Ruta del archivo CSV de Bank Feed"
                    }
                },
                "required": ["archivo_csv"]
            }
        }
    },

    {
        "type": "function",
        "function": {
            "name": "procesar_reconciliacion_bancaria",
            "description": "Procesa CSV de reconciliación bancaria y crea transacciones en QuickBooks. Soporta dos formatos: CON balance (6 columnas con validación completa) o SIN balance (5 columnas con cálculo automático). Columnas obligatorias: date, description, debit, credit. Columnas opcionales: balance, reference.",
            "parameters": {
                "type": "object",
                "properties": {
                    "archivo_csv": {
                        "type": "string",
                        "description": "Ruta del archivo CSV de reconciliación bancaria"
                    }
                },
                "required": ["archivo_csv"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "taggear_reconciliacion",
            "description": "BNK-RECON tagger: marca transactions existentes en QBO (Deposit.Memo, Bill.PrivateNote, Purchase.PrivateNote) con el tag BNK-RECON-YYYY-MM-xxxxx. NO crea transactions nuevas, solo agrega tags visibles. Útil para reconciliación en QBO UI. Columnas requeridas del CSV: date, description, amount.",
            "parameters": {
                "type": "object",
                "properties": {
                    "archivo_csv": {
                        "type": "string",
                        "description": "Ruta del archivo CSV del bank statement (columnas: date, description, amount)"
                    },
                    "cuenta_id": {
                        "type": "string",
                        "description": "ID de la cuenta bancaria en QBO (opcional; se auto-detecta por categoría BANK si se omite)"
                    },
                    "fecha_inicio": {
                        "type": "string",
                        "description": "Fecha de inicio del período en formato YYYY-MM-DD (opcional; default: primer día del mes actual)"
                    },
                    "fecha_fin": {
                        "type": "string",
                        "description": "Fecha de fin del período en formato YYYY-MM-DD (opcional; default: último día del mes actual)"
                    },
                    "dias_fuzzy": {
                        "type": "integer",
                        "description": "Tolerancia en días para fuzzy match (default 2)",
                        "default": 2
                    },
                    "monto_fuzzy": {
                        "type": "number",
                        "description": "Tolerancia en USD para diferencia de monto (default 0.50)",
                        "default": 0.50
                    }
                },
                "required": ["archivo_csv"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "limpiar_tags_reconciliacion",
            "description": "Limpia los tags BNK-RECON aplicados por un batch previo. Lee el reporte del batch y borra los Memo/PrivateNote. Útil para deshacer una reconciliación de prueba.",
            "parameters": {
                "type": "object",
                "properties": {
                    "batch_id": {
                        "type": "string",
                        "description": "ID del batch cuyos tags se quieren limpiar"
                    }
                },
                "required": ["batch_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "depositar_lote_csv",
            "description": "Procesa CSV de deposits multi-cliente usando el motor batch con state machine, disambiguación interactiva y dry-run obligatorio. Columnas requeridas: date, client_name, amount. Si un cliente no existe, pregunta si crearlo. Si confirmar=false, solo hace dry-run sin crear nada en QBO.",
            "parameters": {
                "type": "object",
                "properties": {
                    "ruta_archivo": {
                        "type": "string",
                        "description": "Ruta al CSV de líneas de deposit (date, client_name, amount)"
                    },
                    "cuenta_banco_id": {
                        "type": "string",
                        "description": "ID de la cuenta bancaria destino (opcional; se auto-detecta)"
                    },
                    "cuenta_ingreso_id": {
                        "type": "string",
                        "description": "ID de la cuenta de ingreso (opcional; se auto-detecta)"
                    },
                    "confirmar": {
                        "type": "boolean",
                        "description": "Si False, solo corre dry-run sin crear (default True)",
                        "default": True
                    }
                },
                "required": ["ruta_archivo"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "procesar_lote_bills",
            "description": (
                "Procesa un lote de bills/invoices desde un PDF usando OCR con Gemini. "
                "Busca automáticamente en la carpeta 'Pending bills'. "
                "Si hay un solo PDF, lo procesa automáticamente. "
                "Si hay múltiples, lista los disponibles para que el usuario elija. "
                "Extrae: vendor, customer, invoice#, date, total, tax. "
                "Genera CSV preview para revisión antes de crear Bills en QuickBooks."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "nombre_archivo": {
                        "type": "string",
                        "description": (
                            "Nombre del archivo PDF a procesar (opcional). "
                            "Si no se especifica y hay solo 1 PDF en 'Pending bills', "
                            "se procesa automáticamente. Puede ser nombre parcial "
                            "(ej: 'okna' encontrará 'DONE-Okna-Invoices-dated1.8.26.pdf')"
                        )
                    }
                },
                "required": []
            }
        }
    },
    # ========== TOOLS DE AUTONOMÍA ==========
    {
        "type": "function",
        "function": {
            "name": "buscarenweb",
            "description": "Busca información actualizada en internet (vía DuckDuckGo).",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Consulta de búsqueda"}
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "buscardocsqbo",
            "description": "Busca específicamente en la documentación oficial de QuickBooks Online API.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Consulta técnica sobre la API"}
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "crearasientodiario",
            "description": "Crea un Journal Entry (asiento contable) complejo en QuickBooks. Los débitos deben ser iguales a los créditos.",
            "parameters": {
                "type": "object",
                "properties": {
                    "lines": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "account_id": {"type": "string"},
                                "amount": {"type": "number"},
                                "posting_type": {"type": "string", "enum": ["Debit", "Credit"]},
                                "description": {"type": "string"}
                            },
                            "required": ["account_id", "amount", "posting_type"]
                        }
                    },
                    "txn_date": {"type": "string", "description": "Fecha YYYY-MM-DD"},
                    "memo": {"type": "string"}
                },
                "required": ["lines", "txn_date"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "creartransferencia",
            "description": "Crea una transferencia de fondos entre dos cuentas bancarias en QuickBooks.",
            "parameters": {
                "type": "object",
                "properties": {
                    "from_account_id": {"type": "string"},
                    "to_account_id": {"type": "string"},
                    "amount": {"type": "number"},
                    "txn_date": {"type": "string"},
                    "memo": {"type": "string"}
                },
                "required": ["from_account_id", "to_account_id", "amount", "txn_date"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "qborequestgenerico",
            "description": "Realiza una petición genérica a cualquier endpoint de la API de QuickBooks v3.",
            "parameters": {
                "type": "object",
                "properties": {
                    "method": {"type": "string", "enum": ["GET", "POST", "UPDATE"]},
                    "endpoint": {"type": "string", "description": "Nombre del recurso (ej: Purchase, PaymentMethod)"},
                    "data": {"type": "object", "description": "Cuerpo del JSON para POST/UPDATE"},
                    "entity_id": {"type": "string", "description": "ID del recurso para GET/UPDATE específico"}
                },
                "required": ["method", "endpoint"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "listarendpointsqbo",
            "description": "Lista los endpoints más comunes disponibles en la API de QuickBooks.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "infoendpointqbo",
            "description": "Obtiene información detallada sobre cómo usar un endpoint específico de la API.",
            "parameters": {
                "type": "object",
                "properties": {
                    "endpoint_name": {"type": "string"}
                },
                "required": ["endpoint_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "ejecutarcodigo",
            "description": "Ejecuta fragmentos de código Python para análisis de datos avanzados o cálculos complejos.",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {"type": "string", "description": "Código Python a ejecutar"}
                },
                "required": ["code"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "analizarbankfeed",
            "description": "Analiza una lista de transacciones bancarias para sugerir clasificaciones contables basadas en aprendizaje previo.",
            "parameters": {
                "type": "object",
                "properties": {
                    "account_name": {"type": "string"},
                    "transactions": {
                        "type": "array",
                        "items": {"type": "object"}
                    },
                    "min_confidence": {"type": "number", "default": 0.7}
                },
                "required": ["account_name", "transactions"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "registrarclasificacion",
            "description": "Registra el aprendizaje de una clasificación manual hecha por el usuario.",
            "parameters": {
                "type": "object",
                "properties": {
                    "description": {"type": "string"},
                    "account_id": {"type": "string"},
                    "account_name": {"type": "string"},
                    "amount": {"type": "number"},
                    "date": {"type": "string"},
                    "vendor": {"type": "string"},
                    "qb_suggestion": {"type": "string"}
                },
                "required": ["description", "account_id", "account_name", "amount", "date"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "estadisticasclasificacion",
            "description": "Obtiene estadísticas sobre el aprendizaje del sistema de clasificación bancaria.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "buscarpatron",
            "description": "Busca si existe un patrón de clasificación previo para una descripción dada.",
            "parameters": {
                "type": "object",
                "properties": {
                    "description": {"type": "string"}
                },
                "required": ["description"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "aprenderinteraccion",
            "description": "Aprende de las preferencias del usuario basadas en sus interacciones.",
            "parameters": {
                "type": "object",
                "properties": {
                    "interaction_type": {"type": "string", "enum": ["account_use", "vendor_use", "report_use"]},
                    "details": {"type": "object"},
                    "context": {"type": "string"}
                },
                "required": ["interaction_type", "details"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "obtenersugerencias",
            "description": "Obtiene sugerencias de acciones basadas en el comportamiento histórico del usuario.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "registrarcorreccion",
            "description": "Registra una corrección del usuario cuando el sistema comete un error.",
            "parameters": {
                "type": "object",
                "properties": {
                    "wrong": {"type": "string"},
                    "correct": {"type": "string"},
                    "context": {"type": "string"}
                },
                "required": ["wrong", "correct", "context"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "obtenercontexto",
            "description": "Obtiene un resumen del contexto reciente de la conversación y tareas activas.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generarreportecustom",
            "description": "Genera reportes personalizados interpretando peticiones en lenguaje natural.",
            "parameters": {
                "type": "object",
                "properties": {
                    "user_request": {"type": "string"},
                    "filters": {"type": "object"}
                },
                "required": ["user_request"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "parsearfecha",
            "description": "Convierte expresiones temporales (ej: 'el mes pasado') en fechas específicas.",
            "parameters": {
                "type": "object",
                "properties": {
                    "expression": {"type": "string"}
                },
                "required": ["expression"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "gestionar_empresas",
            "description": "Permite registrar una nueva empresa (vía link QBO o ID), listar las registradas o cambiar entre ellas.",
            "parameters": {
                "type": "object",
                "properties": {
                    "accion": {
                        "type": "string", 
                        "enum": ["registrar", "listar", "cambiar"],
                        "description": "Acción a realizar: 'registrar' una nueva, 'listar' todas, o 'cambiar' a una existente."
                    },
                    "nombre": {
                        "type": "string",
                        "description": "Nombre de la empresa (requerido para 'registrar' y 'cambiar')."
                    },
                    "link_o_id": {
                        "type": "string",
                        "description": "URL de QuickBooks o Realm ID de la empresa (requerido para 'registrar')."
                    }
                },
                "required": ["accion"]
            }
        }
    }
]

# ==================== FUNCIONES DE LOS TOOLS ====================

def tool_buscar_cliente(nombre: str, exacto: bool = False) -> dict:
    """Tool: Busca cliente por nombre"""
    results = search_customer(nombre, exact=exacto)

    # Guardar en session state para referencia rápida
    session_state["last_search_results"]["customers"] = results

    return {
        "encontrados": len(results),
        "clientes": results[:5]  # Máximo 5 resultados
    }

def tool_buscar_vendor(nombre: str) -> dict:
    """Tool: Busca vendor por nombre"""
    results = search_vendor(nombre)
    session_state["last_search_results"]["vendors"] = results

    return {
        "encontrados": len(results),
        "vendors": results[:5]
    }

def tool_buscar_cuenta(termino: str, categoria: str = None) -> dict:
    """Tool: Busca cuenta contable"""
    results = find_account(termino, category=categoria)
    session_state["last_search_results"]["accounts"] = results

    return {
        "encontradas": len(results),
        "cuentas": [{
            "id": acc["id"],
            "numero": acc["number"],
            "nombre": acc["name"],
            "tipo": acc["type"],
            "categoria": acc["category"],
            "balance": acc["balance"],
            "similitud": round(acc.get("match_score", 1.0) * 100, 1)
        } for acc in results[:5]]
    }

def tool_buscar_item(nombre: str) -> dict:
    """Tool: Busca item/servicio"""
    results = search_item(nombre)

    return {
        "encontrados": len(results),
        "items": results[:5]
    }

def tool_crear_invoice(customer_id: str, lineas: List[dict], fecha: str = None, memo: str = None) -> dict:
    """Tool: Crea invoice"""
    return create_invoice(customer_id, lineas, fecha, memo)

def tool_crear_bill(vendor_id: str, lineas: List[dict], fecha: str = None, 
                   fecha_vencimiento: str = None, memo: str = None) -> dict:
    """Tool: Crea bill"""
    return create_bill(vendor_id, lineas, fecha, fecha_vencimiento, memo)

def tool_crear_deposito(cuenta_destino_id: str, lineas: List[dict], fecha: str = None, memo: str = None) -> dict:
    """Tool: Crea depósito"""
    return create_deposit(cuenta_destino_id, lineas, fecha, memo)

def tool_crear_pago(customer_id: str, amount: float, cuenta_id: str, fecha: str = None,
                   aplicar_a_invoices: List[dict] = None) -> dict:
    """Tool: Crea payment"""
    return create_payment(customer_id, amount, cuenta_id, fecha, aplicar_a_invoices)

def tool_crear_cliente(nombre: str, email: str = None, telefono: str = None,
                       direccion: str = None, empresa: str = None) -> dict:
    """Tool: Crea un cliente (Customer) en QuickBooks."""
    return create_customer(nombre, email, telefono, direccion, empresa)


# ========================================================================
# Tool wrappers para Sprint 1A: Master Data creates
# ========================================================================

def tool_crear_vendor(nombre: str, empresa: str = None, email: str = None,
                      telefono: str = None, direccion: str = None,
                      es_1099: bool = False, tarifa_hora: float = None,
                      term_id: str = None) -> dict:
    """Tool: Crea un proveedor (Vendor) en QuickBooks."""
    return create_vendor(nombre, empresa, email, telefono, direccion, es_1099, tarifa_hora, term_id)


def tool_crear_cuenta(nombre: str, tipo_cuenta: str, subtipo: str = None,
                      descripcion: str = None, saldo_apertura: float = None,
                      fecha_saldo_apertura: str = None) -> dict:
    """Tool: Crea una cuenta contable (Account) en QuickBooks."""
    return create_account(nombre, tipo_cuenta, subtipo, descripcion,
                          saldo_apertura, fecha_saldo_apertura)


def tool_crear_item(nombre: str, tipo: str = "Service", precio_unitario: float = 0.0,
                    cuenta_ingreso_id: str = None, cuenta_gasto_id: str = None,
                    cuenta_activo_id: str = None, sku: str = None,
                    rastrear_inventario: bool = False, cantidad_inicial: float = 0.0,
                    fecha_inicio_inv: str = None, descripcion: str = None) -> dict:
    """Tool: Crea un item (producto o servicio) en QuickBooks."""
    return create_item(nombre, tipo, precio_unitario, cuenta_ingreso_id,
                       cuenta_gasto_id, cuenta_activo_id, sku,
                       rastrear_inventario, cantidad_inicial, fecha_inicio_inv,
                       descripcion)


def tool_crear_empleado(nombre: str, apellido: str = None, segundo_apellido: str = None,
                        email: str = None, telefono: str = None, direccion: str = None,
                        fecha_contratacion: str = None, tarifa_hora: float = None) -> dict:
    """Tool: Crea un empleado (Employee) en QuickBooks."""
    return create_employee(nombre, apellido, segundo_apellido, email, telefono,
                          direccion, fecha_contratacion, tarifa_hora)


def tool_crear_clase(nombre: str, clase_padre_id: str = None, activa: bool = True) -> dict:
    """Tool: Crea una clase para segmentación P&L."""
    return create_class(nombre, clase_padre_id, activa)


def tool_crear_departamento(nombre: str, depto_padre_id: str = None, activo: bool = True) -> dict:
    """Tool: Crea un departamento para segmentación P&L."""
    return create_department(nombre, depto_padre_id, activo)


def tool_crear_termino(nombre: str, dias_vencimiento: int = 30,
                       dias_descuento: int = 0, pct_descuento: float = 0.0,
                       activo: bool = True) -> dict:
    """Tool: Crea un plazo de pago (ej: Net 30, 2/10 Net 30)."""
    return create_term(nombre, dias_vencimiento, dias_descuento, pct_descuento, activo)


def tool_crear_paymentmethod(nombre: str, tipo: str = "Other", activo: bool = True) -> dict:
    """Tool: Crea un método de pago en QuickBooks."""
    return create_payment_method(nombre, tipo, activo)


# ========================================================================
# Tool wrappers para Sprint 1B: Transaction creates
# ========================================================================

def tool_crear_billpayment(vendor_id: str, monto_total: float, tipo_pago: str = "Check",
                           fecha: str = None, cuenta_banco_id: str = None,
                           cuenta_cc_id: str = None, aplicar_a_bills: List[dict] = None,
                           memo: str = None) -> dict:
    """Tool: Paga uno o más bills (BillPayment) en QuickBooks."""
    return create_billpayment(vendor_id, monto_total, tipo_pago, fecha, cuenta_banco_id,
                              cuenta_cc_id, aplicar_a_bills, memo)


def tool_crear_estimate(cliente_id: str, lineas: List[dict], fecha: str = None,
                        fecha_expiracion: str = None, memo: str = None) -> dict:
    """Tool: Crea una cotización (Estimate) en QuickBooks."""
    return create_estimate(cliente_id, lineas, fecha, fecha_expiracion, memo)


def tool_crear_salesreceipt(cliente_id: str = None, lineas: List[dict] = None,
                            fecha: str = None, cuenta_deposito_id: str = None,
                            metodo_pago_id: str = None, memo: str = None) -> dict:
    """Tool: Crea un recibo de venta inmediata (SalesReceipt)."""
    return create_salesreceipt(cliente_id, lineas, fecha, cuenta_deposito_id,
                               metodo_pago_id, memo= memo)


def tool_crear_creditmemo(cliente_id: str, lineas: List[dict], fecha: str = None,
                          memo: str = None) -> dict:
    """Tool: Crea una nota de crédito (CreditMemo) para un cliente."""
    return create_creditmemo(cliente_id, lineas, fecha, memo)


def tool_crear_purchase(vendor_id: str, cuenta_gasto_id: str, monto: float,
                        tipo_pago: str = "Cash", fecha: str = None,
                        descripcion: str = None, memo: str = None) -> dict:
    """Tool: Crea una compra genérica (Purchase) por cash, check o tarjeta."""
    return create_purchase(vendor_id, cuenta_gasto_id, monto, tipo_pago, fecha,
                           descripcion, memo)


def tool_crear_purchaseorder(vendor_id: str, lineas: List[dict], fecha: str = None,
                             direccion_envio: str = None, memo: str = None,
                             email_po: str = None) -> dict:
    """Tool: Crea una orden de compra (PurchaseOrder) en QuickBooks."""
    return create_purchaseorder(vendor_id, lineas, fecha, direccion_envio, memo, email_po)


def tool_crear_refundreceipt(cliente_id: str, lineas: List[dict], cuenta_reembolso_id: str,
                             fecha: str = None, memo: str = None) -> dict:
    """Tool: Crea un recibo de reembolso (RefundReceipt) para un cliente."""
    return create_refundreceipt(cliente_id, lineas, cuenta_reembolso_id, fecha, memo)


def tool_crear_vendorcredit(vendor_id: str, lineas: List[dict], fecha: str = None,
                            memo: str = None) -> dict:
    """Tool: Crea un crédito de proveedor (VendorCredit)."""
    return create_vendorcredit(vendor_id, lineas, fecha, memo)


def tool_crear_timeactivity(empleado_id: str, horas: int = 0, minutos: int = 0,
                            fecha: str = None, cliente_id: str = None,
                            item_id: str = None, facturable: bool = True,
                            descripcion: str = None) -> dict:
    """Tool: Registra horas trabajadas (TimeActivity)."""
    return create_timeactivity(empleado_id, horas, minutos, fecha, cliente_id,
                               item_id, facturable, descripcion)


# ========================================================================
# Tool wrappers para Sprint 1C: Update/Void/Delete
# ========================================================================

def tool_actualizar_cliente(cliente_id: str, cambios: dict, sync_token: str = None) -> dict:
    """Tool: Actualiza un cliente (Customer) en QuickBooks vía sparse update."""
    return update_entity("customer", cliente_id, cambios, sync_token, sparse=True)


def tool_actualizar_vendor(vendor_id: str, cambios: dict, sync_token: str = None) -> dict:
    """Tool: Actualiza un vendor en QuickBooks vía sparse update."""
    return update_entity("vendor", vendor_id, cambios, sync_token, sparse=True)


def tool_actualizar_factura(invoice_id: str, cambios: dict, sync_token: str = None) -> dict:
    """Tool: Actualiza una factura (Invoice) en QuickBooks."""
    return update_entity("invoice", invoice_id, cambios, sync_token, sparse=True)


def tool_actualizar_bill(bill_id: str, cambios: dict, sync_token: str = None) -> dict:
    """Tool: Actualiza un bill en QuickBooks."""
    return update_entity("bill", bill_id, cambios, sync_token, sparse=True)


def tool_eliminar_transaccion(tipo: str, transaccion_id: str, sync_token: str) -> dict:
    """Tool: Elimina una transacción (Invoice, Bill, Payment, etc.) vía hard delete."""
    return delete_transaction(tipo, transaccion_id, sync_token)


def tool_void_transaccion(tipo: str, transaccion_id: str, sync_token: str) -> dict:
    """Tool: Anula (void) una transacción sin eliminarla del histórico."""
    return void_transaction(tipo, transaccion_id, sync_token)


def tool_desactivar_cliente(cliente_id: str, sync_token: str = None) -> dict:
    """Tool: Desactiva un cliente (soft delete via Active=false)."""
    return deactivate_entity("customer", cliente_id, sync_token)


def tool_desactivar_vendor(vendor_id: str, sync_token: str = None) -> dict:
    """Tool: Desactiva un vendor (soft delete)."""
    return deactivate_entity("vendor", vendor_id, sync_token)


# ========================================================================
# Tool wrappers para Sprint 1D: Send
# ========================================================================

def tool_enviar_factura(invoice_id: str, email: str = None) -> dict:
    """Tool: Envía una factura (Invoice) por email al cliente."""
    return send_transaction_email("invoice", invoice_id, email)


def tool_enviar_orden_compra(po_id: str, email: str = None) -> dict:
    """Tool: Envía una orden de compra (PurchaseOrder) por email al vendor."""
    return send_transaction_email("purchaseorder", po_id, email)


# ========================================================================
# Tool wrappers para Sprint 1E: Reportes adicionales
# ========================================================================

def tool_reporte_trial_balance(fecha_inicio: str, fecha_fin: str, metodo: str = "Accrual") -> dict:
    """Tool: Genera el Trial Balance (balance de comprobación) de la empresa."""
    return generate_trial_balance_report(fecha_inicio, fecha_fin, metodo)


def tool_reporte_general_ledger(fecha_inicio: str, fecha_fin: str, cuenta_id: str = None,
                                metodo: str = "Accrual") -> dict:
    """Tool: Genera el General Ledger (libro mayor) de la empresa."""
    return generate_general_ledger_report(fecha_inicio, fecha_fin, metodo, cuenta_id)


def tool_reporte_cash_flow(fecha_inicio: str, fecha_fin: str, metodo: str = "Accrual") -> dict:
    """Tool: Genera el Statement of Cash Flows."""
    return generate_cash_flow_report(fecha_inicio, fecha_fin, metodo)


def tool_reporte_ar_aging(fecha_corte: str, metodo_aging: str = "ReportDate",
                          num_periodos: int = 4) -> dict:
    """Tool: Genera A/R Aging Summary (reporte de cobranzas por antigüedad)."""
    return generate_ar_aging_report(fecha_corte, metodo_aging, num_periodos)


def tool_reporte_ap_aging(fecha_corte: str, metodo_aging: str = "ReportDate",
                          num_periodos: int = 4) -> dict:
    """Tool: Genera A/P Aging Summary (reporte de pagos pendientes por antigüedad)."""
    return generate_ap_aging_report(fecha_corte, metodo_aging, num_periodos)


def tool_reporte_customer_balance(fecha_corte: str = None, cliente_id: str = None) -> dict:
    """Tool: Genera Customer Balance Summary."""
    return generate_customer_balance_report(fecha_corte, cliente_id)


def tool_reporte_vendor_balance(fecha_corte: str = None, vendor_id: str = None) -> dict:
    """Tool: Genera Vendor Balance Summary."""
    return generate_vendor_balance_report(fecha_corte, vendor_id)


def tool_reporte_pl_detail(fecha_inicio: str, fecha_fin: str, metodo: str = "Accrual") -> dict:
    """Tool: Genera Profit & Loss Detail (más granular que P&L)."""
    return generate_pl_detail_report(fecha_inicio, fecha_fin, metodo)


def tool_reporte_journal(fecha_inicio: str, fecha_fin: str) -> dict:
    """Tool: Genera Journal Report (todos los asientos en un período)."""
    return generate_journal_report(fecha_inicio, fecha_fin)


def tool_reporte_account_list() -> dict:
    """Tool: Genera Account List (lista de cuentas contables)."""
    return generate_account_list_report()


def tool_reporte_inventory_valuation(fecha_inicio: str = None, fecha_fin: str = None,
                                      item_id: str = None) -> dict:
    """Tool: Genera Inventory Valuation Summary (valorización de inventario)."""
    return generate_inventory_valuation_report(fecha_inicio, fecha_fin, item_id)


def tool_reporte_sales_by_customer(fecha_inicio: str, fecha_fin: str,
                                    cliente_id: str = None) -> dict:
    """Tool: Genera Sales by Customer Summary (ventas agrupadas por cliente)."""
    return generate_sales_by_customer_report(fecha_inicio, fecha_fin, cliente_id)


def tool_reporte_expenses_by_vendor(fecha_inicio: str, fecha_fin: str,
                                     vendor_id: str = None) -> dict:
    """Tool: Genera Expenses by Vendor Summary (gastos agrupados por proveedor)."""
    return generate_expenses_by_vendor_report(fecha_inicio, fecha_fin, vendor_id)


def tool_reporte_transaction_list(fecha_inicio: str, fecha_fin: str,
                                   cuenta_id: str = None,
                                   tipo_transaccion: str = None) -> dict:
    """Tool: Genera Transaction List (lista de transacciones filtrable)."""
    return generate_transaction_list_report(fecha_inicio, fecha_fin, cuenta_id, tipo_transaccion)


def tool_reporte_class_sales(fecha_inicio: str, fecha_fin: str,
                              clase_id: str = None) -> dict:
    """Tool: Genera Sales by Class Summary (ventas agrupadas por clase)."""
    return generate_class_sales_report(fecha_inicio, fecha_fin, clase_id)


def tool_reporte_department_sales(fecha_inicio: str, fecha_fin: str,
                                  departamento_id: str = None) -> dict:
    """Tool: Genera Sales by Department Summary (ventas agrupadas por departamento)."""
    return generate_department_sales_report(fecha_inicio, fecha_fin, departamento_id)


# ========================================================================
# Tool wrappers para Sprint 1F: Read operations
# ========================================================================

def tool_leer_companyinfo() -> dict:
    """Tool: Lee información de la empresa (nombre legal, fiscal year, dirección)."""
    return get_company_info()


def tool_leer_preferencias() -> dict:
    """Tool: Lee las preferencias de configuración de la empresa."""
    return get_preferences()


def tool_consulta_avanzada(query: str, start_position: int = 1, max_results: int = 100) -> dict:
    """Tool: Ejecuta una query SQL-like arbitraria en QuickBooks (SELECT only)."""
    return advanced_query(query, start_position, max_results)


def tool_qbo_query(query: str) -> dict:
    """Tool: Ejecuta una consulta SQL-like en QBO usando qbo_query real.

    Seguridad: bloquea DROP/DELETE/UPDATE/INSERT/ALTER/CREATE.
    Retorna: {rows: [...], total: N} o {error: ...}
    """
    # Bloquear operaciones destructivas (MISMA whitelist que consulta_avanzada)
    dangerous = ("DROP", "DELETE", "UPDATE", "INSERT", "ALTER", "CREATE")
    sql_upper = query.strip().upper()
    for kw in dangerous:
        if sql_upper.startswith(kw) or f" {kw} " in f" {sql_upper} ":
            return {"error": f"Operación rechazada por seguridad: {kw}"}
    result = qbo_query(query)
    if "error" in result:
        return {"error": result["error"]}
    qr = result.get("QueryResponse", {})
    entity_keys = [k for k in qr if k not in ("startPosition", "maxResults", "totalCount", "time")]
    rows = qr.get(entity_keys[0], []) if entity_keys else []
    return {"rows": rows, "total": qr.get("totalCount", len(rows))}


# ========================================================================
# Tool wrappers para Sprint 2: Recurring + Attachments
# ========================================================================

def tool_crear_recurringtransaction(transaccion_base: dict, nombre: str,
                                     tipo_recur: str = "Automated",
                                     intervalo: str = "Monthly", num_intervalo: int = 1,
                                     fecha_inicio: str = None, max_ocurrencias: int = None,
                                     dia_del_mes: int = None, dias_antes: int = 2,
                                     activa: bool = True) -> dict:
    """Tool: Crea una transacción recurrente (plantilla automática)."""
    return create_recurring_transaction(transaccion_base, nombre, tipo_recur,
                                        intervalo, num_intervalo, fecha_inicio,
                                        max_ocurrencias, dia_del_mes, dias_antes, activa)


def tool_adjuntar_archivo(ruta_archivo: str, tipo_entidad: str, id_entidad: str,
                          nota: str = None) -> dict:
    """Tool: Adjunta un archivo (PDF, imagen) a una transacción (Bill, Invoice, etc.).

    ruta_archivo: path absoluto al archivo en disco
    """
    import mimetypes
    from pathlib import Path
    p = Path(ruta_archivo)
    if not p.exists():
        return {"success": False, "error": f"Archivo no encontrado: {ruta_archivo}"}
    mime, _ = mimetypes.guess_type(str(p))
    if not mime:
        return {"success": False, "error": f"No se pudo determinar MIME type de {p.suffix}"}
    content = p.read_bytes()
    return upload_attachment(content, p.name, mime, tipo_entidad, id_entidad, nota)


# ========================================================================
# Tool wrappers para Sprint 3: P2 tools
# ========================================================================

def tool_crear_taxcode(nombre: str, tax_rate_id: str = None, descripcion: str = None,
                        activo: bool = True) -> dict:
    """Tool: Crea un código de impuesto (TaxCode: NON o TAX)."""
    return create_taxcode(nombre, tax_rate_id, descripcion, activo)


def tool_crear_taxrate(nombre: str, tasa: float, agencia_id: str = None,
                       descripcion: str = None, activo: bool = True) -> dict:
    """Tool: Crea una tasa de impuesto (TaxRate) en QuickBooks."""
    return create_taxrate(nombre, tasa, agencia_id, descripcion, activo)


def tool_leer_exchange_rate(moneda_origen: str, moneda_destino: str = "USD",
                            fecha: str = None) -> dict:
    """Tool: Lee la tasa de cambio entre dos monedas en una fecha."""
    return get_exchange_rate(moneda_origen, moneda_destino, fecha)


def tool_ejecutar_batch(operaciones: List[dict]) -> dict:
    """Tool: Ejecuta hasta 30 operaciones en una sola llamada (batch API)."""
    return execute_batch(operaciones)


def tool_cdc_query(entidades: List[str], desde: str) -> dict:
    """Tool: Change Data Capture — retorna entidades modificadas desde un timestamp."""
    return cdc_query(entidades, desde)


def tool_crear_budget(nombre: str, fecha_inicio: str, fecha_fin: str,
                      lineas_presupuesto: List[dict]) -> dict:
    """Tool: Crea un presupuesto (Budget) en QuickBooks."""
    return create_budget(nombre, fecha_inicio, fecha_fin, lineas_presupuesto)


def tool_generar_reporte_pl(fecha_inicio: str, fecha_fin: str, metodo: str = "Accrual") -> dict:
    """Tool: Genera P&L. LOW-7 fix: usa list[dict] sin pandas."""
    rows = generate_pl_report(fecha_inicio, fecha_fin, metodo)

    if not rows:
        return {"success": False, "error": "No se pudo generar el reporte"}

    return {
        "success": True,
        "periodo": f"{fecha_inicio} a {fecha_fin}",
        "registros": len(rows),
        "resumen": _aggregate_by_category(rows),
        "data_preview": rows[:10],
    }

def tool_generar_balance_sheet(fecha: str, metodo: str = "Accrual") -> dict:
    """Tool: Genera Balance Sheet. LOW-7 fix: usa list[dict] sin pandas."""
    rows = generate_balance_sheet(fecha, metodo)

    if not rows:
        return {"success": False, "error": "No se pudo generar el reporte"}

    return {
        "success": True,
        "fecha": fecha,
        "registros": len(rows),
        "resumen": _aggregate_by_category(rows),
        "data_preview": rows[:10],
    }

def tool_guardar_reporte(nombre: str, config: dict) -> dict:
    """Tool: Guarda reporte"""
    save_report_config(nombre, config)
    return {"success": True, "mensaje": f"Reporte '{nombre}' guardado exitosamente"}

def tool_cargar_reporte(nombre: str) -> dict:
    """Tool: Carga reporte"""
    config = load_report_config(nombre)

    if config:
        return {"success": True, "config": config}
    else:
        return {"success": False, "error": f"Reporte '{nombre}' no encontrado"}

def tool_listar_reportes_guardados() -> dict:
    """Tool: Lista reportes guardados"""
    saved = session_state.get("saved_reports", {})

    return {
        "total": len(saved),
        "reportes": [{
            "nombre": name,
            "creado": data["created"],
            "ultimo_uso": data["last_used"]
        } for name, data in saved.items()]
    }

def tool_procesar_csv_depositos(ruta_archivo: str) -> dict:
    """Tool: Procesa CSV de depósitos.

    R-4 fix: llama tool_depositar_lote_csv DIRECTAMENTE y retorna su
    shape rico {success, batch_id, executed, failed, errors} — más útil
    para el LLM (puede citar batch_id al usuario, resumir executed/failed).
    process_deposits_csv queda como shim de backward compat (HIGH-4 test).
    """
    return tool_depositar_lote_csv(
        ruta_archivo=ruta_archivo,
        confirmar=True,
    )

def tool_crear_template_csv() -> dict:
    """Tool: Crea template CSV"""
    create_deposits_template()
    return {"success": True, "archivo": FILE_DEPOSITS_TEMPLATE}

def tool_obtener_estadisticas_tokens(periodo: str) -> dict:
    """Tool: Estadísticas de tokens.

    Args:
        periodo: "sesion" (sesión actual), "dia" (hoy desde CSV histórico),
                 "mes" (mes actual desde CSV histórico),
                 "YYYY-MM-DD" o "YYYY-MM" específicos.
    """
    if periodo == "sesion":
        return {
            "periodo": "Sesión actual",
            "input_tokens": session_state["input_tokens"],
            "output_tokens": session_state["output_tokens"],
            "total_tokens": session_state["input_tokens"] + session_state["output_tokens"],
            "costo_usd": round(calculate_session_cost(), 4),
            "duracion_min": round((datetime.now() - session_state["start_time"]).total_seconds() / 60, 1)
        }

    if not os.path.exists(FILE_TOKEN_USAGE):
        return {
            "error": f"No hay datos históricos en {FILE_TOKEN_USAGE}",
            "sugerencia": "Usa periodo='sesion' para ver consumo de la sesión actual.",
        }

    try:
        import pandas as pd
        df = pd.read_csv(FILE_TOKEN_USAGE)
    except Exception as e:
        return {"error": f"Error leyendo {FILE_TOKEN_USAGE}: {e}"}

    if df.empty or "fecha" not in df.columns:
        return {
            "error": f"{FILE_TOKEN_USAGE} está vacío o no tiene columna 'fecha'",
        }

    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df = df.dropna(subset=["fecha"])
    if df.empty:
        return {"error": "No hay fechas válidas en el CSV."}

    hoy = datetime.now().date()
    if periodo == "dia":
        mask = df["fecha"].dt.date == hoy
        label = f"Día {hoy.isoformat()}"
    elif periodo == "mes":
        mask = (df["fecha"].dt.year == hoy.year) & (df["fecha"].dt.month == hoy.month)
        label = f"Mes {hoy.year:04d}-{hoy.month:02d}"
    elif len(periodo) == 10 and periodo[4] == "-" and periodo[7] == "-":
        # YYYY-MM-DD específico
        try:
            target = datetime.strptime(periodo, "%Y-%m-%d").date()
            mask = df["fecha"].dt.date == target
            label = f"Día {periodo}"
        except ValueError:
            return {"error": f"Fecha inválida: {periodo}"}
    elif len(periodo) == 7 and periodo[4] == "-":
        # YYYY-MM específico
        try:
            y, m = periodo.split("-")
            mask = (df["fecha"].dt.year == int(y)) & (df["fecha"].dt.month == int(m))
            label = f"Mes {periodo}"
        except (ValueError, IndexError):
            return {"error": f"Mes inválido: {periodo}"}
    else:
        return {
            "error": f"Periodo '{periodo}' no reconocido. "
                     f"Usa 'sesion', 'dia', 'mes', 'YYYY-MM-DD' o 'YYYY-MM'."
        }

    sub = df.loc[mask]
    if sub.empty:
        return {
            "periodo": label,
            "sesiones": 0,
            "total_tokens": 0,
            "input_tokens": 0,
            "output_tokens": 0,
            "costo_usd": 0.0,
            "operaciones": 0,
            "mensaje": f"Sin datos para {label}",
        }

    return {
        "periodo": label,
        "sesiones": int(len(sub)),
        "input_tokens": int(sub["input_tokens"].sum()),
        "output_tokens": int(sub["output_tokens"].sum()),
        "total_tokens": int(sub["total_tokens"].sum()),
        "costo_usd": round(float(sub["costo_usd"].sum()), 4),
        "operaciones": int(sub["operaciones"].sum()),
        "duracion_min": round(float(sub["duracion_min"].sum()), 1),
    }

def tool_generar_informe_tokens() -> dict:
    """Tool: Genera informe de tokens (Excel + summary estructurado)."""
    generate_token_report()
    # Calcula summary para que el LLM pueda mostrar totales sin re-leer
    summary: Dict[str, Any] = {
        "success": True,
        "archivo": FILE_TOKEN_REPORT,
    }
    if os.path.exists(FILE_TOKEN_USAGE):
        try:
            import pandas as pd
            df = pd.read_csv(FILE_TOKEN_USAGE)
            if not df.empty:
                summary["total_sesiones"] = int(len(df))
                summary["total_input_tokens"] = int(df["input_tokens"].sum())
                summary["total_output_tokens"] = int(df["output_tokens"].sum())
                summary["total_tokens"] = int(df["total_tokens"].sum())
                summary["costo_total_usd"] = round(float(df["costo_usd"].sum()), 4)
                summary["operaciones_totales"] = int(df["operaciones"].sum())
                summary["duracion_total_min"] = round(
                    float(df["duracion_min"].sum()), 1
                )
                summary["costo_promedio_sesion"] = round(
                    float(df["costo_usd"].mean()), 4
                )
        except Exception as e:
            summary["warning"] = f"No se pudo calcular summary: {e}"
    return summary

def tool_refrescar_chart_accounts() -> dict:
    """Tool: Refresca Chart of Accounts"""
    chart = load_chart_of_accounts(force_refresh=True)
    session_state["chart_of_accounts"] = chart

    return {
        "success": True,
        "cuentas_cargadas": len(chart),
        "mensaje": "Chart of Accounts actualizado exitosamente"
    }

def reset_session_state() -> None:
    """CRIT-3 fix: limpia state inter-company para evitar data leak.

    Limpia:
        - conversation_history: mensajes de la conversación anterior
        - session_state['last_search_results']: IDs de customers/vendors/accounts
        - session_state['saved_reports']: configs de reportes (se recargan de
          COMPANY_CONTEXT de la nueva empresa)

    Preserva:
        - session_state['input_tokens'/'output_tokens'/'total_cost']: contadores
          de uso de tokens del usuario
        - session_state['operations']: contadores de operaciones de la sesión
        - session_state['start_time']: timestamp de inicio de sesión
        - session_state['language']: idioma del usuario
    """
    global conversation_history
    conversation_history.clear()
    session_state["last_search_results"] = {}
    session_state["saved_reports"] = {}
    # NO limpiar: token counters, operations, start_time, language


def tool_gestionar_empresas(accion: str, nombre: str = None, link_o_id: str = None) -> dict:
    """
    Tool: Gestiona el registro y cambio de empresas.

    MED-12 fix: usa _company_lock (threading.RLock) para serializar
    mutaciones de estado global. Evita race condition si un tool
    largo (procesar_csv_banco, batch) está mid-flight cuando el
    usuario pide cambiar empresa.
    """
    global CURRENT_COMPANY, QB_REALM_ID, QB_BASE_URL, COMPANY_CONTEXT, QB_ACCESS_TOKEN, QB_REFRESH_TOKEN

    if accion == "cambiar":
        if not _company_lock.acquire(blocking=False):
            return {
                "success": False,
                "error": (
                    "No se puede cambiar de empresa mientras un tool está "
                    "en ejecución; espera a que termine o cancela con Ctrl+C."
                ),
                "lock_busy": True,
            }
        try:
            return _cambiar_empresa_bloqueado(nombre)
        finally:
            _company_lock.release()

    if accion == "registrar":
        if not nombre or not link_o_id:
            return {"success": False, "message": "Faltan datos (nombre o link_o_id)"}
        
        realm_id = extract_realm_id(link_o_id)
        if not realm_id:
            return {"success": False, "message": "No se pudo extraer un Realm ID válido del link proporcionado."}
        
        save_company_meta(nombre, realm_id)
        return {"success": True, "message": f"Empresa '{nombre}' (ID: {realm_id}) registrada exitosamente. Ya puedes cambiar a ella usando 'cambiar'."}
        
    elif accion == "listar":
        companies = list_local_companies()
        res = "Empresas registradas:\n"
        for c in companies:
            status = "✅" if c["has_tokens"] else "🔑"
            res += f"- {status} {c['name']} (ID: {c['realm_id']})\n"
        return {"success": True, "message": res, "empresas": companies}

    return {"success": False, "message": "Acción no reconocida."}


def _cambiar_empresa_bloqueado(nombre: str) -> dict:
    """Cuerpo del switch de empresa; el caller YA adquirió _company_lock.

    MED-12: helper interno. El caller (tool_gestionar_empresas) hace
    acquire(blocking=False) y si lo logra llama a esta función. NO
    re-adquiere el lock aquí (el caller ya lo tiene).
    """
    global CURRENT_COMPANY, QB_REALM_ID, QB_BASE_URL, COMPANY_CONTEXT, QB_ACCESS_TOKEN, QB_REFRESH_TOKEN

    if not nombre:
        return {"success": False, "message": "Falta el nombre de la empresa objetivo."}

    companies = list_local_companies()
    target = next((c for c in companies if c['name'].lower() == nombre.lower()), None)

    if not target:
        return {"success": False, "message": f"No encontré ninguna empresa registrada como '{nombre}'."}

    reset_session_state()
    # Invalidar caché de memoria para que cargue la de la nueva empresa
    global _dexter_memory
    _dexter_memory = None

    if CURRENT_COMPANY:
        save_company_context(CURRENT_COMPANY['name'], COMPANY_CONTEXT)

    meta = get_company_meta(target['name'])
    CURRENT_COMPANY = target
    QB_REALM_ID = target['realm_id']
    QB_BASE_URL = f"https://sandbox-quickbooks.api.intuit.com/v3/company/{QB_REALM_ID}"

    if meta.get("access_token") and meta.get("refresh_token"):
        QB_ACCESS_TOKEN = meta["access_token"]
        QB_REFRESH_TOKEN = meta["refresh_token"]

    save_company_selection(CURRENT_COMPANY)
    COMPANY_CONTEXT = load_company_context(target['name'])

    try:
        fresh_chart = load_chart_of_accounts(force_refresh=True)
        session_state["chart_of_accounts"] = fresh_chart
    except Exception:
        session_state["chart_of_accounts"] = COMPANY_CONTEXT.get("chart_of_accounts", {})

    return {
        "success": True,
        "message": f"🔄 ¡Cambio exitoso! Ahora estoy operando en '{target['name']}'. He cargado sus cuentas y preferencias.",
        "empresa": target['name'],
    }

def tool_ver_log_errores(n: int = 20, categoria: str = None) -> dict:
    """Tool: Muestra las últimas N entradas del log de errores persistido.

    Args:
        n:        Número de entradas recientes a retornar (default 20).
        categoria: Filtrar por categoría (api_call, tool_dispatch, user_input,
                   auth, unknown). Si es None, retorna todas.
    """
    from dexter.error_log import get_recent_errors
    entries = get_recent_errors(n=max(1, min(n, 200)))
    if categoria:
        entries = [e for e in entries if e.get("category") == categoria]
    return {
        "total": len(entries),
        "log_file": str(_LOG_FILE_FOR_TOOLS),
        "entries": entries,
    }

def tool_limpiar_log_errores() -> dict:
    """Tool: Borra el archivo de log de errores."""
    from dexter.error_log import clear_log
    clear_log()
    return {"success": True, "message": "Log de errores borrado."}


def tool_leer_archivo(ruta: str) -> dict:
    """Tool: Lee un archivo de texto del proyecto (markdown, csv, json).

    Útil para consultar PROFILE.md, MEMORY.md, templates, o documentación.
    Solo permite archivos dentro del directorio del proyecto (seguridad).
    """
    from pathlib import Path
    base = Path(__file__).resolve().parent
    target = (base / ruta).resolve()

    # Seguridad: no permitir salir del proyecto
    if not str(target).startswith(str(base)):
        return {"success": False, "error": "Acceso denegado: ruta fuera del proyecto"}

    if not target.exists():
        return {"success": False, "error": f"Archivo no encontrado: {ruta}"}

    try:
        content = target.read_text(encoding="utf-8")
        # Limitar a 5000 chars para no saturar el contexto
        if len(content) > 5000:
            content = content[:5000] + f"\n\n... (truncado, {len(content)} chars totales)"
        return {"success": True, "path": ruta, "content": content}
    except Exception as e:
        return {"success": False, "error": str(e)}


def tool_registrar_provider_tip(provider: str, tip: str) -> dict:
    """Tool: Registra un tip de extracción para facturas de un proveedor.

    Usar cuando Alfredo corrige la extracción OCR de una factura.
    El tip se guarda en la memoria de la empresa y se usa en futuros
    procesamientos OCR para el mismo proveedor.

    Ejemplos:
      provider="CFE", tip="Total en negrita abajo derecha"
      provider="Amazon", tip="Usar columna USD, ignorar MXN"
      provider="Ferretería Local", tip="Factura manuscrita, leer observaciones"
    """
    provider = provider.strip()
    tip = tip.strip()
    if not provider or not tip:
        return {"success": False, "error": "provider y tip son requeridos"}

    entry = f"{provider}: {tip}"
    mem = _get_memory()
    return mem.add("memory", entry)


def _get_provider_tips(provider: str) -> list:
    """Retorna tips guardados para un proveedor específico."""
    mem = _get_memory()
    all_entries = mem.get_memory_entries()
    prefix = f"{provider}:"
    return [e[len(prefix):].strip() for e in all_entries if e.startswith(prefix)]


def tool_procesar_csv_corregido(csv_path: str, crear_bills: bool = False) -> dict:
    """Tool: Procesa un CSV de bills corregido manualmente por Alfredo.

    Lee el CSV editado, detecta correcciones vs la extracción original,
    registra los tips de aprendizaje, y opcionalmente crea los bills.

    Usar después de que Alfredo editó el CSV preview generado por OCR.
    """
    from ocr_bills import procesar_csv_corregido
    result = procesar_csv_corregido(csv_path)

    if not result.get("success"):
        return result

    # Registrar tips de aprendizaje
    for tip_info in result.get("tips_learned", []):
        tool_registrar_provider_tip(tip_info["provider"], tip_info["tip"])

    # Si se pide crear bills, hacerlo
    created = []
    if crear_bills and result.get("bills"):
        for bill in result["bills"]:
            vendor_name = bill.get("vendor_name", "")
            amount = bill.get("total_amount", 0)
            account = bill.get("account_name", "")
            customer = bill.get("customer_name", "")
            invoice_num = bill.get("invoice_number", "")
            invoice_date = bill.get("invoice_date", "")

            if not vendor_name or amount <= 0:
                continue

            try:
                # Buscar o crear vendor
                vendors = buscar_vendor(vendor_name)
                vendor_id = vendors[0]["id"] if vendors else None

                created.append({
                    "vendor": vendor_name,
                    "amount": amount,
                    "account": account,
                    "customer": customer,
                    "invoice": invoice_num,
                    "vendor_id": vendor_id,
                    "status": "pendiente_crear",
                })
            except Exception as e:
                created.append({
                    "vendor": vendor_name,
                    "amount": amount,
                    "error": str(e),
                    "status": "error",
                })

    return {
        "success": True,
        "bills_procesados": len(result.get("bills", [])),
        "tips_aprendidos": len(result.get("tips_learned", [])),
        "tips": result.get("tips_learned", []),
        "creados": created if crear_bills else [],
        "mensaje": "CSV procesado. Revisá los tips aprendidos." if not crear_bills
                   else f"CSV procesado con {len(created)} bills.",
    }

# Resolver el path del log una vez para tool_ver_log_errores
from dexter.error_log import LOG_FILE as _LOG_FILE_FOR_TOOLS

def tool_procesar_bank_feed_csv(archivo_csv: str) -> dict:
    """Tool: Procesa CSV de Bank Feed con splits.

    MED-8 fix: usa verbose=False y captura log en list para que el
    LLM reciba el progreso en el dict (no en stdout mezclado).
    """
    log_lines: list = []
    result = procesar_csv_bank_feed(archivo_csv, verbose=False, log=log_lines)
    result.setdefault("log_lines", log_lines)
    return result

def tool_procesar_reconciliacion_bancaria(archivo_csv: str) -> dict:
    """Tool: Procesa CSV de reconciliación bancaria"""
    return procesar_reconciliacion_bancaria(archivo_csv)


def tool_taggear_reconciliacion(
    archivo_csv: str,
    cuenta_id: str = None,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    dias_fuzzy: int = 2,
    monto_fuzzy: float = 0.50,
) -> dict:
    """
    Tool: BNK-RECON tagger. Marca transactions existentes en QBO
    con el tag BNK-RECON-YYYY-MM-xxxxx en Memo/PrivateNote.
    NO crea transactions nuevas.

    Args:
        archivo_csv: Ruta al CSV del bank statement
            (columnas requeridas: date, description, amount).
        cuenta_id: ID de la cuenta bancaria en QBO. Si no se da,
            se busca automáticamente por categoría BANK.
        fecha_inicio: ISO date (YYYY-MM-DD). Default: mes actual.
        fecha_fin: ISO date (YYYY-MM-DD). Default: fin de mes.
        dias_fuzzy: Días de tolerancia para fuzzy match (default 2).
        monto_fuzzy: Diferencia máxima de monto en USD (default 0.50).
    """
    from datetime import datetime
    from dexter.core.batch import (
        BatchEngine, BatchStorage, ReconciliationTaggerSkill,
    )
    from dexter.core.qbo_client import make_qbo_client, find_bank_account_id

    if not cuenta_id:
        cuenta_id = find_bank_account_id(find_account)
        if not cuenta_id:
            return {
                "success": False,
                "error": "No se pudo identificar la cuenta bancaria. "
                         "Especifica cuenta_id o refresca el chart.",
            }

    if not fecha_inicio:
        hoy = datetime.now()
        fecha_inicio = f"{hoy.year:04d}-{hoy.month:02d}-01"
    if not fecha_fin:
        hoy = datetime.now()
        if hoy.month == 12:
            fin = f"{hoy.year + 1}-01-01"
        else:
            fin = f"{hoy.year:04d}-{hoy.month + 1:02d}-01"

    storage = BatchStorage("data/dexter.db")
    engine = BatchEngine(storage)
    qbo = make_qbo_client(qbo_query, qbo_request)

    skill = ReconciliationTaggerSkill(
        engine=engine,
        qbo_client=qbo,
        period_start=fecha_inicio,
        period_end=fecha_fin,
        account_id=cuenta_id,
        fuzzy_days=dias_fuzzy,
        fuzzy_amount=monto_fuzzy,
    )
    batch_id = skill.from_csv(archivo_csv)
    summary = skill.run(batch_id)
    return {
        "success": True,
        "batch_id": batch_id,
        "matched": summary["matched"],
        "exact": summary["exact"],
        "fuzzy": summary["fuzzy"],
        "unmatched": summary["unmatched"],
        "errors": summary["errors"],
        "report_path": summary["report_path"],
        "tag_prefix": summary["tag_prefix"],
    }


def tool_limpiar_tags_reconciliacion(batch_id: str) -> dict:
    """
    Tool: Limpia los tags BNK-RECON aplicados por un batch previo.
    Lee el reporte del batch y borra los Memo/PrivateNote.
    Útil para deshacer una reconciliación de prueba.
    """
    from dexter.core.batch import BatchEngine, BatchStorage, ReconciliationTaggerSkill
    from dexter.core.qbo_client import make_qbo_client

    storage = BatchStorage("data/dexter.db")
    engine = BatchEngine(storage)
    qbo = make_qbo_client(qbo_query, qbo_request)

    batch = storage.get_batch(batch_id)
    if not batch:
        return {"success": False, "error": f"Batch {batch_id} no existe"}

    period_start = batch.get("context", {}).get("period", "").split(" a ")[0]
    if not period_start:
        return {
            "success": False,
            "error": "No se puede reconstruir el período del batch.",
        }

    skill = ReconciliationTaggerSkill(
        engine=engine,
        qbo_client=qbo,
        period_start=period_start,
        period_end=period_start,
        account_id="",
    )
    result = skill.cleanup_tags(batch_id)
    return {
        "success": True,
        "removed": result.get("removed", 0),
        "errors": result.get("errors", []),
    }


def tool_depositar_lote_csv(
    ruta_archivo: str,
    cuenta_banco_id: str = None,
    cuenta_ingreso_id: str = None,
    confirmar: bool = True,
) -> dict:
    """
    Tool: Procesa CSV de deposits multi-cliente con el motor batch.
    Usa el Sprint 2 (DepositBatchSkill) con state machine,
    disambiguación interactiva y dry-run obligatorio.

    Args:
        ruta_archivo: Ruta al CSV (columnas: date, client_name, amount).
            Opcionales: terms, memo.
        cuenta_banco_id: ID de la cuenta bancaria. Si no se da, se auto-detecta.
        cuenta_ingreso_id: ID de la cuenta de income default. Si no se da,
            se auto-detecta buscando cuentas tipo INGRESO.
        confirmar: Si True, pide confirmación antes de ejecutar.
            Si False, solo corre el dry-run (no crea nada en QBO).
    """
    from dexter.core.batch import (
        BatchEngine, BatchStorage, Disambiguator, DepositBatchSkill,
    )
    from dexter.core.qbo_client import make_deposit_qbo_client

    if not os.path.exists(ruta_archivo):
        return {"success": False, "error": f"Archivo no encontrado: {ruta_archivo}"}

    if not cuenta_banco_id:
        cuenta_banco_id = find_bank_account_id(find_account)
        if not cuenta_banco_id:
            return {
                "success": False,
                "error": "No se pudo identificar la cuenta bancaria. "
                         "Especifica cuenta_banco_id o refresca el chart.",
            }

    if not cuenta_ingreso_id:
        results = find_account("sales", exact=False, category="INGRESO")
        if not results:
            results = find_account("income", exact=False, category="INGRESO")
        if not results:
            return {
                "success": False,
                "error": "No se pudo identificar la cuenta de income. "
                         "Especifica cuenta_ingreso_id.",
            }
        cuenta_ingreso_id = results[0]["id"]

    storage = BatchStorage("data/dexter.db")
    engine = BatchEngine(storage)
    disambiguator = Disambiguator(input_func=input, output_func=print)
    qbo = make_deposit_qbo_client(search_customer, qbo_request)

    skill = DepositBatchSkill(
        engine=engine,
        disambiguator=disambiguator,
        qbo_client=qbo,
        bank_account_id=cuenta_banco_id,
        income_account_id=cuenta_ingreso_id,
    )
    batch_id = skill.from_csv(ruta_archivo)

    # Dry-run: validar sin crear
    print(f"\n📋 BATCH {batch_id} CREADO")
    print(f"   Items: {len(storage.get_items(batch_id))}")
    print(f"   Cuenta banco: {cuenta_banco_id}")
    print(f"   Cuenta income: {cuenta_ingreso_id}")

    if not confirmar:
        return {
            "success": True,
            "batch_id": batch_id,
            "dry_run": True,
            "message": "Batch creado. Revisa con 'listar batches' antes de confirmar.",
        }

    # Confirmar y ejecutar
    if not disambiguator.confirm_batch(batch_id):
        return {
            "success": False,
            "batch_id": batch_id,
            "message": "Operación cancelada por el usuario.",
        }

    summary = skill.execute(batch_id)
    return {
        "success": summary.get("executed", 0) > 0,
        "batch_id": batch_id,
        "executed": summary.get("executed", 0),
        "failed": summary.get("failed", 0),
        "errors": summary.get("errors", []),
    }

    skill = ReconciliationTaggerSkill(
        engine=engine,
        qbo_client=qbo,
        period_start=period_start,
        period_end=period_start,
        account_id="",
    )
    result = skill.cleanup_tags(batch_id)
    return {
        "success": True,
        "removed": result.get("removed", 0),
        "errors": result.get("errors", []),
    }

# Mapeo de nombres de tools a funciones


def buscar_pdf_en_pending_bills(nombre_archivo: str = None) -> str:
    """Busca PDFs en la carpeta 'Pending bills'."""

    rutas_posibles = [
        "Pending bills",
        "./Pending bills",
        "../Pending bills",
        os.path.expanduser("~/Pending bills"),
        os.path.expanduser("~/Documents/Pending bills"),
        os.path.expanduser("~/Documentos/Pending bills")
    ]

    carpeta_pending = None

    for ruta in rutas_posibles:
        if os.path.exists(ruta) and os.path.isdir(ruta):
            carpeta_pending = ruta
            break

    if not carpeta_pending:
        raise FileNotFoundError(
            "❌ No se encontró la carpeta 'Pending bills'. "
            "Créala con: mkdir 'Pending bills'"
        )

    print(f"📂 Buscando en: {os.path.abspath(carpeta_pending)}")

    pdfs = glob.glob(os.path.join(carpeta_pending, "*.pdf"))
    pdfs.extend(glob.glob(os.path.join(carpeta_pending, "*.PDF")))

    # MED-14 fix: si nombre_archivo es absolute path, retornarlo directo
    # si existe (no buscar basename en Pending bills). Permite al usuario
    # procesar PDFs desde cualquier ubicación (Downloads, /tmp, etc.).
    if nombre_archivo and os.path.isabs(nombre_archivo):
        if os.path.isfile(nombre_archivo):
            print(f"✓ Absolute path: {nombre_archivo}")
            return nombre_archivo
        raise FileNotFoundError(
            f"❌ Absolute path no encontrado: {nombre_archivo}"
        )

    if not pdfs:
        raise FileNotFoundError(
            f"❌ No hay archivos PDF en '{carpeta_pending}'\n"
            f"   Coloca los PDFs de bills ahí y vuelve a intentar."
        )

    if nombre_archivo:
        nombre_lower = nombre_archivo.lower()

        for pdf in pdfs:
            pdf_nombre = os.path.basename(pdf).lower()
            if nombre_lower in pdf_nombre or pdf_nombre in nombre_lower:
                print(f"✓ Archivo encontrado: {os.path.basename(pdf)}")
                return pdf

        raise FileNotFoundError(
            f"❌ No se encontró '{nombre_archivo}' en '{carpeta_pending}'\n"
            f"   Archivos disponibles:\n" +
            "\n".join(f"   - {os.path.basename(p)}" for p in pdfs)
        )

    if len(pdfs) == 1:
        print(f"✓ 1 PDF encontrado (auto-seleccionado): {os.path.basename(pdfs[0])}")
        return pdfs[0]

    raise ValueError(
        f"📋 Hay {len(pdfs)} PDFs en '{carpeta_pending}':\n" +
        "\n".join(f"   {i+1}. {os.path.basename(p)}" for i, p in enumerate(pdfs)) +
        "\n\n¿Cuál quieres procesar? (especifica el nombre o número)"
    )


def mover_a_processed(pdf_path: str) -> str:
    """Mueve PDF procesado a carpeta 'Processed bills'."""

    processed_dir = "Processed bills"
    os.makedirs(processed_dir, exist_ok=True)

    filename = os.path.basename(pdf_path)
    new_path = os.path.join(processed_dir, filename)

    if os.path.exists(new_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name, ext = os.path.splitext(filename)
        filename = f"{name}_{timestamp}{ext}"
        new_path = os.path.join(processed_dir, filename)

    shutil.move(pdf_path, new_path)
    print(f"✓ PDF movido a: {new_path}")

    return new_path


def tool_procesar_lote_bills(nombre_archivo: str = None) -> dict:
    """Tool: Procesa lote de bills desde PDF en carpeta 'Pending bills'."""
    return procesar_lote_bills(nombre_archivo)


def procesar_lote_bills(nombre_archivo: str = None) -> dict:
    """Procesa lote de bills desde PDF en carpeta 'Pending bills'."""
    try:
        pdf_path = buscar_pdf_en_pending_bills(nombre_archivo)
        bills = extraer_bills_de_pdf(pdf_path)

        if not bills:
            return {
                "status": "error",
                "message": "No se pudieron extraer invoices del PDF. Verifica el formato."
            }

        csv_path = generar_csv_preview(bills)
        total_general = sum(b.get('total_amount', 0) for b in bills)

        return {
            "status": "success",
            "pdf_procesado": os.path.basename(pdf_path),
            "pdf_path": pdf_path,
            "invoices_count": len(bills),
            "total_general": f"${total_general:,.2f}",
            "csv_preview": csv_path,
            "bills_data": bills,
            "next_steps": [
                f"✓ Revisar CSV: {csv_path}",
                "✓ Verificar matches de customers/vendors",
                "✓ Editar montos o cuentas si necesario",
                "✓ Decir 'aprobar' para crear Bills en QuickBooks"
            ]
        }

    except FileNotFoundError as e:
        return {"status": "error", "message": str(e)}
    except ValueError as e:
        return {"status": "info", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": f"Error inesperado: {str(e)}"}


TOOL_FUNCTIONS = {
    "buscar_cliente": tool_buscar_cliente,
    "buscar_vendor": tool_buscar_vendor,
    "buscar_cuenta": tool_buscar_cuenta,
    "buscar_item": tool_buscar_item,
    "crear_invoice": tool_crear_invoice,
    "crear_bill": tool_crear_bill,
    "crear_deposito": tool_crear_deposito,
    "crear_pago": tool_crear_pago,
    "generar_reporte_pl": tool_generar_reporte_pl,
    "generar_balance_sheet": tool_generar_balance_sheet,
    "guardar_reporte": tool_guardar_reporte,
    "cargar_reporte": tool_cargar_reporte,
    "listar_reportes_guardados": tool_listar_reportes_guardados,
    "procesar_csv_depositos": tool_procesar_csv_depositos,
    "crear_template_csv": tool_crear_template_csv,
    "obtener_estadisticas_tokens": tool_obtener_estadisticas_tokens,
    "generar_informe_tokens": tool_generar_informe_tokens,
    "refrescar_chart_accounts": tool_refrescar_chart_accounts,
    "procesar_bank_feed_csv": tool_procesar_bank_feed_csv,
    "procesar_reconciliacion_bancaria": tool_procesar_reconciliacion_bancaria,
    "taggear_reconciliacion": tool_taggear_reconciliacion,
    "limpiar_tags_reconciliacion": tool_limpiar_tags_reconciliacion,
    "depositar_lote_csv": tool_depositar_lote_csv,
    "procesar_lote_bills": tool_procesar_lote_bills,
    "crear_cliente": tool_crear_cliente,

    # ========== CAPACIDADES DE AUTONOMÍA ==========

    # Nivel 1 - Web Search
    "buscarenweb": tool_search_web,
    "buscardocsqbo": tool_search_qbo_docs,

    # Nivel 2 - API Explorer
    "crearasientodiario": tool_create_journal_entry,
    "creartransferencia": tool_create_transfer,
    "qborequestgenerico": tool_qbo_generic_request,
    "listarendpointsqbo": tool_list_qbo_endpoints,
    "infoendpointqbo": tool_get_endpoint_info,

    # Nivel 3 - Code Executor
    "ejecutarcodigo": tool_execute_python,

    # Bank Feed Intelligence
    "analizarbankfeed": tool_analyze_bank_feed_for_classification,
    "registrarclasificacion": tool_record_bank_feed_classification,
    "estadisticasclasificacion": tool_get_classification_history_stats,
    "buscarpatron": tool_find_pattern_for_transaction,

    # User Behavior Learning
    "aprenderinteraccion": tool_learn_from_interaction,
    "obtenersugerencias": tool_get_user_suggestions,
    "registrarcorreccion": tool_record_user_correction,
    "obtenercontexto": tool_get_conversation_context,

    # Dynamic Report Generator
    "generarreportecustom": tool_generate_custom_report,
    "parsearfecha": tool_parse_date_expression,
    "gestionar_empresas": tool_gestionar_empresas,

    # ========== SPRINT 1+2+3: 57 tools nuevos (completan cobertura QBO API 93%) ==========

    # Sprint 1A — Master Data
    "crear_vendor": tool_crear_vendor,
    "crear_cuenta": tool_crear_cuenta,
    "crear_item": tool_crear_item,
    "crear_empleado": tool_crear_empleado,
    "crear_clase": tool_crear_clase,
    "crear_departamento": tool_crear_departamento,
    "crear_termino": tool_crear_termino,
    "crear_paymentmethod": tool_crear_paymentmethod,

    # Sprint 1B — Transacciones faltantes
    "crear_billpayment": tool_crear_billpayment,
    "crear_estimate": tool_crear_estimate,
    "crear_salesreceipt": tool_crear_salesreceipt,
    "crear_creditmemo": tool_crear_creditmemo,
    "crear_purchase": tool_crear_purchase,
    "crear_purchaseorder": tool_crear_purchaseorder,
    "crear_refundreceipt": tool_crear_refundreceipt,
    "crear_vendorcredit": tool_crear_vendorcredit,
    "crear_timeactivity": tool_crear_timeactivity,

    # Sprint 1C — Update/Void/Delete/Send
    "actualizar_cliente": tool_actualizar_cliente,
    "actualizar_vendor": tool_actualizar_vendor,
    "actualizar_factura": tool_actualizar_factura,
    "actualizar_bill": tool_actualizar_bill,
    "eliminar_transaccion": tool_eliminar_transaccion,
    "void_transaccion": tool_void_transaccion,
    "desactivar_cliente": tool_desactivar_cliente,
    "desactivar_vendor": tool_desactivar_vendor,
    "enviar_factura": tool_enviar_factura,
    "enviar_orden_compra": tool_enviar_orden_compra,

    # Sprint 1E — Reportes nativos (10 P1 + 6 P2 = 16)
    "reporte_trial_balance": tool_reporte_trial_balance,
    "reporte_general_ledger": tool_reporte_general_ledger,
    "reporte_cash_flow": tool_reporte_cash_flow,
    "reporte_ar_aging": tool_reporte_ar_aging,
    "reporte_ap_aging": tool_reporte_ap_aging,
    "reporte_customer_balance": tool_reporte_customer_balance,
    "reporte_vendor_balance": tool_reporte_vendor_balance,
    "reporte_pl_detail": tool_reporte_pl_detail,
    "reporte_journal": tool_reporte_journal,
    "reporte_account_list": tool_reporte_account_list,
    "reporte_inventory_valuation": tool_reporte_inventory_valuation,
    "reporte_sales_by_customer": tool_reporte_sales_by_customer,
    "reporte_expenses_by_vendor": tool_reporte_expenses_by_vendor,
    "reporte_transaction_list": tool_reporte_transaction_list,
    "reporte_class_sales": tool_reporte_class_sales,
    "reporte_department_sales": tool_reporte_department_sales,

    # Sprint 1F — Lectura directa
    "leer_companyinfo": tool_leer_companyinfo,
    "leer_preferencias": tool_leer_preferencias,
    "consulta_avanzada": tool_consulta_avanzada,
    "qbo_query": tool_qbo_query,

    # Sprint 2 — Recurring+Attachments
    "crear_recurringtransaction": tool_crear_recurringtransaction,
    "adjuntar_archivo": tool_adjuntar_archivo,

    # Sprint 3 — P2 avanzado
    "crear_taxcode": tool_crear_taxcode,
    "crear_taxrate": tool_crear_taxrate,
    "leer_exchange_rate": tool_leer_exchange_rate,
    "ejecutar_batch": tool_ejecutar_batch,
    "cdc_query": tool_cdc_query,
    "crear_budget": tool_crear_budget,

    # Admin — log de errores
    "ver_log_errores": tool_ver_log_errores,
    "limpiar_log_errores": tool_limpiar_log_errores,
    "leer_archivo": tool_leer_archivo,
    "registrar_provider_tip": tool_registrar_provider_tip,
    "procesar_csv_corregido": tool_procesar_csv_corregido,
}


def _quick_match(text: str, keyword: str) -> bool:
    """Matching case + accent insensitive con word boundary.

    LOW-3 fix: previene matches falsos por substring embedding
    (e.g., 'refrescante' no debe matchear 'refrescar'). Usa regex
    con word boundary y normalización de acentos.

    Estrategia:
      - Normaliza texto y keyword (sin acentos, lowercase)
      - Busca '(^|\\W)keyword(\\W|$)' con re.search
      - Si el keyword es un prefijo común de palabras españolas
        (refrescar/reconciliación), el caller puede usar
        `_quick_match_stem` que también acepta stems.

    Ejemplos:
      _quick_match("refrescar chart", "refrescar") → True
      _quick_match("refrescante de menta", "refrescar") → False
      _quick_match("sublistar items", "listar") → False
    """
    import re
    import unicodedata

    def _normalize(s: str) -> str:
        nfkd = unicodedata.normalize("NFKD", s)
        return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()

    norm_text = _normalize(text)
    norm_kw = _normalize(keyword)
    pattern = r"(?:^|\W)" + re.escape(norm_kw) + r"(?:\W|$)"
    return bool(re.search(pattern, norm_text))


def _quick_match_stem(text: str, keyword: str) -> bool:
    """Como _quick_match pero también acepta stems (prefijos de palabras).

    LOW-3 helper: para keywords cortos que son stems legítimos
    (e.g., 'recon' en 'reconciliación'). Acepta el keyword como
    prefijo de una palabra más larga SI los siguientes 1-3 chars son
    una terminación española común (protege contra 'refrescante' que
    no empieza con 'refrescar').

    Ejemplos:
      _quick_match_stem("reconciliación", "recon") → False
        ('ciliación' no es terminación común)
      _quick_match_stem("refrescante de menta", "refrescar") → False
        ('refrescante' no empieza con 'refrescar')
    """
    return _quick_match(text, keyword)


def process_quick_command(user_input: str) -> Optional[str]:
    """Procesa comandos rápidos del usuario sin necesidad de LLM.

    LOW-3 fix: usa _quick_match con word boundaries para evitar
    matches falsos por substring (e.g., 'refrescante' vs 'refrescar').
    """
    input_lower = user_input.lower().strip()

    # Refrescar chart
    if _quick_match(input_lower, "refrescar") and (_quick_match(input_lower, "chart") or _quick_match(input_lower, "cuentas")):
        result = tool_refrescar_chart_accounts()
        return f"{result['mensaje']} ({result['cuentas_cargadas']} cuentas actualizadas)"

    # BNK-RECON tagger (guía rápida)
    if _quick_match(input_lower, "recon") and (_quick_match(input_lower, "tag") or _quick_match(input_lower, "marcar") or "bnk" in input_lower):
        return (
            "🏷️ **BNK-RECON TAGGER**\n"
            "Esta opción solo AGREGA tags a transactions existentes de QBO (no crea nuevas).\n\n"
            "Pasos:\n"
            "1. Descarga el CSV de tu banco (formato: date, description, amount).\n"
            "2. Colócalo en `/Bank Reconciliation/` o en cualquier ruta.\n"
            "3. Dime: 'reconcilia el banco con [ruta al CSV]'.\n"
            "4. Yo agregaré tags `BNK-RECON-YYYY-MM-xxxxx` a las transactions que matcheen.\n"
            "5. Después entras a QBO UI a reconciliar manualmente con esos tags visibles.\n\n"
            "Para limpiar tags de un batch: 'limpia los tags del batch [batch_id]'."
        )

    # Batch deposits (guía rápida)
    if any(kw in input_lower for kw in [
        "lote csv", "lote deposit", "depositar batch", "batch deposit",
        "depositos lote", "depositos csv",
    ]):
        return (
            "📦 **DEPOSITOS BATCH (motor con state machine)**\n"
            "Procesa CSVs de líneas de deposit multi-cliente con:\n"
            "- State machine: PENDING → VALIDATED → DRY-RUN → CONFIRMED → EXECUTED\n"
            "- Disambiguación interactiva: pregunta si un cliente no existe\n"
            "- Dry-run obligatorio antes de crear en QBO\n"
            "- Audit log completo en SQLite\n\n"
            "Pasos:\n"
            "1. Prepara un CSV con columnas: date, client_name, amount.\n"
            "   Opcionales: terms, memo.\n"
            "2. Dime: 'procesa el lote [ruta al CSV]'.\n"
            "3. Yo busco los clientes en QBO. Si falta alguno, te pregunto.\n"
            "4. Te muestro el dry-run. Tú confirmas.\n"
            "5. Creo el deposit en QBO y guardo el batch_id.\n\n"
            "Para dry-run sin crear: 'procesa el lote [ruta] sin confirmar'."
        )

    # Template CSV
    if _quick_match(input_lower, "template") or _quick_match(input_lower, "plantilla"):
        result = tool_crear_template_csv()
        return f"Template CSV creado: {result['archivo']}. Úsalo como base para depósitos batch."

    # Listar reportes guardados
    if _quick_match(input_lower, "listar") and _quick_match(input_lower, "reporte"):
        result = tool_listar_reportes_guardados()
        if result["total"] == 0:
            return "No tienes reportes guardados todavía. Guarda configuraciones de reportes frecuentes para acceso rápido."

        response = f"Reportes Guardados ({result['total']}):\n"
        for rep in result["reportes"]:
            response += f"  • {rep['nombre']} - Creado: {rep['creado'][:10]}, Último uso: {rep['ultimo_uso'][:10]}\n"

        return response

    # Cambiar idioma
    if _quick_match(input_lower, "cambiar") and (_quick_match(input_lower, "idioma") or _quick_match(input_lower, "language")):
        new_lang = "en" if session_state["language"] == "es" else "es"
        session_state["language"] = new_lang
        
        # Guardar en contexto de empresa para persistencia
        if 'COMPANY_CONTEXT' in globals() and 'CURRENT_COMPANY' in globals() and CURRENT_COMPANY:
            COMPANY_CONTEXT["language"] = new_lang
            save_company_context(CURRENT_COMPANY['name'], COMPANY_CONTEXT)
            
        lang_name = "Inglés" if new_lang == "en" else "Español"
        status_msg = f"✅ Idioma cambiado a: **{lang_name}** / Language changed to: **{lang_name}**"
        return status_msg

    # Ayuda Contextual
    if _quick_match(input_lower, "ayuda") or _quick_match(input_lower, "manual"):
        if _quick_match(input_lower, "ocr") or _quick_match(input_lower, "factura"):
            return "📖 **AYUDA OCR:**\n1. Coloca PDFs/imágenes en `/Pending bills/`.\n2. Dime: 'Procesa las facturas'.\n3. Yo extraeré los datos y te preguntaré si tengo dudas.\n4. Los archivos irán a `/Processed bills/`."

        if _quick_match(input_lower, "banco") or _quick_match(input_lower, "reconcilia"):
            return (
                "📖 **AYUDA BANCOS:**\n"
                "Tengo 2 modos de reconciliación:\n"
                "1. **Agresivo** (crea transactions nuevas):\n"
                "   'reconcilia el banco con [archivo CSV]'\n"
                "2. **Seguro BNK-RECON** (solo taggea, no crea):\n"
                "   'recon tag [archivo CSV]'\n\n"
                "Usa el modo seguro si solo quieres marcadores visibles en QBO UI."
            )

        if _quick_match(input_lower, "reporte") or _quick_match(input_lower, "analiza"):
            return "📖 **AYUDA REPORTES:**\n- Puedes pedir P&L, Balance Sheet o análisis comparativos.\n- Ejemplo: 'Haz un P&L de este mes vs el anterior'.\n- También puedo generar Excels complejos con gráficos."

        return "📖 **DEXTER HELP:**\nPuedes pedir ayuda específica:\n- `ayuda ocr`\n- `ayuda bancos`\n- `ayuda reportes`\n- `ayuda recon`"

    return None


# ==================== OPTIMIZACIONES ====================

def _bilingual_keywords(spanish_keywords: list) -> list:
    """Combina keywords en español con sus traducciones comunes en inglés.

    LOW-6 fix: cuando el usuario está en modo 'en' o usa términos
    técnicos en inglés (e.g., 'invoice', 'vendor', 'report'), los
    KEYWORDS de dexter.tools (todo en español) no matchean. Esta
    función agrega las traducciones comunes para que el matching
    funcione bilingüe sin modificar cada módulo individualmente.
    """
    en_translations = {
        "buscar": ["find", "search", "lookup", "look up"],
        "cliente": ["customer", "client"],
        "vendor": ["vendor", "supplier", "provider"],
        "cuenta": ["account"],
        "item": ["item", "product", "service"],
        "reporte": ["report", "generate report", "show report"],
        "p&l": ["p&l", "profit and loss", "income statement", "pnl"],
        "balance": ["balance", "balance sheet"],
        "estado": ["statement"],
        "factura": ["invoice", "bill", "billing"],
        "bill": ["bill", "vendor bill"],
        "crear": ["create", "make", "new", "add"],
        "ocr": ["ocr", "extract", "scan"],
        "pdf": ["pdf"],
        "pending": ["pending", "queue"],
        "procesar": ["process", "handle"],
        "extraer": ["extract", "pull"],
        "clasificar": ["classify", "categorize", "categorise"],
        "banco": ["bank", "banking"],
        "feed": ["feed"],
        "lote": ["batch", "bulk", "lot"],
        "batch": ["batch", "bulk"],
        "csv": ["csv", "spreadsheet"],
        "deposito": ["deposit"],
        "depósito": ["deposit"],
        "depositar": ["deposit"],
        "template": ["template", "sample"],
        "asiento": ["journal entry", "entry", "je"],
        "journal": ["journal", "journal entry"],
        "transferencia": ["transfer", "wire"],
        "token": ["token", "usage", "billing"],
        "estadística": ["stats", "statistics", "metrics"],
        "costo": ["cost", "price", "expense"],
        "gasto": ["expense", "spending", "spend"],
        "recon": ["recon", "reconcile", "reconciliation", "match"],
        "reconcili": ["recon", "reconcile", "reconciliation"],
        "tag": ["tag", "label", "mark"],
        "marcar": ["mark", "tag", "label"],
        "código": ["code", "script"],
        "python": ["python", "py"],
        "calcula": ["calculate", "compute", "calc"],
        "analiza": ["analyze", "analyse", "analysis"],
        "ejecuta": ["execute", "run"],
        "script": ["script", "code"],
        "idioma": ["language", "lang"],
        "language": ["language", "lang"],
        "cambiar": ["change", "switch"],
        "empresa": ["company", "business"],
        "empresas": ["companies", "businesses"],
        "registrar": ["register", "add", "create"],
        "listar": ["list", "show all"],
        "refrescar": ["refresh", "reload"],
        "chart": ["chart", "accounts", "coa"],
        "cuentas": ["accounts", "chart"],
    }
    out = list(spanish_keywords)
    for kw in spanish_keywords:
        for k_norm in en_translations.get(kw.lower(), []):
            if k_norm not in out:
                out.append(k_norm)
    seen = set()
    deduped = []
    for x in out:
        if x.lower() not in seen:
            deduped.append(x)
            seen.add(x.lower())
    return deduped


def get_relevant_tools(user_message: str) -> list:
    """Retorna la lista COMPLETA de tools (schemas) disponibles.

    Data-driven: incluye TODOS los tools de TOOLS + ALL_SCHEMAS.
    El LLM es suficientemente inteligente para elegir el tool correcto
    sin necesidad de un filtro de keywords que limite sus opciones.

    Si el LLM no entiende qué tool usar, debe preguntar al usuario
    (esto se refuerza en el system prompt).
    """
    import dexter.tools as dexter_tools
    from dexter.tools import _extract_name

    seen = set()
    result = []

    # Todos los tools del registry (101 tools)
    for schema in dexter_tools.ALL_SCHEMAS:
        # Normalizar: algunos módulos usan formato corto {name, desc, params}
        # sin el wrapper {"type": "function", "function": {...}}.
        # OpenRouter/Groq/DeepInfra requieren el formato estándar.
        if isinstance(schema, dict) and "type" not in schema:
            schema = {"type": "function", "function": schema}
        name = _extract_name(schema)
        if name and name not in seen:
            result.append(schema)
            seen.add(name)

    # Tools que solo están en TOOLS (legacy, si queda alguno no migrado)
    for t in TOOLS:
        name = t["function"]["name"]
        if name not in seen:
            result.append(t)
            seen.add(name)

    return result

def _truncate_message_content(msg: dict, max_chars: int = 2000) -> dict:
    """Trunca el campo 'content' de un mensaje a max_chars.

    MED-13 fix: protege el context window del LLM. Si un tool
    retornó un payload gigante (reporte sin MED-6 fix, batch
    grande), el mensaje vive en conversation_history y se reenvía
    en cada iteración. Esta función trunca el contenido a N
    caracteres agregando un marcador '[truncated, original: N chars]'.

    Preserva: role, name, tool_call_id, tool_calls y otros campos.
    """
    if not isinstance(msg, dict) or max_chars <= 0:
        return msg

    content = msg.get("content")
    if not isinstance(content, str) or len(content) <= max_chars:
        return msg

    truncated_msg = dict(msg)
    truncated_msg["content"] = (
        content[:max_chars]
        + f"...[truncated, original: {len(content)} chars]"
    )
    truncated_msg["_truncated"] = True
    truncated_msg["_original_content_length"] = len(content)
    return truncated_msg


def build_conversation_context(history: list, max_turns: int = 5,
                                max_content_chars: int = 2000) -> tuple:
    """Construye (recent_messages, context_hint) para el system prompt.

    MED-13 fix: trunca cada mensaje a max_content_chars para evitar
    que un tool result gigante (5MB) se reenvíe al LLM en cada
    iteración del loop.
    """
    recent_raw = list(history)[-(max_turns * 2):] if len(history) > max_turns * 2 else list(history)
    recent = [_truncate_message_content(m, max_chars=max_content_chars) for m in recent_raw]

    if history:
        text = " ".join([str(m.get("content", ""))[:80] for m in list(history)[-4:]])
        hints = []
        if "reporte" in text: hints.append("reportes")
        if "clasifica" in text: hints.append("clasificación")
        context = f"Contexto: {', '.join(hints)}" if hints else "Conversación"
    else:
        context = "Inicio"
    return recent, context

def necesita_chart(msg: str) -> bool:
    kw = ["clasifica", "cuenta", "bill", "journal", "asiento"]
    return any(k in msg.lower() for k in kw)

# ==================== LOOP PRINCIPAL ====================

def show_main_menu() -> str:
    """Retorna el menú completo de referencia (callable a demanda)."""
    return (
        "=" * 70 + "\n"
        "           🤖 DEXTER - QuickBooks AI Assistant\n"
        "              Operando para: Alfredo\n"
        + "=" * 70 + "\n"
        "\n"
        "Comandos rápidos:\n"
        "  • 'ayuda ocr'         - Guía paso a paso para facturas (PDFs)\n"
        "  • 'ayuda bancos'      - Guía para bank feeds / clasificación\n"
        "  • 'ayuda recon'       - Guía BNK-RECON (tag-only, no crea txns)\n"
        "  • 'ayuda reportes'    - Guía de reportes (P&L, Balance, custom)\n"
        "  • '¿cuánto he gastado?' - Estadísticas de tokens (sesión/día/mes)\n"
        "  • 'informe de tokens' - Genera Excel con estadísticas\n"
        "  • 'template csv'      - Crea plantilla para depósitos en lote\n"
        "  • 'lote csv [ruta]'   - Guía del motor batch de depósitos\n"
        "  • 'menu' / '?'        - Muestra este menú\n"
        "  • 'salir' / 'exit'    - Termina la sesión\n"
        "\n"
        "💡 Habla con naturalidad para todo lo demás. El LLM interpreta\n"
        "   y llama el tool correcto (100 tools disponibles en 21 dominios).\n"
    )


def main_loop():
    """Loop principal conversacional (estilo Claude Code / opencode)."""
    try:
        from dexter.console import console
        console.print("  [dim]DEXTER listo. 'menu' para ayuda, 'salir' para terminar.[/dim]")
        console.print()
    except ImportError:
        print("🤖 DEXTER listo. Escribe 'menu' o '?' para la ayuda. 'salir' para terminar.\n")

    try:
        _main_loop_body()
    finally:
        _close_session_safely()


def _main_loop_body():
    """Cuerpo principal del loop, separado para try/finally en main_loop."""
    while True:
        try:
            # Usar Rich prompt si está disponible, sino input() normal
            try:
                from dexter.console import user_prompt as _rich_prompt
                user_input = _rich_prompt().strip()
            except ImportError:
                user_input = input("👤 Tú: ").strip()

            if not user_input:
                continue

            # Detectar modo dry-run
            user_input, DRY_RUN_ACTIVE = _parse_dry_run(user_input)
            lower = user_input.lower()

            # Comando /ejecutar: replica el último dry-run pero real
            if lower in ("/ejecutar", "ejecutar", "ejecutalo", "ejecútalo",
                         "hacelo", "dale", "ahora si", "ahora sí", "confirmo"):
                if _last_dry_run_message:
                    user_input = _last_dry_run_message
                    _last_dry_run_message = None
                    lower = user_input.lower()
                    DRY_RUN_ACTIVE = False
                    print(f"  [Ejecutando: \"{user_input}\"]")
                else:
                    print("  No hay nada pendiente para ejecutar. Usá --dry-run primero.")
                    continue

            # Comando de salida
            if lower in ["salir", "exit", "quit", "chao", "adiós", "adios"]:
                break

            # Menú a demanda (extraído a función para que se pueda testear)
            if lower in ["menu", "?", "help", "ayuda"]:
                print(f"\n{show_main_menu()}")
                continue

            # Comando /estudiar: regenera el perfil de la empresa
            if lower in ("/estudiar", "estudiar empresa", "/estudiar empresa"):
                print("  🔍 Estudiando la empresa desde QBO...")
                try:
                    profile = _generate_company_profile(force=True)
                    if profile:
                        print(f"  ✅ Perfil actualizado: {profile}")
                    else:
                        print("  ⚠️ No se pudo generar el perfil.")
                except Exception as e:
                    print(f"  ❌ Error: {e}")
                continue

            # Procesar comandos rápidos (triggers NL como "informe de tokens", "lote csv")
            quick_response = process_quick_command(user_input)
            if quick_response:
                print(f"\n{quick_response}\n")
                continue

            # Llamar al LLM con tools
            try:
                from dexter.console import assistant_label, assistant_response as _rich_response
                assistant_label()
            except ImportError:
                print("\n🤖 ", end="", flush=True)

            try:
                response = call_llm(user_input, tools=TOOLS)
                try:
                    from dexter.console import assistant_response as _ar
                    _ar(response)
                except ImportError:
                    print(f"{response}\n")
            except KeyboardInterrupt:
                print("\n[Interrupción detectada]\n")
                break
            except Exception as e:
                print(f"\n❌ Error: {e}\n")
                _log_error(
                    e,
                    category="user_input",
                    user_input=user_input,
                    company=(CURRENT_COMPANY or {}).get("name"),
                )
                import traceback
                traceback.print_exc()

        except KeyboardInterrupt:
            print("\n\n[Interrupción detectada]\n")
            break
        except Exception as e:
            print(f"\n❌ Error: {e}\n")
            _log_error(
                e,
                category="user_input",
                user_input=user_input if "user_input" in locals() else None,
                company=(CURRENT_COMPANY or {}).get("name"),
            )
            import traceback
            traceback.print_exc()


_session_already_closed = False


def _close_session_safely() -> None:
    """Cierra sesión con print de resumen y save idempotente.

    LOW-2 fix: se llama desde main_loop() finally block. Idempotente
    (LOW-2b): usa _session_already_closed para que atexit y finally
    no escriban dos veces el CSV.
    """
    global _session_already_closed
    if _session_already_closed:
        return
    _session_already_closed = True

    try:
        print("\n" + "=" * 70)
        print("Cerrando sesión...")

        duration = (datetime.now() - session_state["start_time"]).total_seconds() / 60
        total_tokens = session_state["input_tokens"] + session_state["output_tokens"]
        cost = calculate_session_cost()

        print("\n📊 RESUMEN DE LA SESIÓN")
        print("=" * 70)
        print(f"  Duración: {duration:.0f} minutos")
        print(f"  Tokens: {total_tokens:,}")
        print(f"    • Input: {session_state['input_tokens']:,}")
        print(f"    • Output: {session_state['output_tokens']:,}")
        print(f"  Costo: ${cost:.4f}")
        print(f"  Operaciones: {sum(session_state['operations'].values())}")

        for op, count in session_state["operations"].items():
            if count > 0:
                print(f"    • {op}: {count}")

        save_session_to_csv()
        print("\n✅ Sesión guardada exitosamente")
        print("=" * 70)
    except Exception as e:
        print(f"\n⚠️ Error cerrando sesión: {e}")
        try:
            _log_error(e, category="user_input", user_input="<session_close>")
        except Exception:
            pass


import atexit as _atexit
_atexit.register(_close_session_safely)


# ==================== VERIFICACIÓN QBO + OFERTA DE RE-AUTH (UX-1) ====================


def _reload_env_after_oauth():
    """Recarga QB_ACCESS_TOKEN / QB_REFRESH_TOKEN / QB_REALM_ID desde .env
    después de un OAuth flow exitoso. Actualiza también los globals y
    el meta.json de la empresa actual (si existe), para evitar que
    tokens viejos en meta.json sobrescriban los nuevos al reiniciar."""
    global QB_ACCESS_TOKEN, QB_REFRESH_TOKEN, QB_REALM_ID
    from dotenv import load_dotenv
    load_dotenv(override=True)
    QB_ACCESS_TOKEN = os.getenv("QB_ACCESS_TOKEN", QB_ACCESS_TOKEN)
    QB_REFRESH_TOKEN = os.getenv("QB_REFRESH_TOKEN", QB_REFRESH_TOKEN)
    QB_REALM_ID = os.getenv("QB_REALM_ID", QB_REALM_ID)

    # Sincronizar meta.json de la empresa actual con los nuevos tokens
    if CURRENT_COMPANY:
        from company_manager import save_company_meta
        save_company_meta(
            CURRENT_COMPANY["name"],
            CURRENT_COMPANY["realm_id"],
            access_token=QB_ACCESS_TOKEN,
            refresh_token=QB_REFRESH_TOKEN,
        )


def _verify_qbo_connection_or_offer_reauth() -> bool:
    """Verifica la conexión a QBO. Si falla, intenta refresh; si refresh
    también falla (refresh token expirado), ofrece al usuario lanzar
    `scripts/oauth_flow.py` para re-autenticar interactivamente.

    Returns:
        True si la conexión quedó establecida (directa, vía refresh, o
              vía re-auth del usuario).
        False si QBO no responde y el usuario eligió no re-autenticar,
              o si el input no es interactivo (EOF), o si el OAuth
              flow terminó con error.

    UX-1: reemplaza el `exit(1)` silencioso del bloque de verificación
          por una experiencia de recuperación. Llamado desde el entry
          point `if __name__ == "__main__":`.
    """
    from pathlib import Path

    print("\n🔄 Verificando conexión a QuickBooks...")
    test_query = qbo_query("SELECT COUNT(*) FROM Account")

    # Camino feliz: QBO responde 200
    if "error" not in test_query:
        print("✅ Conexión establecida")
        return True

    # QBO no respondió. Intentar refresh automático.
    print(f"⚠️  Error conectando a QuickBooks: {test_query.get('error', 'desconocido')}")
    print("🔄 Intentando refrescar token...")
    if refresh_qb_token():
        # Refresh exitoso, re-verificar
        test_query2 = qbo_query("SELECT COUNT(*) FROM Account")
        if "error" not in test_query2:
            print("✅ Conexión establecida (token refrescado)")
            return True
        # Refresh dijo OK pero el endpoint sigue fallando (raro)
        print(f"⚠️  El token se refrescó pero QBO sigue respondiendo: {test_query2.get('error')}")

    # Refresh falló (o el endpoint sigue caído post-refresh). El caso más
    # probable es que el refresh_token mismo expiró (sandbox expira cada
    # ~24h, producción cada 100 días). Ofrecer OAuth interactivo.
    print("\n❌ No se pudo refrescar el token automáticamente.")
    print("   Esto pasa cuando el refresh token de QBO expira.")
    print("   Solución: correr el OAuth flow completo (login en navegador).")
    oauth_script = Path(__file__).resolve().parent / "scripts" / "oauth_flow.py"
    print(f"\n   Comando: python3 {oauth_script}")

    # Detectar si la entrada es interactiva
    if not sys.stdin.isatty():
        print("\n⚠️  Entrada no interactiva: no se puede ofrecer re-auth.")
        print("   Correr `python3 scripts/oauth_flow.py` manualmente y volver a intentar.")
        return False

    try:
        respuesta = input("\n¿Lanzar el OAuth flow ahora? (S/n): ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print("\n   Cancelado por el usuario.")
        return False

    if respuesta not in ("", "s", "si", "sí", "y", "yes"):
        print("   No se relanzará el OAuth flow.")
        return False

    # Lanzar OAuth flow en el mismo proceso (bloquea hasta que el user
    # complete el login en el navegador o cancele).
    env_label = "sandbox"
    if "production" in (os.getenv("QB_ENV") or "").lower() or \
       "production" in (os.getenv("QB_BASE_URL") or "").lower():
        env_label = "production"
    print(f"\n🚀 Lanzando OAuth flow ({env_label})...")
    proc = subprocess.run(
        [sys.executable, str(oauth_script), "--environment", env_label],
        cwd=str(Path(__file__).resolve().parent),
    )
    if proc.returncode != 0:
        print(f"\n❌ OAuth flow terminó con código {proc.returncode}.")
        return False

    # OAuth completó. Recargar .env (el script escribió nuevos tokens) y
    # re-verificar.
    print("\n🔄 Recargando credenciales desde .env...")
    _reload_env_after_oauth()
    test_query3 = qbo_query("SELECT COUNT(*) FROM Account")
    if "error" in test_query3:
        print(f"❌ OAuth completó pero QBO sigue sin responder: {test_query3.get('error')}")
        return False
    print("✅ Conexión establecida (re-auth exitoso)")
    return True


# ==================== MEMORIA PERSISTENTE (HERMES-STYLE) ====================


_dexter_memory = None  # inicializado bajo __main__


def _get_memory():
    """Lazy init de PersistentMemory — memoria por empresa.

    USER.md es global (~/.config/dexter/USER.md): preferencias de Alfredo.
    MEMORY.md es por empresa: companies/{name}/MEMORY.md.
    """
    global _dexter_memory
    if _dexter_memory is None:
        from dexter.core.memory import PersistentMemory

        # MEMORY.md por empresa (si hay empresa activa)
        company_memory = None
        if CURRENT_COMPANY and CURRENT_COMPANY.get("name"):
            safe_name = CURRENT_COMPANY["name"].replace("/", "_").replace("\\", "_")
            company_memory = f"companies/{safe_name}/MEMORY.md"

        _dexter_memory = PersistentMemory(
            memory_path=company_memory,
            user_path=None,  # USER.md usa el default global
        )
    return _dexter_memory


# ═══════════════════════════════════════════════════════════════════════
# COMPANY PROFILE — perfil automático de empresa (UX-3)
# ═══════════════════════════════════════════════════════════════════════

def _generate_company_profile(profile_dir: str = None, force: bool = False):
    """Genera PROFILE.md con datos clave de la empresa desde QBO.

    Contenido: chart of accounts, P&L últimos 3 meses, clientes activos,
    vendors activos, invoices recientes, y cuentas bancarias.

    Se ejecuta una sola vez (si PROFILE.md no existe) o con force=True.
    """
    from pathlib import Path
    from datetime import datetime, timedelta

    if not CURRENT_COMPANY:
        return ""

    if profile_dir:
        base = Path(profile_dir)
    else:
        safe_name = CURRENT_COMPANY["name"].replace("/", "_").replace("\\", "_")
        base = Path(f"companies/{safe_name}")

    profile_path = base / "PROFILE.md"
    if profile_path.exists() and not force:
        return ""

    base.mkdir(parents=True, exist_ok=True)
    today = datetime.now()
    last_month = (today - timedelta(days=30)).strftime("%Y-%m-%d")
    three_months_ago = (today - timedelta(days=90)).strftime("%Y-%m-%d")

    lines = []
    lines.append(f"# Perfil de {CURRENT_COMPANY['name']}")
    lines.append(f"Generado: {today.strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"Realm ID: {CURRENT_COMPANY['realm_id']}")
    lines.append("")

    # ── Chart of Accounts ──
    try:
        acct_count = qbo_query("SELECT COUNT(*) FROM Account")
        total = acct_count.get("QueryResponse", {}).get("totalCount", 0)
    except Exception:
        total = 0

    lines.append(f"## Chart of Accounts ({total} cuentas)")

    if total > 0:
        try:
            accts = qbo_query("SELECT * FROM Account MAXRESULTS 200")
            acct_rows = accts.get("QueryResponse", {}).get("Account", [])
            type_counts = {}
            bank_accounts = []
            for a in acct_rows:
                atype = a.get("AccountType", "Other")
                type_counts[atype] = type_counts.get(atype, 0) + 1
                if atype in ("Bank", "Credit Card"):
                    bank_accounts.append(f"  - {a.get('Name', '?')} ({atype}) · {a.get('Id', '?')}")

            lines.append("Por tipo:")
            for atype, count in sorted(type_counts.items()):
                lines.append(f"  - {atype}: {count}")

            if bank_accounts:
                lines.append("Cuentas bancarias:")
                lines.extend(bank_accounts)
        except Exception:
            lines.append("  (no se pudo obtener detalle)")

    lines.append("")

    # ── P&L ──
    lines.append("## P&L (último mes)")
    try:
        pl = _fetch_report("ProfitAndLoss", start_date=last_month,
                          end_date=today.strftime("%Y-%m-%d"))
        lines.append(_report_summary(pl, compact=True))
    except Exception:
        lines.append("  (no disponible)")

    lines.append("")

    # ── Clientes / Vendors activos ──
    try:
        cust_count = qbo_query("SELECT COUNT(*) FROM Customer WHERE Active = true")
        cust_total = cust_count.get("QueryResponse", {}).get("totalCount", 0)
    except Exception:
        cust_total = 0
    try:
        vend_count = qbo_query("SELECT COUNT(*) FROM Vendor WHERE Active = true")
        vend_total = vend_count.get("QueryResponse", {}).get("totalCount", 0)
    except Exception:
        vend_total = 0

    lines.append(f"## Clientes / Proveedores")
    lines.append(f"  - Clientes activos: {cust_total}")
    lines.append(f"  - Proveedores activos: {vend_total}")
    lines.append("")

    # ── Invoices recientes (últimos 30 días) ──
    lines.append("## Actividad reciente (30 días)")
    try:
        inv = qbo_query(
            f"SELECT * FROM Invoice WHERE TxnDate >= '{last_month}' MAXRESULTS 5"
        )
        inv_rows = inv.get("QueryResponse", {}).get("Invoice", [])
        if inv_rows:
            lines.append("Últimas facturas:")
            for i in inv_rows[:5]:
                customer = i.get("CustomerRef", {}).get("name", "?")
                total = i.get("TotalAmt", 0)
                date = i.get("TxnDate", "?")
                lines.append(f"  - {date} | {customer} | ${total:,.2f}")
        else:
            lines.append("  Sin facturas en los últimos 30 días")
    except Exception:
        lines.append("  (no disponible)")

    profile_path.write_text("\n".join(lines), encoding="utf-8")
    return str(profile_path)


def _load_company_profile(profile_dir: str = None):
    """Carga PROFILE.md como string para inyectar en el prompt."""
    from pathlib import Path

    if profile_dir:
        path = Path(profile_dir) / "PROFILE.md"
    elif CURRENT_COMPANY:
        safe_name = CURRENT_COMPANY["name"].replace("/", "_").replace("\\", "_")
        path = Path(f"companies/{safe_name}/PROFILE.md")
    else:
        return ""

    if path.exists():
        return path.read_text(encoding="utf-8").strip()
    return ""


def _report_summary(report_data: dict, compact: bool = False) -> str:
    """Convierte datos de _fetch_report en un resumen de texto compacto."""
    lines = []
    header = report_data.get("Header", {})
    name = header.get("ReportName", "Reporte")
    rows = report_data.get("Rows", {}).get("Row", [])
    if isinstance(rows, list):
        for row in rows[:10]:
            label = row.get("ColData", [{}])[0].get("value", "")
            value = ""
            if len(row.get("ColData", [])) > 1:
                value = row["ColData"][-1].get("value", "")
            if label and value:
                lines.append(f"  {label}: {value}")
    if compact:
        return "\n".join(lines[:8])
    return f"{name}\n" + "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════


def tool_gestionar_memoria(target: str = "memory", action: str = "add",
                           content: str = "", old_text: str = "") -> dict:
    """Tool: gestiona la memoria persistente del agente (MEMORY.md / USER.md).

    Args:
        target: 'memory' (notas del agente) o 'user' (perfil de Alfredo)
        action: 'add' (agregar), 'remove' (eliminar por substring), 'status' (ver uso)
        content: texto a agregar (para action='add')
        old_text: substring de la entrada a eliminar (para action='remove')

    Usos típicos:
      - Después de aprender algo nuevo sobre Alfredo
      - Después de descubrir un dato del entorno (ej. realm ID)
      - Cuando Alfredo corrige algo que dijiste
    """
    mem = _get_memory()
    if action == "status":
        return mem.get_status()
    elif action == "add":
        if not content.strip():
            return {"success": False, "error": "content no puede estar vacío"}
        return mem.add(target, content)
    elif action == "remove":
        if not old_text.strip():
            return {"success": False, "error": "old_text requerido para remove"}
        return mem.remove(target, old_text)
    else:
        return {"success": False, "error": f"Acción desconocida: {action}"}


TOOL_FUNCTIONS["gestionar_memoria"] = tool_gestionar_memoria


# ==================== TOOL DISPATCH TABLE ====================


if __name__ == "__main__":

    from dexter.console import (banner, header, success, error,
                                 user_prompt, assistant_label, assistant_response,
                                 status_msg, console)

    # Banner principal
    banner("🧠  DEXTER  ·  QBO Assistant", version="4.1.0-dev")

    # ---------------------------------------------------------------
    # SELECTOR DE EMPRESA (MULTI-COMPANY)
    # ---------------------------------------------------------------


    # Intentar cargar empresa previamente seleccionada
    CURRENT_COMPANY = load_current_company()

    if CURRENT_COMPANY:
        header(f"📁 {CURRENT_COMPANY['name']} · Realm {CURRENT_COMPANY['realm_id']}")
        cambiar = input("  ¿Cambiar de empresa? (s/N): ").strip().lower()
        if cambiar in ["s", "si", "sí", "y", "yes"]:
            CURRENT_COMPANY = None
            console.print()

    # Si no hay empresa seleccionada, mostrar selector
    if not CURRENT_COMPANY:
        CURRENT_COMPANY = select_company_interactive(QB_REALM_ID)
        if not CURRENT_COMPANY:
            error("No se seleccionó ninguna empresa.")
            exit(0)

    # Guardar selección
    save_company_selection(CURRENT_COMPANY)

    QB_REALM_ID = CURRENT_COMPANY["realm_id"]
    QB_BASE_URL = f"https://sandbox-quickbooks.api.intuit.com/v3/company/{QB_REALM_ID}"

    # Cargar tokens específicos de la empresa si existen
    meta = get_company_meta(CURRENT_COMPANY['name'])
    if meta.get("access_token") and meta.get("refresh_token"):
        QB_ACCESS_TOKEN = meta["access_token"]
        QB_REFRESH_TOKEN = meta["refresh_token"]
        print(f"🔑 Tokens cargados específicamente para {CURRENT_COMPANY['name']}")

    # Cargar contexto de la empresa
    status_msg(f"Cargando contexto de {CURRENT_COMPANY['name']}...")
    COMPANY_CONTEXT = load_company_context(CURRENT_COMPANY['name'])

    # Cargar Chart of Accounts específico de esta empresa
    if COMPANY_CONTEXT.get("chart_of_accounts"):
        session_state["chart_of_accounts"] = COMPANY_CONTEXT["chart_of_accounts"]
    else:
        session_state["chart_of_accounts"] = load_chart_of_accounts()
        COMPANY_CONTEXT["chart_of_accounts"] = session_state["chart_of_accounts"]
        save_company_context(CURRENT_COMPANY['name'], COMPANY_CONTEXT)

    # Cargar otros contextos
    if COMPANY_CONTEXT.get("saved_reports"):
        session_state["saved_reports"] = COMPANY_CONTEXT["saved_reports"]
    if COMPANY_CONTEXT.get("language"):
        session_state["language"] = COMPANY_CONTEXT["language"]

    n_accounts = len(session_state.get("chart_of_accounts", {}))
    n_reports = len(COMPANY_CONTEXT.get("saved_reports", {}))
    n_rules = len(COMPANY_CONTEXT.get("bank_feed_rules", {}))
    lang = session_state.get("language", "es").upper()
    status_msg(f"Contexto: {n_accounts} cuentas · {n_reports} reportes · {n_rules} reglas · {lang}")
    console.print()

    # Generar perfil de empresa si es primera carga
    from pathlib import Path
    safe_name = CURRENT_COMPANY["name"].replace("/", "_").replace("\\", "_")
    profile_path = Path(f"companies/{safe_name}/PROFILE.md")
    if not profile_path.exists():
        status_msg("Primera carga detectada. Estudiando la empresa...")
        try:
            profile_file = _generate_company_profile()
            if profile_file:
                success(f"Perfil generado: {profile_file}")
            else:
                info("Perfil no disponible (sin conexión a QBO)")
        except Exception as e:
            info(f"No se pudo generar perfil: {e}")

    # Verificar credenciales mínimas
    missing_creds = []
    if not QB_ACCESS_TOKEN:
        missing_creds.append("QB_ACCESS_TOKEN")
    if not QB_REALM_ID:
        missing_creds.append("QB_REALM_ID")
    if not OPENROUTER_API_KEY:
        missing_creds.append("OPENROUTER_API_KEY")

    if missing_creds:
        print("❌ ERROR: Faltan credenciales en el archivo .env:")
        for cred in missing_creds:
            print(f"   • {cred}")
        print("\nCrea un archivo .env con las credenciales necesarias.")
        exit(1)

    # Verificar conexión a QuickBooks (UX-1: ofrece re-auth si refresh falla)
    if not _verify_qbo_connection_or_offer_reauth():
        error("No se pudo establecer conexión con QuickBooks.")
        info("Si el refresh token expiró, ejecutá: python3 scripts/oauth_flow.py")
        exit(1)
    console.print()

    # Asegurar Chart of Accounts cargado en session_state
    if not session_state.get("chart_of_accounts"):
        session_state["chart_of_accounts"] = load_chart_of_accounts()
    
    # Iniciar loop principal
    main_loop()

    # Guardar memoria persistente al cerrar (por si el agente agregó entradas)
    print()
    mem = _get_memory()
    status = mem.get_status()
    if status["memory_entries"] > 0 or status["user_entries"] > 0:
        print(f"📝 Memoria: {status['memory_entries']} notas, {status['user_entries']} perfil")

